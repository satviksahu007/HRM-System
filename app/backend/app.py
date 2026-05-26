from flask import Flask ,jsonify ,request ,session,send_from_directory ,send_file
from flask_cors import CORS         #comms between differnet hosts
import os ,random
from werkzeug.utils import secure_filename
from db import get_db
import json
import hashlib
from datetime import date
import secrets
from datetime import timedelta
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
import cv2
import numpy as np
from datetime import datetime, date, time, timedelta
import openpyxl
from io import BytesIO
import calendar as cal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import timezone, timedelta







#APP CONFIG FILES////////////////////////////////////////////////////////////////////////////////////////////////

app = Flask(__name__)
CORS(app ,supports_credentials = True,origins = ["http://localhost:3000"])
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_PERMANENT'] = True

app.secret_key ="torToies123"
app.config.update(
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=False   # True only if using HTTPS
)

app.config['CALENDAR_UPLOAD_FOLDER'] = 'uploads/calendar'
app.config['UPLOAD_FOLDER'] = 'uploads/documents'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')


basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, "instance", "hr.db")
load_dotenv()
EMAIL_SENDER = "2328118@kiit.ac.in"
EMAIL_PASSWORD = "oikq ofng vgfn cili"







active_users = {}



#SESSION MANAGEMENT LOGIN LOGOUT

@app.route("/")
def home():
    return "Backend Successful"


def hash_password(password):
    """Hash password using SHA-256 (or use bcrypt for better security)"""
    return hashlib.sha256(password.encode()).hexdigest()

@app.route("/login", methods=["POST"])
def login():
    try:
        data = request.get_json()
        empid = data.get('empid', '').strip().upper()
        mobile = data.get('mobile', '').strip()
        password = data.get('password', '')

        if not password:
            return jsonify({"message": "Enter Password"}), 400

        if not empid and not mobile:
            return jsonify({"message": "Enter Employee Code or Mobile Number"}), 400

        conn = get_db()
        cursor = conn.cursor()

        # Single query: lock check, user data, and employee name
        cursor.execute("""
            SELECT 
                u.eha_id,
                u.is_locked,
                u.failed_attempts,
                u.password,

                LOWER(p.profile_name) AS role,

                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM team_emp_combination_all tec
                        WHERE tec.eha_id = u.eha_id
                        AND tec.link_type = 2
                        AND tec.unlinked_on IS NULL
                    )
                    THEN 1
                    ELSE 0
                END AS is_team_lead,

                u.status,
                u.is_first_login,
                e.first_name,
                e.last_name

            FROM user_login u

            JOIN employee_header_all e 
                ON u.eha_id = e.eha_id

            JOIN profile_header_all p
                ON e.type = p.profile_id

            WHERE (u.eha_id = %s OR u.mobile_number = %s)
            AND u.status = 1
            AND e.status = 1
            AND p.status = 1

            LIMIT 1
        """, (empid or None, mobile or None))

        user = cursor.fetchone()

        if not user:
            conn.close()
            return jsonify({"message": "Invalid credentials"}), 401

        if user['is_locked'] == 1:
            conn.close()
            return jsonify({"message": "Account is locked. Contact HR."}), 403

        # Verify password
        hashed_pass = hashlib.sha256(password.encode()).hexdigest()

        if user['password'] != hashed_pass:
            new_attempts = (user['failed_attempts'] or 0) + 1
            is_locked = 1 if new_attempts >= 5 else 0
            cursor.execute("""
                UPDATE user_login
                SET failed_attempts = %s, is_locked = %s, updated_on = NOW()
                WHERE eha_id = %s
            """, (new_attempts, is_locked, user['eha_id']))
            conn.commit()
            conn.close()

            attempts_left = 5 - new_attempts
            msg = f"Invalid credentials. {attempts_left} attempts left." if attempts_left > 0 else "Account locked."
            return jsonify({"message": msg}), 401

        # --- Successful login ---
        new_token = secrets.token_hex(32)
        login_timestamp = datetime.now()

        # Update login token and reset failed attempts (use resolved eha_id)
        cursor.execute("""
            UPDATE user_login
            SET login_token = %s, last_login = NOW(),
                failed_attempts = 0, is_locked = 0, updated_on = NOW()
            WHERE eha_id = %s
        """, (new_token, user['eha_id']))
        conn.commit()
        conn.close()

        # Build full name and set session
        full_name = f"{user['first_name']} {user['last_name']}"
        session['empid'] = user['eha_id']
        session['name'] = full_name
        session['role'] = user['role']
        session['token'] = new_token
        session['login_time'] = datetime.now().isoformat()
        session['session_expiry'] = (login_timestamp + timedelta(hours=24)).isoformat()
        active_users[user['eha_id']] = datetime.now()
        session['is_team_lead'] = user['is_team_lead']
        print(session)

        if user['is_first_login'] == 1:
            session['is_first_login'] = 1
            return jsonify({
                "message": "First login. Please change your password.",
                "requires_password_change": True,
                "is_first_login": True,
                "empid": user['eha_id'],
                "name": full_name,
                "is_team_lead": user['is_team_lead']
            }), 200

        session['is_first_login'] = 0
        return jsonify({
            "success": True,
            "name": full_name,
            "role": user['role'],
            "is_team_lead": user['is_team_lead'],
            "is_first_login": False
        }), 200

    except Exception as e:
        print(f"Login error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": "Login failed. Please try again."}), 500


@app.route("/logout", methods=["POST"])
def logout():
    if 'empid' in session:
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE user_login 
                SET login_token = NULL, 
                    reset_token = NULL, 
                    reset_token_expiry = NULL,
                    updated_on = NOW()
                WHERE eha_id = %s
            """, (session['empid'],))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Logout error: {e}")
    
    session.clear()
    return jsonify({"message": "Logged out"}), 200

active_users = {}  # { eha_id: last_activity_datetime }

@app.route("/check-session", methods=["GET"])
def check_session():
    session_expiry_str = session.get('session_expiry')
    if 'empid' not in session or 'token' not in session:
        return jsonify({"valid": False, "code": "NO_SESSION"}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT login_token, status, is_locked FROM user_login WHERE eha_id = %s", (session['empid'],))
    user = cursor.fetchone()
    conn.close()
    
    if not user or user['status'] != 1 or user['is_locked'] == 1:
        active_users.pop(session['empid'], None)
        session.clear()
        return jsonify({"valid": False, "code": "ACCOUNT_INACTIVE"}), 401
    
    if user['login_token'] != session['token']:
        active_users.pop(session['empid'], None)
        session.clear()
        return jsonify({"valid": False, "code": "TOKEN_MISMATCH"}), 401
    
    # Check absolute expiry (24 hours)
    session_expiry_str = session.get('session_expiry')      
    if session_expiry_str:
        session_expiry = datetime.fromisoformat(session_expiry_str)
        if datetime.now() > session_expiry:
            session.clear()
            return jsonify({"valid": False, "code": "SESSION_EXPIRED_24H"}), 401
    
    # Check idle timeout
    now = datetime.now()
    empid = session['empid']
    is_ping = request.args.get('ping') == 'true' 

    if empid in active_users:
        idle_time = now - active_users[empid]
        print(idle_time)
        if idle_time > timedelta(minutes=10):
            active_users.pop(empid, None)
            session.clear()
            return jsonify({"valid": False, "code": "IDLE_TIMEOUT"}), 401

    if not is_ping:                     
        active_users[empid] = now

    return jsonify({"valid": True}), 200


def send_otp_email(to_email,emp_name,otp,expiry):
    try:
        msg = MIMEMultipart()
        msg['From'] = f"Hive Mind<{EMAIL_SENDER}>"
        msg['To'] = to_email
        msg['Subject'] = "Password Change"

        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px;">
            <p>Hello <strong>{emp_name}</strong>,</p>
            <p>Your OTP for password change is:</p>
            <div style="background: #F3F4F6; padding: 15px; border-radius: 8px; text-align: center; margin: 15px 0;">
                <span style="font-size: 20px; font-weight: bold; letter-spacing: 8px; color: #4F46E5;">{otp}</span>
            </div>
            <p>This OTP is valid for <strong>10 minutes</strong>.</p>
            <p><strong>If this wasn't you, Please contact HR <strong></p>
            <p>Micro Integrated Semiconductor Systems Pvt.Ltd</p>
        </body>
        </html>
        """

        msg.attach(MIMEText(body,"html"))

        server = smtplib.SMTP('smtp.gmail.com',587)
        server.starttls()
        server.login(EMAIL_SENDER,EMAIL_PASSWORD.replace(' ',''))
        server.send_message(msg)
        server.quit()

        print(f" OTP email sent to {to_email}")
        return True
    except Exception as e:
        print(f" Email error: {e}")
        return False
    
def send_def_pass(to_email,emp_name,empcode,def_pass):
    try:
        msg = MIMEMultipart()
        msg['From'] = f"HiveMind<{EMAIL_SENDER}>"
        msg['To'] = to_email
        msg['Subject'] = "Welcome To The HiveMind"

        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; font-size: 14px; line-height: 1.5; color: #333; padding: 20px; margin: 0;">
            <p>Hello <strong>{emp_name}</strong>,</p>
            <p>Welcome to <strong>Micro Integrated Semiconductor Systems Pvt Ltd</strong>.</p>
            <p>We're absolutely delighted to have you join us.</p>
            <p>Your Employee Code and default password are as follows:</p>

            <div style="background: #F3F4F6; padding: 12px 15px; border-radius: 8px; text-align: center; margin: 16px 0;">
                <span style="font-size: 20px; font-weight: bold; letter-spacing: 4px; color: #4F46E5;">{empcode}</span>
            </div>
            <div style="background: #F3F4F6; padding: 12px 15px; border-radius: 8px; text-align: center; margin: 16px 0;">
                <span style="font-size: 20px; font-weight: bold; letter-spacing: 4px; color: #4F46E5;">{def_pass}</span>
            </div>

            <p><strong>Please log in and change your password at first login.</strong></p>
            <p style="margin-top: 24px; font-size: 12px; color: #666;">Micro Integrated Semiconductor Systems Pvt. Ltd.</p>
            </body>
        </html>
                """

        msg.attach(MIMEText(body,"html"))

        server = smtplib.SMTP('smtp.gmail.com',587)
        server.starttls()
        server.login(EMAIL_SENDER,EMAIL_PASSWORD.replace(' ',''))
        server.send_message(msg)
        server.quit()

        print(f"Default pass sent to {to_email}")
        return True
    except Exception as e:
        print(f" Email error: {e}")
        return False
    
@app.route("/forgot-password/request-otp", methods=["POST"])
def forgot_password_request_otp():
    try:
        print("byebye")
        data = request.get_json()
        empid = data.get('empid', '').strip().upper() or None
        mobile = data.get('mobile', '').strip() or None

        if not empid and not mobile:
            return jsonify({"message": "Please provide employee code or mobile number"}), 400

        conn = get_db()
        cursor = conn.cursor()

        # Find the user by empid or mobile (must be active)
        cursor.execute("""
            SELECT u.eha_id, u.reset_token, e.email, e.first_name
            FROM user_login u
            JOIN employee_header_all e ON u.eha_id = e.eha_id
            WHERE (u.eha_id = %s OR u.mobile_number = %s)
              AND u.status = 1 AND e.status = 1
            LIMIT 1
        """, (empid, mobile))

        user = cursor.fetchone()

        if not user:
            conn.close()
            return jsonify({"message": "No active account found with that identifier"}), 400

        if not user['email']:
            conn.close()
            return jsonify({"message": "No email address on file. Please contact HR."}), 400

        # Generate OTP and expiry
        otp = str(random.randint(100000, 999999))
        expiry = datetime.now() + timedelta(minutes =10)

        # Store OTP / expiry in the reset_token fields
        cursor.execute("""
            UPDATE user_login
            SET reset_token = %s, reset_token_expiry = %s, updated_on = NOW()
            WHERE eha_id = %s
        """, (otp, expiry, user['eha_id']))
        conn.commit()
        conn.close()

        # Send the OTP email (your existing send_otp_email function)
        email_sent = send_otp_email(user['email'], user['first_name'], otp, expiry)

        # For security, always return a generic message
        return jsonify({"message": "If the account exists, an OTP has been sent to the registered email."}), 200

    except Exception as e:
        print(f"Forgot password OTP error: {e}")
        return jsonify({"message": "Unable to process request"}), 500
    
@app.route("/forgot-password/reset", methods=["POST"])
def forgot_password_reset():
    try:
        data = request.get_json()
        empid = data.get('empid', '').strip().upper() or None
        mobile = data.get('mobile', '').strip() or None
        otp = data.get('otp', '').strip()
        new_password = data.get('new_password', '')

        if not empid and not mobile:
            return jsonify({"message": "Identifier required"}), 400
        if not otp or not new_password:
            return jsonify({"message": "OTP and new password are required"}), 400
        if len(otp) != 6 or not otp.isdigit():
            return jsonify({"message": "Invalid OTP format"}), 400
        if len(new_password) < 8:
            return jsonify({"message": "Password must be at least 8 characters"}), 400

        conn = get_db()
        cursor = conn.cursor()

        # Find user again
        cursor.execute("""
            SELECT eha_id, reset_token, reset_token_expiry, failed_attempts, is_locked, password
            FROM user_login
            WHERE (eha_id = %s OR mobile_number = %s) AND status = 1
            LIMIT 1
        """, (empid, mobile))
        user = cursor.fetchone()

        if not user:
            conn.close()
            return jsonify({"message": "Invalid request"}), 400

        if user['is_locked'] == 1:
            conn.close()
            return jsonify({"message": "Account is locked. Contact HR."}), 403

        # Check OTP
        if user['reset_token'] != otp:
            # Increment failed attempts to prevent brute force
            new_attempts = (user['failed_attempts'] or 0) + 1
            is_locked = 1 if new_attempts >= 5 else 0
            cursor.execute("""
                UPDATE user_login
                SET failed_attempts = %s, is_locked = %s, updated_on = NOW()
                WHERE eha_id = %s
            """, (new_attempts, is_locked, user['eha_id']))
            if is_locked:
                cursor.execute("UPDATE employee_header_all SET status = 0, valid_till = CURDATE() WHERE eha_id = %s", (user['eha_id'],))
            conn.commit()
            conn.close()
            return jsonify({"message": "Invalid OTP"}), 400

        if user['reset_token_expiry'] and datetime.now() > user['reset_token_expiry']:
            conn.close()
            return jsonify({"message": "OTP expired. Please request a new one."}), 400

        # OTP valid – update password and clear OTP fields
        hashed_new = hashlib.sha256(new_password.encode()).hexdigest()

        # Prevent using the same password
        if user['password'] == hashed_new:
            conn.close()
            return jsonify({"message": "New password cannot be same as current password"}), 400

        cursor.execute("""
            UPDATE user_login
            SET password = %s,
                reset_token = NULL,
                reset_token_expiry = NULL,
                is_first_login = 0,
                failed_attempts = 0,
                is_locked = 0,
                password_changed_on = NOW(),
                updated_on = NOW()
            WHERE eha_id = %s
        """, (hashed_new, user['eha_id']))
        conn.commit()
        conn.close()

        return jsonify({"message": "Password reset successfully. You can now log in.", "success": True}), 200

    except Exception as e:
        print(f"Password reset error: {e}")
        return jsonify({"message": "Something went wrong"}), 500

@app.route("/send-otp", methods=["POST"])
def send_otp():
    try:
        if 'empid' not in session:
            return jsonify({"message": "Not logged in"}), 401
        
        empid = session['empid']
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get user's email
        cursor.execute("""
            SELECT e.email, e.first_name, e.last_name
            FROM user_login u
            JOIN employee_header_all e ON u.eha_id = e.eha_id
            WHERE u.eha_id = %s AND u.status = 1 AND e.status = 1
        """, (empid,))
        
        user = cursor.fetchone()
        
        if not user or not user['email']:
            conn.close()
            return jsonify({"message": "No email found for your account"}), 400
        
        # Generate 6-digit OTP
        otp = str(random.randint(100000, 999999))
        expiry = datetime.now() + timedelta(minutes=10)
        
        # Store OTP in reset_token column
        cursor.execute("""
            UPDATE user_login 
            SET reset_token = %s, reset_token_expiry = %s, updated_on = NOW()
            WHERE eha_id = %s
        """, (otp, expiry, empid))
        
        conn.commit()
        conn.close()
        
        email_sent = send_otp_email(user['email'], user['first_name'], otp, expiry)
        
        return jsonify({"message": "OTP sent to your registered email"}), 200
        
    except Exception as e:
        print(f"Error sending OTP: {e}")
        return jsonify({"message": "Failed to send OTP"}), 500


@app.route("/verify-otp", methods=["POST"])
def verify_otp():
    try:
        if 'empid' not in session:
            return jsonify({"message": "Not logged in"}), 401
        
        data = request.get_json()
        otp = data.get('otp', '').strip()
        new_password = data.get('new_password', '')
        
        if not otp or not new_password:
            return jsonify({"message": "OTP and new password are required"}), 400
        
        if len(otp) != 6 or not otp.isdigit():
            return jsonify({"message": "Invalid OTP format"}), 401
        
        if len(new_password) < 8:
            return jsonify({"message": "Password must be at least 8 characters"}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        empid = session['empid']
        
        cursor.execute("""
            SELECT reset_token, reset_token_expiry, failed_attempts
            FROM user_login 
            WHERE eha_id = %s AND status = 1
        """, (empid,))
        
        user = cursor.fetchone()
        
        if not user:
            conn.close()
            return jsonify({"message": "Account not found"}), 401
        
        # Check if OTP matches
        if user['reset_token'] != otp:
            # Count failed OTP attempts
            new_attempts = user['failed_attempts'] + 1
            should_lock = new_attempts >= 5
            
            cursor.execute("""
                UPDATE user_login 
                SET failed_attempts = %s, is_locked = %s, updated_on = NOW()
                WHERE eha_id = %s
            """, (new_attempts, should_lock, empid))
            
            if should_lock:
                cursor.execute("""
                    UPDATE employee_header_all 
                    SET status = 0, valid_till = CURDATE() 
                    WHERE eha_id = %s
                """, (empid,))
            
            conn.commit()
            conn.close()
            
            return jsonify({"message": "Invalid OTP"}), 400
        
        # Check if OTP expired
        if user['reset_token_expiry'] and datetime.now() > user['reset_token_expiry']:
            conn.close()
            return jsonify({"message": "OTP has expired. Please request a new one."}), 400
        
        # OTP valid - update password
        hashed_password = hashlib.sha256(new_password.encode()).hexdigest()
        
        cursor.execute("""
            UPDATE user_login 
            SET password = %s, 
                reset_token = NULL, 
                reset_token_expiry = NULL,
                is_first_login = 0, 
                failed_attempts = 0, 
                is_locked = 0,
                password_changed_on = NOW(),
                updated_on = NOW()
            WHERE eha_id = %s
        """, (hashed_password, empid))
        
        conn.commit()
        conn.close()
        
        session['is_first_login'] = 0
        
        return jsonify({"message": "Password changed successfully!", "success": True}), 200
        
    except Exception as e:
        print(f"Error verifying OTP: {e}")
        return jsonify({"message": "Something went wrong"}), 500

@app.route("/change_password", methods=["POST"])
def change_password():
    if "empid" not in session:
        return jsonify({"message": "Not logged in"}), 401
    
    data = request.json
    new_password = data.get("new_password")
    otp = data.get('otp', '').strip()
    empid = session['empid']

    if not otp or not new_password:
        return jsonify({"message": "OTP and new password are required"}), 400
    
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get user (ignore password check first)
    cursor.execute("""
        SELECT eha_id, password, is_first_login, status, is_locked, failed_attempts,reset_token,reset_token_expiry
        FROM user_login 
        WHERE eha_id = %s
    """, (session["empid"],))
    
    user = cursor.fetchone()

    
    
    if not user or user['status'] != 1:
        conn.close()
        return jsonify({"message": "Account not found or inactive"}), 401
    
    hashed_new = hashlib.sha256(new_password.encode()).hexdigest()
    if user['password'] == hashed_new:
        conn.close()
        return jsonify({"message": "New password cannot be same as current password"}), 400
    
    if user['is_locked'] == 1:
        return jsonify({"message": "Account locked. Contact admin"}), 401
    
    if user['reset_token']!= otp:
        new_attempts = user['failed_attempts'] +1
        should_lock = new_attempts>=5
        cursor.execute("UPDATE user_login SET failed_attempts = %s, is_locked = %s WHERE eha_id = %s", (new_attempts, should_lock, empid))
        if should_lock:
            cursor.execute("UPDATE employee_header_all SET status = 0, valid_till = CURDATE() WHERE eha_id = %s", (empid,))
        conn.commit()
        conn.close()
        return jsonify({"message": "Invalid OTP"}), 400
    
    if user['reset_token_expiry'] and datetime.now() > user['reset_token_expiry']:
        conn.close()
        return jsonify({"message": "OTP expired. Request a new one."}), 401
    
    
    
    cursor.execute("""
        UPDATE user_login 
        SET password = %s, reset_token = NULL, reset_token_expiry = NULL,
            is_first_login = 0, failed_attempts = 0, is_locked = 0,
            password_changed_on = NOW(), updated_on = NOW()
        WHERE eha_id = %s
    """, (hashed_new, empid))
    conn.commit()
    conn.close()
    
        
    session['is_first_login'] = 0
        
    return jsonify({
        "message": "Password changed successfully",
        "is_first_login": False
    }), 200
    

def verify_password(plain, hashed):
    return hash_password(plain) == hashed

@app.route("/auto-terminate", methods=["POST"])
def auto_terminate():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE employee_header_all 
            SET status = 0 
            WHERE valid_till IS NOT NULL 
              AND valid_till <= CURDATE() 
              AND status = 1
        """)
        
        affected = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "terminated": affected}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500
    





#EMPLOYEE MANAGEMENT////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


    
@app.route("/addemp", methods=["POST"])
def add_emp():
    try:
        if 'empid' not in session:
            return jsonify({"message": "Not logged in!"}), 401
        
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        data = request.form
        files = request.files
        status_value = int(data.get('status', 1))
        trial_eha_id = data.get('trial_eha_id')
        
        # Save uploaded files function
        def save_file(file_field, field_name):
            if file_field in files and files[file_field]:
                file = files[file_field]
                if file and allowed_file(file.filename):
                    filename = f"{field_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    return filepath
            return None

        # Save all files
        aadhar_image = save_file('aadhar_image', 'aadhar')
        pan_image = save_file('pan_image', 'pan')
        tenth_marksheet = save_file('tenth_marksheet', 'tenth')
        twelfth_marksheet = save_file('twelfth_marksheet', 'twelfth')
        ug_degree_image = save_file('ug_degree_image', 'ug')
        pg_degree_image = save_file('pg_degree_image', 'pg')
        photo = save_file('photo', 'photo')

        conn = get_db()
        cursor = conn.cursor()

        # ------------------------------------------------------------
        # TRIAL CONFIRMATION (existing trial employee becomes active)
        # ------------------------------------------------------------
        print("All form keys:", list(data.keys()))
        print("trial_eha_id value:", repr(trial_eha_id))
        if trial_eha_id and trial_eha_id.strip():
            print("byebuyte")
            # Fetch existing trial employee + their file paths
            cursor.execute("""
                SELECT eha_id, aadhar_image_path, pan_image_path, photo_path,
                    tenth_marksheet_path, twelfth_marksheet_path,
                    ug_degree_image_path, pg_degree_image_path
                FROM employee_header_all WHERE eha_id = %s AND status = 2
            """, (trial_eha_id,))
            trial_emp = cursor.fetchone()

            if not trial_emp:
                return jsonify({"message": "Trial employee not found"}), 404

            # Prevent double confirmation
            cursor.execute("SELECT ulh_id FROM user_login WHERE eha_id = %s", (trial_eha_id,))
            if cursor.fetchone():
                return jsonify({"message": "Employee already confirmed"}), 400

            # Use existing files if no new ones uploaded
            final_aadhar_image      = aadhar_image      or trial_emp['aadhar_image_path']
            final_pan_image         = pan_image         or trial_emp['pan_image_path']
            final_photo             = photo             or trial_emp['photo_path']
            final_tenth_marksheet   = tenth_marksheet   or trial_emp['tenth_marksheet_path']
            final_twelfth_marksheet = twelfth_marksheet or trial_emp['twelfth_marksheet_path']
            final_ug_degree         = ug_degree_image   or trial_emp['ug_degree_image_path']
            final_pg_degree         = pg_degree_image   or trial_emp['pg_degree_image_path']

            # Update the existing row: set status=1, valid_till=NULL
            cursor.execute("""
                UPDATE employee_header_all SET
                    first_name=%s, middle_name=%s, last_name=%s,
                    marital_status=%s, gender=%s, gender_other=%s, dob=%s,
                    mob_no=%s, email=%s,
                    aadhar=%s, pan=%s,
                    aadhar_image_path=%s, pan_image_path=%s, photo_path=%s,
                    bank_name=%s, bank_account_number=%s, bank_ifsc=%s,
                    salary=%s, expertise=%s,
                    spouse_name=%s, spouse_mob=%s, spouse_email=%s,
                    tenth_school=%s, tenth_board=%s, tenth_year=%s, tenth_marks=%s, tenth_marksheet_path=%s,
                    twelfth_school=%s, twelfth_board=%s, twelfth_year=%s, twelfth_marks=%s, twelfth_marksheet_path=%s,
                    ug_college=%s, ug_degree=%s, ug_year=%s, ug_marks=%s, ug_degree_image_path=%s,
                    pg_college=%s, pg_degree=%s, pg_year=%s, pg_marks=%s, pg_degree_image_path=%s,
                    father_name=%s, father_mobile=%s, mother_name=%s, mother_mobile=%s,
                    joining_date=%s, status=1, valid_till=NULL
                WHERE eha_id=%s AND status=2
            """, (
                data.get('first_name', ''),
                data.get('middle_name') or None,
                data.get('last_name', ''),
                data.get('marital_status', 'single'),
                data.get('gender', ''),
                data.get('gender_other') or None,
                data.get('dob') or None,
                data.get('mob_no', ''),
                data.get('email', ''),
                data.get('aadhar', ''),
                data.get('pan', ''),
                final_aadhar_image, final_pan_image, final_photo,
                data.get('bank_name', ''),
                data.get('bank_account_number', ''),
                data.get('bank_ifsc', ''),
                data.get('salary', 0),
                data.get('expertise', ''),
                data.get('spouse_name') or None,
                data.get('spouse_mob') or None,
                data.get('spouse_email') or None,
                data.get('tenth_school') or None, data.get('tenth_board') or None,
                data.get('tenth_year') or None, data.get('tenth_marks') or None, final_tenth_marksheet,
                data.get('twelfth_school') or None, data.get('twelfth_board') or None,
                data.get('twelfth_year') or None, data.get('twelfth_marks') or None, final_twelfth_marksheet,
                data.get('ug_college') or None, data.get('ug_degree') or None,
                data.get('ug_year') or None, data.get('ug_marks') or None, final_ug_degree,
                data.get('pg_college') or None, data.get('pg_degree') or None,
                data.get('pg_year') or None, data.get('pg_marks') or None, final_pg_degree,
                data.get('father_name') or None, data.get('father_mobile') or None,
                data.get('mother_name') or None, data.get('mother_mobile') or None,
                data.get('joining_date') or None,
                trial_eha_id
            ))

            # Create login for the now active employee
            cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'user_login'")
            login_header = cursor.fetchone()
            if not login_header or login_header['last_id'] is None:
                cursor.execute("INSERT INTO unique_header_all (table_name, prefix, last_id) VALUES ('user_login', 'ULH', '0')")
                conn.commit()
                new_login_id = 1
                ulh_id = "ULH-00001"
            else:
                new_login_id = int(login_header['last_id']) + 1
                ulh_id = f"{login_header['prefix']}-{new_login_id:05d}"
                cursor.execute(
                    "UPDATE unique_header_all SET last_id=%s, modified_on=NOW() WHERE table_name='user_login'",
                    (str(new_login_id),)
                )

            default_pass = str(secrets.randbelow(90_000_000) + 10_000_000)
            hashed_pass = hashlib.sha256(default_pass.encode()).hexdigest()

            cursor.execute("""
                INSERT INTO user_login (ulh_id, eha_id, mobile_number, password, role, status, is_first_login)
                VALUES (%s, %s, %s, %s, 'employee', 1, 1)
            """, (ulh_id, trial_eha_id, data.get('mob_no'), hashed_pass))

            send_def_pass(data.get('email'), data.get('first_name'), trial_eha_id, default_pass)
            conn.commit()
            conn.close()

            return jsonify({
                "success": True,
                "message": "Trial employee confirmed as active",
                "eha_id": trial_eha_id,
                "password": default_pass,
                "name": f"{data.get('first_name')} {data.get('last_name')}",
                "email": data.get('email')
            }), 200

        # ------------------------------------------------------------
        # BRAND NEW EMPLOYEE
        # ------------------------------------------------------------
        # Check for duplicate email
        cursor.execute("SELECT eha_id FROM employee_header_all WHERE email = %s", (data.get('email'),))
        if cursor.fetchone():
            conn.close()
            return jsonify({"message": "Email already in use"}), 400

        # Generate new eha_id
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'employee_header_all'")
        header = cursor.fetchone()
        if not header:
            cursor.execute("""
                INSERT INTO unique_header_all (table_name, prefix, last_id) 
                VALUES ('employee_header_all', 'EHA', '0')
            """)
            conn.commit()
            new_last_id = 1
            eha_id = "EHA-00001"
        else:
            new_last_id = int(header['last_id']) + 1
            eha_id = f"{header['prefix']}-{new_last_id:05d}"
            cursor.execute("""
                UPDATE unique_header_all 
                SET last_id = %s, modified_on = NOW() 
                WHERE table_name = 'employee_header_all'
            """, (str(new_last_id),))

        # If not a trial, generate login credentials now
        if status_value != 2:
            cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'user_login'")
            login_header = cursor.fetchone()
            if not login_header or login_header['last_id'] is None:
                cursor.execute("""
                    INSERT INTO unique_header_all (table_name, prefix, last_id) 
                    VALUES ('user_login', 'ULH', '0')
                """)
                conn.commit()
                new_login_id = 1
                ulh_id = "ULH-00001"
            else:
                new_login_id = int(login_header['last_id']) + 1
                ulh_id = f"{login_header['prefix']}-{new_login_id:05d}"
                cursor.execute("""
                    UPDATE unique_header_all 
                    SET last_id = %s, modified_on = NOW() 
                    WHERE table_name = 'user_login'
                """, (str(new_login_id),))
            default_pass = str(secrets.randbelow(90_000_000) + 10_000_000)
            hashed_pass = hashlib.sha256(default_pass.encode()).hexdigest()

        # Insert employee record
        sql = """
            INSERT INTO employee_header_all (
                eha_id, first_name, middle_name, last_name, 
                marital_status, gender, gender_other, dob, 
                mob_no, email, type, status,
                present_address_line1, present_address_line2, present_city, present_state, present_pincode, present_country,
                permanent_address_line1, permanent_address_line2, permanent_city, permanent_state, permanent_pincode, permanent_country,
                aadhar, pan, aadhar_image_path, pan_image_path,
                bank_name, bank_account_number, bank_ifsc,
                salary, expertise, photo_path,
                spouse_name, spouse_mob, spouse_email,
                tenth_school, tenth_board, tenth_year, tenth_marks, tenth_marksheet_path,
                twelfth_school, twelfth_board, twelfth_year, twelfth_marks, twelfth_marksheet_path,
                ug_college, ug_degree, ug_year, ug_marks, ug_degree_image_path,
                pg_college, pg_degree, pg_year, pg_marks, pg_degree_image_path,
                father_name, father_mobile, mother_name, mother_mobile, joining_date
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s
            )
        """
        values = (
            eha_id,
            data.get('first_name', ''),
            data.get('middle_name') or None,
            data.get('last_name', ''),
            data.get('marital_status', 'single'),
            data.get('gender', ''),
            data.get('gender_other') or None,
            data.get('dob') or None,
            data.get('mob_no', ''),
            data.get('email', ''),
            'PHA-00003',   # type – adjust as needed
            1,             # temporary status – will be updated if trial
            data.get('present_address_line1', ''),
            data.get('present_address_line2', ''),
            data.get('present_city', ''),
            data.get('present_state', ''),
            data.get('present_pincode', ''),
            data.get('present_country', 'India'),
            data.get('permanent_address_line1', ''),
            data.get('permanent_address_line2', ''),
            data.get('permanent_city', ''),
            data.get('permanent_state', ''),
            data.get('permanent_pincode', ''),
            data.get('permanent_country', 'India'),
            data.get('aadhar', ''),
            data.get('pan', ''),
            aadhar_image,
            pan_image,
            data.get('bank_name', ''),
            data.get('bank_account_number', ''),
            data.get('bank_ifsc', ''),
            data.get('salary', 0),
            data.get('expertise', ''),
            photo,
            data.get('spouse_name') or None,
            data.get('spouse_mob') or None,
            data.get('spouse_email') or None,
            data.get('tenth_school') or None,
            data.get('tenth_board') or None,
            data.get('tenth_year') or None,
            data.get('tenth_marks') or None,
            tenth_marksheet,
            data.get('twelfth_school') or None,
            data.get('twelfth_board') or None,
            data.get('twelfth_year') or None,
            data.get('twelfth_marks') or None,
            twelfth_marksheet,
            data.get('ug_college') or None,
            data.get('ug_degree') or None,
            data.get('ug_year') or None,
            data.get('ug_marks') or None,
            ug_degree_image,
            data.get('pg_college') or None,
            data.get('pg_degree') or None,
            data.get('pg_year') or None,
            data.get('pg_marks') or None,
            pg_degree_image,
            data.get('father_name') or None,
            data.get('father_mobile') or None,
            data.get('mother_name') or None,
            data.get('mother_mobile') or None,
            data.get('joining_date') or None
        )
        cursor.execute(sql, values)

        # If trial, set status=2 and valid_till = joining_date + 7 days
        if status_value == 2:
            cursor.execute("""
                UPDATE employee_header_all 
                SET valid_till = DATE_ADD(joining_date, INTERVAL 7 DAY), status = 2
                WHERE eha_id = %s
            """, (eha_id,))
        else:
            # Normal active employee: create user login
            cursor.execute("""
                INSERT INTO user_login (ulh_id, eha_id, mobile_number, password, role, status, is_first_login)
                VALUES (%s, %s, %s, %s, 'employee', 1, 1)
            """, (ulh_id, eha_id, data.get('mob_no'), hashed_pass))
            send_def_pass(data.get('email'), data.get('first_name'), eha_id, default_pass)

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Employee added successfully",
            "eha_id": eha_id,
            **({"password": default_pass} if status_value != 2 else {}),
            "name": f"{data.get('first_name')} {data.get('last_name')}",
            "email": data.get('email')
        }), 201

    except Exception as e:
        print(f"Error adding employee: {str(e)}")
        import traceback
        traceback.print_exc()
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({
            "success": False,
            "message": f"Error adding employee: {str(e)}"
        }), 500


@app.route('/verify-face', methods=['POST'])
def verify_face():
    if 'photo' not in request.files:
        return jsonify({'faceDetected': False, 'message': 'No photo uploaded'}), 400

    file = request.files['photo']

    # Read file into numpy array
    npimg = np.frombuffer(file.read(), np.uint8)
    img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)

    if img is None:
        return jsonify({'faceDetected': False, 'message': 'Invalid image file'}), 400

    # Convert to grayscale for detection
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Detect faces
    faces = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.1,
        minNeighbors=5,
        minSize=(30, 30)
    )

    if len(faces) == 0:
        return jsonify({'faceDetected': False, 'message': 'No face detected in photo'})
    
    if len(faces) > 1:
        return jsonify({'faceDetected': False, 'message': 'Multiple faces detected, upload a solo photo'})

    return jsonify({'faceDetected': True, 'message': 'Face detected'})


# GET /employees
#FOR PAGE
@app.route("/employees", methods=["GET"])
def get_employees():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        status = request.args.get('status', 1, type=int)
        offset = (page - 1) * per_page
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT
                emp.eha_id,
                emp.first_name,
                emp.middle_name,
                emp.last_name,
                emp.expertise,
                emp.salary,
                emp.mob_no,
                emp.email,
                emp.valid_till,
                emp.status,

                GROUP_CONCAT(
                    DISTINCT CONCAT(
                        tha.team_name,
                        ' (',
                        tec.designation,
                        ')'
                    )
                    SEPARATOR ', '
                ) AS team_designations

            FROM employee_header_all emp

            LEFT JOIN team_emp_combination_all tec
                ON emp.eha_id = tec.eha_id
                AND tec.unlinked_on IS NULL

            LEFT JOIN team_header_all tha
                ON tec.tha_id = tha.tha_id

            WHERE emp.status = %s

            GROUP BY emp.eha_id

            ORDER BY emp.id DESC

            LIMIT %s OFFSET %s
        """, (status, per_page, offset))
        
        employees = cursor.fetchall()
        conn.close()
        
        return jsonify({"employees": employees}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500
    
 
#CHECK IF NOTHING EXISTS ON THAT DAY FOR GEN
@app.route("/check-valid-fromGen", methods=["GET"])
def check_valid_from():
    try:
        exclude_id = request.args.get("exclude_id")
        date = request.args.get('date')
        if not date:
            return jsonify({"exists": False}), 200
        
        conn = get_db()
        cursor = conn.cursor()
        if exclude_id:
            cursor.execute("SELECT hgs_id FROM hr_general_settings_all WHERE valid_from = %s", (date,))
        else:
            cursor.execute("""
                SELECT COUNT(*) as count FROM hr_general_settings_all 
                WHERE valid_from = %s
            """, (date,))
        existing = cursor.fetchone()
        conn.close()
        
        return jsonify({"exists": existing is not None}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500

@app.route("/check-special-emp", methods=["GET"])
def check_special_emp():
    try:
        eha_id = request.args.get('eha_id')
        if not eha_id:
            return jsonify({"exists": False}), 200
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT hsa_id FROM hr_special_approvals_settings_all WHERE eha_id = %s", (eha_id,))
        existing = cursor.fetchone()
        conn.close()
        
        return jsonify({"exists": existing is not None}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500





#FOR SEARCH MODAL
@app.route("/employees/search", methods=["GET"])
def search_employees():
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        search = request.args.get('search', '')
        
        conn = get_db()
        cursor = conn.cursor()
        
        if search:
            cursor.execute("""
                SELECT DISTINCT
                    e.eha_id,
                    e.first_name,
                    e.last_name

                FROM employee_header_all e

                WHERE e.status = 1

                AND (
                    e.eha_id LIKE %s
                    OR e.first_name LIKE %s
                    OR e.last_name LIKE %s
                )

                ORDER BY e.eha_id ASC
                LIMIT 50
            """, (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ))
        else:
            cursor.execute("""
                SELECT
                    e.eha_id,
                    e.first_name,
                    e.last_name

                FROM employee_header_all e

                WHERE e.status = 1

                ORDER BY e.eha_id ASC
                LIMIT 50
            """)
        
        employees = cursor.fetchall()
        conn.close()
        
        return jsonify({"employees": employees}), 200
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500



    
#FOR ALL DOWNLOAD BY STATUS
@app.route("/getemployees/all", methods=["GET"])
def get_all_employees():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT *
            FROM employee_header_all 
            WHERE status = 1
            ORDER BY id DESC
        """)
        active_employees = cursor.fetchall()
        
        cursor.execute("""
            SELECT *
            FROM employee_header_all 
            WHERE status = 0
            ORDER BY id DESC
        """)
        
        inactive_employees = cursor.fetchall()

        cursor.execute("""
            SELECT *
            FROM employee_header_all 
            WHERE status = 2
            ORDER BY id DESC
        """)

        trial_employees = cursor.fetchall()

        cursor.execute("""
            SELECT *
            FROM employee_header_all 
            WHERE status = 3
            ORDER BY id DESC
        """)

        rejected_employees = cursor.fetchall()
        
        conn.close()
        
        return jsonify({
            "active": [dict(emp) for emp in active_employees],
            "inactive": [dict(emp) for emp in inactive_employees],
            "trial" : [dict(emp) for emp in trial_employees],
            "rejected" : [dict(emp) for emp in rejected_employees]
        }), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500

# PUT /terminate/<eha_id>
@app.route("/terminate/<eha_id>", methods=["PUT"])
def terminate_employee(eha_id):
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403

        data = request.get_json()
        remarks = data.get("remarks", "").strip()

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT status FROM employee_header_all WHERE eha_id = %s",
            (eha_id,)
        )

        row = cursor.fetchone()

        if not row:
            return jsonify({"message": "Employee not found"}), 404

        current_status = row['status']

        # Trial Employee -> Rejected
        if current_status == 2:

            if not remarks:
                return jsonify({
                    "message": "Remarks required for rejection"
                }), 400

            cursor.execute("""
                UPDATE employee_header_all
                SET
                    status = 3,
                    valid_till = CURDATE(),
                    remarks = %s
                WHERE eha_id = %s
            """, (remarks, eha_id))

        # Active Employee -> Terminated
        else:

            cursor.execute("""
                UPDATE employee_header_all
                SET
                    status = 0,
                    valid_till = CURDATE(),
                    remarks = %s
                WHERE eha_id = %s
            """, (remarks, eha_id))

            cursor.execute("""
                UPDATE user_login
                SET
                    status = 0,
                    is_locked = 1,
                    login_token = NULL
                WHERE eha_id = %s
            """, (eha_id,))

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Employee terminated successfully"
        }), 200

    except Exception as e:
        return jsonify({"message": str(e)}), 500
    
@app.route("/reactivate/<eha_id>", methods=["PUT"])
def reactivate_employee(eha_id):
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT status, joining_date FROM employee_header_all WHERE eha_id = %s", (eha_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"message": "Employee not found"}), 404
        
        current_status = row['status']
        joining_date = row['joining_date']

        if current_status == 0:         
            new_status = 1               
            valid_till = None
        elif current_status == 3:       
            new_status = 2              
            if joining_date:
                valid_date = joining_date + timedelta(days=7)
                valid_till = valid_date.strftime('%Y-%m-%d')
            else:
                valid_till = None
        
        
        cursor.execute("""
            UPDATE employee_header_all 
            SET status = %s, valid_till = %s 
            WHERE eha_id = %s
        """, (new_status, valid_till, eha_id))
        
        if current_status ==0:
            cursor.execute("""
                UPDATE user_login 
                SET status = 1, is_locked = 0, failed_attempts = 0 
                WHERE eha_id = %s
            """, (eha_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Employee reactivated"}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500


@app.route("/extend_trial/<eha_id>", methods=["PUT"])
def extend_trial(eha_id):
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403

        data = request.get_json()

        valid_till = data.get('days')

        if not valid_till:
            return jsonify({"message": "Valid date required"}), 400

        try:
            new_valid = datetime.strptime(valid_till, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"message": "Invalid date format"}), 400

        if new_valid <= datetime.now().date():
            return jsonify({"message": "Date must be in the future"}), 400

        conn = get_db()
        cursor = conn.cursor()

        # Get current status
        cursor.execute(
            "SELECT valid_till, status FROM employee_header_all WHERE eha_id = %s",
            (eha_id,)
        )

        row = cursor.fetchone()

        if not row:
            return jsonify({"message": "Employee not found"}), 404

        if row['status'] != 2:
            return jsonify({"message": "Employee is not on trial"}), 400
        
        # Update the database
        cursor.execute("""
            UPDATE employee_header_all 
            SET valid_till = %s 
            WHERE eha_id = %s AND status = 2
        """, (new_valid, eha_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Trial extended till {valid_till} ",
            "new_valid_till": new_valid.strftime('%Y-%m-%d')
        }), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500

@app.route("/getemployee/<eha_id>", methods=["GET"])
def get_employee(eha_id):   
    try:
        if 'empid' not in session:
            return jsonify({"message": "Not logged in"}), 401
        if session['role'] not in ['hr', 'admin', 'director']:
            return jsonify({"message": "Access Denied!"}), 403

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM employee_header_all WHERE eha_id = %s", (eha_id,))
        employee = cursor.fetchone()
        conn.close()

        if not employee:
            return jsonify({"message": "Employee not found"}), 404
        
        return jsonify({
            "employee": dict(employee)
        }), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/updateemployee/<eha_id>', methods=['PUT'])
def update_employee(eha_id):
    if 'empid' not in session:
        return jsonify({"message": "Not logged in"}), 401
    if session['role'] not in ['hr', 'admin', 'director']:
        return jsonify({"message": "Access Denied"}), 403
    
    data = request.form
    files = request.files 

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM employee_header_all WHERE eha_id = %s", (eha_id,))
    employee = cursor.fetchone()

    if not employee:
        conn.close()
        return jsonify({"message": "Employee not found!"}), 404

    def save_file(file_field, field_name):
        if file_field in files and files[file_field]:
            file = files[file_field]
            if file and file.filename and allowed_file(file.filename):
                filename = f"{eha_id}_{field_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                return filepath
        return None

    aadhar_image = save_file('aadhar_image', 'aadhar')
    pan_image = save_file('pan_image', 'pan')
    tenth_marksheet = save_file('tenth_marksheet', 'tenth')
    twelfth_marksheet = save_file('twelfth_marksheet', 'twelfth')
    ug_degree_image = save_file('ug_degree_image', 'ug')
    pg_degree_image = save_file('pg_degree_image', 'pg')
    photo = save_file('photo', 'photo')

    try:
        sql = """
            UPDATE employee_header_all SET 
                first_name = %s,
                middle_name = %s,
                last_name = %s,
                marital_status = %s,
                gender = %s,
                gender_other = %s,
                dob = %s,
                mob_no = %s,
                email = %s,
                type = %s,
                status = %s,
                valid_till = %s,
                joining_date = %s,
                present_address_line1 = %s,
                present_address_line2 = %s,
                present_city = %s,
                present_state = %s,
                present_pincode = %s,
                present_country = %s,
                permanent_address_line1 = %s,
                permanent_address_line2 = %s,
                permanent_city = %s,
                permanent_state = %s,
                permanent_pincode = %s,
                permanent_country = %s,
                aadhar = %s,
                pan = %s,
                bank_name = %s,
                bank_account_number = %s,
                bank_ifsc = %s,
                salary = %s,
                expertise = %s,
                spouse_name = %s,
                spouse_mob = %s,
                spouse_email = %s,
                father_name = %s,
                father_mobile = %s,
                mother_name = %s,
                mother_mobile = %s,
                tenth_school = %s,
                tenth_board = %s,
                tenth_year = %s,
                tenth_marks = %s,
                twelfth_school = %s,
                twelfth_board = %s,
                twelfth_year = %s,
                twelfth_marks = %s,
                ug_college = %s,
                ug_degree = %s,
                ug_year = %s,
                ug_marks = %s,
                pg_college = %s,
                pg_degree = %s,
                pg_year = %s,
                pg_marks = %s,
                aadhar_image_path = COALESCE(NULLIF(%s, ''), aadhar_image_path),
                pan_image_path = COALESCE(NULLIF(%s, ''), pan_image_path),
                photo_path = COALESCE(NULLIF(%s, ''), photo_path),
                tenth_marksheet_path = COALESCE(NULLIF(%s, ''), tenth_marksheet_path),
                twelfth_marksheet_path = COALESCE(NULLIF(%s, ''), twelfth_marksheet_path),
                ug_degree_image_path = COALESCE(NULLIF(%s, ''), ug_degree_image_path),
                pg_degree_image_path = COALESCE(NULLIF(%s, ''), pg_degree_image_path)
            WHERE eha_id = %s
        """
        
        values = (
            data.get('first_name', employee['first_name']),
            data.get('middle_name', employee['middle_name']),
            data.get('last_name', employee['last_name']),
            data.get('marital_status', employee['marital_status']),
            data.get('gender', employee['gender']),
            data.get('gender_other', employee['gender_other']),
            data.get('dob', employee['dob']),
            data.get('mob_no', employee['mob_no']),
            data.get('email', employee['email']),
            data.get('type', employee['type']),
            data.get('status', employee['status']),
            data.get('valid_till', employee['valid_till']),
            data.get('joining_date', employee['joining_date']),
            data.get('present_address_line1', employee['present_address_line1']),
            data.get('present_address_line2', employee['present_address_line2']),
            data.get('present_city', employee['present_city']),
            data.get('present_state', employee['present_state']),
            data.get('present_pincode', employee['present_pincode']),
            data.get('present_country', employee['present_country']),
            data.get('permanent_address_line1', employee['permanent_address_line1']),
            data.get('permanent_address_line2', employee['permanent_address_line2']),
            data.get('permanent_city', employee['permanent_city']),
            data.get('permanent_state', employee['permanent_state']),
            data.get('permanent_pincode', employee['permanent_pincode']),
            data.get('permanent_country', employee['permanent_country']),
            data.get('aadhar', employee['aadhar']),
            data.get('pan', employee['pan']),
            data.get('bank_name', employee['bank_name']),
            data.get('bank_account_number', employee['bank_account_number']),
            data.get('bank_ifsc', employee['bank_ifsc']),
            data.get('salary', employee['salary']),
            data.get('expertise', employee['expertise']),
            data.get('spouse_name', employee['spouse_name']),
            data.get('spouse_mob', employee['spouse_mob']),
            data.get('spouse_email', employee['spouse_email']),
            data.get('father_name', employee['father_name']),
            data.get('father_mobile', employee['father_mobile']),
            data.get('mother_name', employee['mother_name']),
            data.get('mother_mobile', employee['mother_mobile']),
            data.get('tenth_school', employee['tenth_school']),
            data.get('tenth_board', employee['tenth_board']),
            data.get('tenth_year', employee['tenth_year']),
            data.get('tenth_marks', employee['tenth_marks']),
            data.get('twelfth_school', employee['twelfth_school']),
            data.get('twelfth_board', employee['twelfth_board']),
            data.get('twelfth_year', employee['twelfth_year']),
            data.get('twelfth_marks', employee['twelfth_marks']),
            data.get('ug_college', employee['ug_college']),
            data.get('ug_degree', employee['ug_degree']),
            data.get('ug_year', employee['ug_year']),
            data.get('ug_marks', employee['ug_marks']),
            data.get('pg_college', employee['pg_college']),
            data.get('pg_degree', employee['pg_degree']),
            data.get('pg_year', employee['pg_year']),
            data.get('pg_marks', employee['pg_marks']),
            aadhar_image or '',
            pan_image or '',
            photo or '',
            tenth_marksheet or '',
            twelfth_marksheet or '',
            ug_degree_image or '',
            pg_degree_image or '',
            eha_id
        )
        
        cursor.execute(sql, values)
        conn.commit()
        conn.close()
        return jsonify({"message": "Employee updated successfully"}), 200
        
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    

#TEAM MANAGEMENT////////////////////////////////////////////////////////////////////////////////////////////////

@app.route("/create_team", methods=["POST"])
def create_team():
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        data = request.get_json()
        
        team_name = data.get('team_name', '').strip()
        team_leader = data.get('team_leader', {})
        team_members = data.get('team_members', [])

        leader_id = team_leader.get('eha_id')
        leader_designation = team_leader.get('designation', '').strip()
        
        if not team_name:
            return jsonify({"message": "Team name is required"}), 400
        if not leader_id:
            return jsonify({"message": "Team leader is required"}), 400
        if not team_members or len(team_members) == 0:
            return jsonify({"message": "At least one team member is required"}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if team name already exists
        cursor.execute("SELECT id FROM team_header_all WHERE team_name = %s AND status = 1", (team_name,))
        existing = cursor.fetchone()
        if existing:
            conn.close()
            return jsonify({"message": "Team name already exists"}), 409
        
        # Generate THA ID
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'team_header_all'")
        header = cursor.fetchone()
        
        if not header or header['last_id'] is None:
            cursor.execute("INSERT INTO unique_header_all (table_name, prefix, last_id) VALUES ('team_header_all', 'THA', '1')")
            conn.commit()
            tha_id = "THA-00001"
        else:
            new_id = int(header['last_id']) + 1
            tha_id = f"THA-{new_id:05d}"
            cursor.execute("UPDATE unique_header_all SET last_id = %s, modified_on = NOW() WHERE table_name = 'team_header_all'", (new_id,))
            conn.commit()
        
        # Insert team header
        import json
        cursor.execute("""
            INSERT INTO team_header_all (tha_id, team_name, team_members, team_leader, status)
            VALUES (%s, %s, %s, %s, 1)
        """, (tha_id, team_name, json.dumps(team_members), json.dumps([leader_id])))
        
        # Generate TEC entries
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'team_emp_combination_all'")
        tec_header = cursor.fetchone()
        
        if not tec_header or tec_header['last_id'] is None:
            cursor.execute("INSERT INTO unique_header_all (table_name, prefix, last_id) VALUES ('team_emp_combination_all', 'TEC', '0')")
            conn.commit()
            tec_last_id = 0
        else:
            tec_last_id = int(tec_header['last_id'])
        
        # Insert leader (link_type = 2)
        tec_last_id += 1
        tec_id = f"TEC-{tec_last_id:05d}"
        cursor.execute("""
                INSERT INTO team_emp_combination_all
                (
                    tec_id,
                    tha_id,
                    eha_id,
                    designation,
                    linked_on,
                    link_type,
                    line_no
                )
                VALUES (%s, %s, %s, %s, NOW(), 2, 1)
            """, (
                tec_id,
                tha_id,
                leader_id,
                leader_designation
            ))
        
    
        # Insert members (link_type = 1)
        line_no = 2

        for member in team_members:

            member_id = member.get('eha_id')
            designation = member.get('designation', '').strip()

            tec_last_id += 1
            tec_id = f"TEC-{tec_last_id:05d}"

            cursor.execute("""
                INSERT INTO team_emp_combination_all
                (
                    tec_id,
                    tha_id,
                    eha_id,
                    designation,
                    linked_on,
                    link_type,
                    line_no
                )
                VALUES (%s, %s, %s, %s, NOW(), 1, %s)
            """, (
                tec_id,
                tha_id,
                member_id,
                designation,
                line_no
            ))

            line_no += 1
        
        # Update TEC last_id
        cursor.execute("UPDATE unique_header_all SET last_id = %s, modified_on = NOW() WHERE table_name = 'team_emp_combination_all'", (tec_last_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": "Team created successfully",
            "tha_id": tha_id
        }), 201
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500

@app.route("/teams", methods=["GET"])
def get_teams():
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        status_para= request.args.get('status',"active")
        status_map={"active":1,"inactive":0}
        status = status_map.get(status_para,1)

        
        conn = get_db()
        cursor = conn.cursor()
        cursor2 = conn.cursor()
        #active
        cursor.execute("""
            SELECT * FROM team_header_all
            WHERE status = %s
            ORDER BY id DESC
        """,(status,))
        teams = cursor.fetchall()


        
        result = []
        for team in teams:
            team_dict = dict(team)
            
            #leader
            cursor.execute("""
                SELECT
                    tec.eha_id,
                    tec.designation,
                    emp.first_name,
                    emp.last_name
                FROM team_emp_combination_all tec
                JOIN employee_header_all emp
                    ON tec.eha_id = emp.eha_id
                WHERE
                    tec.tha_id = %s
                    AND tec.link_type = 2
                    AND tec.unlinked_on IS NULL
            """, (team_dict['tha_id'],))
            leader = cursor.fetchone()
            
            if leader:
                team_dict['_leader'] = dict(leader)
                team_dict['leader_name'] = f"{leader['first_name']} {leader['last_name']}"
                team_dict['leader_designation'] = leader.get('designation')
            else:
                team_dict['_leader'] = None
                team_dict['leader_name'] = "—"
            
            #members
            cursor2.execute("""
                SELECT
                    tec.eha_id,
                    tec.designation,
                    emp.first_name,
                    emp.last_name
                FROM team_emp_combination_all tec
                JOIN employee_header_all emp
                    ON tec.eha_id = emp.eha_id
                WHERE
                    tec.tha_id = %s
                    AND tec.link_type = 1
                    AND tec.unlinked_on IS NULL
            """, (team_dict['tha_id'],))
            members = cursor2.fetchall()
            
            team_dict['_members'] = [dict(m) for m in members] if members else []
            team_dict['member_count'] = len(members) if members else 0
            
            # Convert datetime to string
            if team_dict.get('inserted_on'):
                team_dict['inserted_on'] = str(team_dict['inserted_on'])
            if team_dict.get('valid_till'):
                team_dict['valid_till'] = str(team_dict['valid_till'])
            
            result.append(team_dict)
        
        conn.close()
        return jsonify({"teams": result}), 200
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500
    
@app.route("/teams/<tha_id>", methods=["PUT"])
def update_team(tha_id):
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        data = request.get_json()
        team_name = data.get('team_name', '').strip()
        team_leader = data.get('team_leader', {})
        team_members = data.get('team_members', [])

        leader_id = team_leader.get('eha_id')
        leader_designation = team_leader.get('designation', '').strip()
        
        if not team_name:
            return jsonify({"message": "Team name is required"}), 400
        if not leader_id:
            return jsonify({"message": "Team leader is required"}), 400
        if not team_members:
            return jsonify({"message": "At least one member is required"}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check team exists
        cursor.execute("SELECT * FROM team_header_all WHERE tha_id = %s", (tha_id,))
        team = cursor.fetchone()
        if not team:
            conn.close()
            return jsonify({"message": "Team not found"}), 404

        
        # Update team header
        cursor.execute("""
            UPDATE team_header_all
            SET team_name = %s, team_members = %s, team_leader = %s
            WHERE tha_id = %s
        """, (team_name, json.dumps(team_members), json.dumps([leader_id]), tha_id))

        
                
        # Unlink all current members (set unlinked_on)
        cursor.execute("""
            UPDATE team_emp_combination_all
            SET unlinked_on = NOW()
            WHERE tha_id = %s AND unlinked_on IS NULL
        """, (tha_id,))
        
        # Generate new TEC entries
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'team_emp_combination_all'")
        tec_header = cursor.fetchone()
        tec_last_id = int(tec_header['last_id']) if tec_header and tec_header['last_id'] else 0
        
        # Insert leader
        tec_last_id += 1
        tec_id = f"TEC-{tec_last_id:05d}"
        cursor.execute("""
            INSERT INTO team_emp_combination_all
            (
                tec_id,
                tha_id,
                eha_id,
                designation,
                linked_on,
                link_type,
                line_no
            )
            VALUES (%s, %s, %s, %s, NOW(), 2, 1)
        """, (
            tec_id,
            tha_id,
            leader_id,
            leader_designation
        ))
        
        # Insert members
        line_no = 2
        for member in team_members:
            member_id = member.get('eha_id')
            designation = member.get('designation', '').strip()

            tec_last_id += 1
            tec_id = f"TEC-{tec_last_id:05d}"

            cursor.execute("""
                INSERT INTO team_emp_combination_all
                (
                    tec_id,
                    tha_id,
                    eha_id,
                    designation,
                    linked_on,
                    link_type,
                    line_no
                )
                VALUES (%s, %s, %s, %s, NOW(), 1, %s)
            """, (
                tec_id,
                tha_id,
                member_id,
                designation,
                line_no
            ))

            line_no += 1
        
        # Update TEC last_id
        cursor.execute("UPDATE unique_header_all SET last_id = %s, modified_on = NOW() WHERE table_name = 'team_emp_combination_all'", (tec_last_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Team updated successfully"}), 200
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500
    
@app.route("/teams/<tha_id>/status", methods=["PUT"])
def toggle_team_status(tha_id):
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        data = request.get_json()
        new_status = data.get('status')
        
        if new_status not in [0, 1]:
            return jsonify({"message": "Invalid status"}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check team exists
        cursor.execute("SELECT * FROM team_header_all WHERE tha_id = %s", (tha_id,))
        team = cursor.fetchone()
        if not team:
            conn.close()
            return jsonify({"message": "Team not found"}), 404
        
        # Update status
        if new_status == 0:
            # Deactivating - set valid_till
            cursor.execute("""
                UPDATE team_header_all
                SET status = 0, valid_till = NOW()
                WHERE tha_id = %s
            """, (tha_id,))
            
        else:
            # Activating - clear valid_till
            cursor.execute("""
                UPDATE team_header_all
                SET status = 1, valid_till = NULL
                WHERE tha_id = %s
            """, (tha_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Team {'activated' if new_status == 1 else 'deactivated'} successfully"
        }), 200
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500






#AGENDAS AND REPORTS/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

#AGENDA
def generate_wa_id(cursor):
    cursor.execute(
        "SELECT * FROM unique_header_all WHERE table_name = 'work_assignment_all'"
    )
    header = cursor.fetchone()
    if not header or header["last_id"] is None:
        cursor.execute(
            "INSERT INTO unique_header_all (table_name, prefix, last_id) "
            "VALUES ('work_assignment_all', 'WA', '0')"
        )
        new_id = 1
    else:
        new_id = int(header["last_id"]) + 1
        cursor.execute(
            "UPDATE unique_header_all SET last_id = %s, modified_on = NOW() "
            "WHERE table_name = 'work_assignment_all'",
            (str(new_id),)
        )
    return f"WA-{new_id:05d}"

def resolve_window(empid, target_date, cursor):
    cursor.execute("""
        SELECT 
            TIME_FORMAT(daily_agenda_start, '%%H:%%i') AS daily_agenda_start,
            TIME_FORMAT(daily_agenda_end, '%%H:%%i') AS daily_agenda_end
        FROM hr_general_settings_all
        WHERE valid_from <= %s
        AND (valid_till IS NULL OR valid_till >= %s)
        ORDER BY valid_from DESC
        LIMIT 1
    """, (target_date, target_date))
    gs =  cursor.fetchone()
    if not gs:
        return None, None

    window_start = gs.get("daily_agenda_start")
    window_end = gs.get("daily_agenda_end")
    
    
    print("hello")
    cursor.execute("""
        SELECT 
            TIME_FORMAT(early_agenda_submit, '%%H:%%i') AS early_agenda_submit,
            TIME_FORMAT(late_agenda_submit,  '%%H:%%i') AS late_agenda_submit
        FROM hr_special_approvals_settings_all
        WHERE eha_id = %s
        AND valid_from <= %s
        AND (valid_till IS NULL OR valid_till >= %s)
        ORDER BY valid_from DESC
        LIMIT 1
    """, (empid, target_date, target_date))
    hsa = cursor.fetchone()
    
    if hsa:
        window_start = hsa.get("early_agenda_submit")
        window_end = hsa.get("late_agenda_submit")
        
        return window_start, window_end,1 
    return window_start, window_end,0 

def to_time(val):
    print("GEGE")
    if isinstance(val, str):
        return datetime.strptime(val, "%H:%M").time()
    elif isinstance(val, timedelta):
        total_seconds = int(val.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes = remainder // 60
        return time(hours, minutes)
    elif isinstance(val, time):
        return val
    return None

@app.route("/window_status", methods=["GET"])
def get_agenda_window():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401
    print(empid)

    now   = datetime.now()
    today = now.date()

    conn   = get_db()
    cursor = conn.cursor()

    try:
        window_start, window_end,_ = resolve_window(empid, today, cursor)
        if window_start is None:
            return jsonify({"error": "No active settings found"}), 500

        # Convert string to time object
        window_start = to_time(window_start)
        window_end   = to_time(window_end)
        if window_start is None or window_end is None:
            return jsonify({"error": "No active settings found"}), 500

        window_open  = datetime.combine(today, window_start)
        window_close = datetime.combine(today, window_end)
        is_open = window_open <= now <= window_close

        return jsonify({
            "is_open":      is_open,
            "window_start": window_start.strftime("%H:%M"),  # 24hr
            "window_end":   window_end.strftime("%H:%M")     # 24hr
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())  # full error in terminal
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@app.route("/daily_agenda/add", methods=["POST"])
def add_daily_agenda():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    data = request.get_json()
    if not data or "tasks" not in data or "tec_id" not in data:
        return jsonify({"error": "Missing tasks or tec_id"}), 400

    task_list = data["tasks"]
    tec_id    = data["tec_id"]

    if not task_list:
        return jsonify({"error": "Tasks array is empty"}), 400

    now       = datetime.now()
    today     = now.date()

    conn   = get_db()
    cursor = conn.cursor()

    try:
        # --- Window checks (unchanged) ---
        window_start_str, window_end_str, _ = resolve_window(empid, today, cursor)
        if window_start_str is None:
            return jsonify({"error": "No active general settings found"}), 500

        window_start = to_time(window_start_str)
        window_end   = to_time(window_end_str)
        if window_start is None or window_end is None:
            return jsonify({"error": "No active settings found"}), 500

        window_open  = datetime.combine(today, window_start)
        window_close = datetime.combine(today, window_end)
        is_open      = window_open <= now <= window_close
        if now < window_open:
            return jsonify({"error": "Window Not Started Yet"}), 410
        submit_status = 0 if is_open else 1

        

        # --- Validate the requested tec_id ---
        cursor.execute("""
            SELECT tec_id, tha_id
            FROM team_emp_combination_all
            WHERE tec_id = %s
            AND eha_id = %s
            AND unlinked_on IS NULL
            AND (valid_till IS NULL OR valid_till >= CURDATE())
        """, (tec_id, empid))
        assign = cursor.fetchone()
        if not assign:
            return jsonify({"error": "TEC ID not valid or assignment inactive"}), 400

        tec_id = assign["tec_id"]
        tha_id = assign["tha_id"]

        day_num = today.day
        day_col = f"day_{day_num}"
        mm_yy   = today.strftime("%m-%Y")

        # --- Find existing row for this employee, this team, this month ---
        cursor.execute(
            f"SELECT id, {day_col} AS existing_day "
            "FROM member_daily_agenda_details_all "
            "WHERE eha_id = %s AND tec_id = %s AND mm_yy = %s",
            (empid, tec_id, mm_yy)
        )
        existing_row = cursor.fetchone()

        # --- Case 1: Row exists and today's column already has data → append tasks ---
        if existing_row and existing_row["existing_day"] is not None:
            if not is_open:
                return jsonify({"error": "Window closed – cannot add more tasks"}), 403

            # Load existing payload
            payload = json.loads(existing_row["existing_day"])
            existing_tasks = payload.get("text", [])

            # Determine next task ID
            next_id = max([t["id"] for t in existing_tasks], default=0) + 1
            new_tasks = [
                {"id": next_id + i, "content": t.get("content", "")}
                for i, t in enumerate(task_list)
            ]

            # Append and save
            payload["text"] = existing_tasks + new_tasks
            payload["in_time"] = now.strftime("%Y-%m-%d %H:%M:%S")   # last modification time

            payload_json = json.dumps(payload, ensure_ascii=False)

            cursor.execute(
                f"UPDATE member_daily_agenda_details_all SET {day_col} = %s "
                "WHERE eha_id = %s AND tec_id = %s AND mm_yy = %s",
                (payload_json, empid, tec_id, mm_yy)
            )
            conn.commit()

            return jsonify({
                "message": "Tasks appended successfully",
                "added_count": len(new_tasks),
                "total_tasks": len(payload["text"]),
                "submit_status": "on-time" if payload.get("submit_status", 0) == 0 else "delayed"
            })

        # --- Case 2: Row exists but today's column is NULL, or no row yet → create/update new payload ---
        # Generate a new work assignment ID (used only when creating a brand‑new day payload)
        wa_id = generate_wa_id(cursor)

        tasks_array = [
            {"id": i + 1, "content": t.get("content", "")}
            for i, t in enumerate(task_list)
        ]
        payload_json = json.dumps({
            "in_time":       now.strftime("%Y-%m-%d %H:%M:%S"),
            "wa_id":         wa_id,
            "text":          tasks_array,
            "submit_status": submit_status      # 0 = on time, 1 = delayed
        }, ensure_ascii=False)

        if existing_row:
            # The row already exists but today's column was empty → update it
            cursor.execute(
                f"UPDATE member_daily_agenda_details_all SET {day_col} = %s "
                "WHERE eha_id = %s AND tec_id = %s AND mm_yy = %s",
                (payload_json, empid, tec_id, mm_yy)
            )
        else:
            # No row for this employee + team + month → insert new
            # Generate mda_id
            cursor.execute(
                "SELECT * FROM unique_header_all WHERE table_name = 'member_daily_agenda_details_all'"
            )
            header = cursor.fetchone()
            if not header or header["last_id"] is None:
                cursor.execute(
                    "INSERT INTO unique_header_all (table_name, prefix, last_id) "
                    "VALUES ('member_daily_agenda_details_all', 'MDA', '0')"
                )
                new_id = 1
            else:
                new_id = int(header["last_id"]) + 1
                cursor.execute(
                    "UPDATE unique_header_all SET last_id = %s, modified_on = NOW() "
                    "WHERE table_name = 'member_daily_agenda_details_all'",
                    (str(new_id),)
                )
            mda_id = f"MDA-{new_id:05d}"

            cursor.execute(
                f"INSERT INTO member_daily_agenda_details_all "
                f"(eha_id, mda_id, tha_id, tec_id, mm_yy, {day_col}) "
                f"VALUES (%s, %s, %s, %s, %s, %s)",
                (empid, mda_id, tha_id, tec_id, mm_yy, payload_json)
            )

        conn.commit()

        return jsonify({
            "message":       "Daily agenda submitted successfully",
            "wa_id":         wa_id,
            "for_date":      today.isoformat(),
            "day":           day_num,
            "submit_status": "on-time" if submit_status == 0 else "delayed"
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/delete_task', methods=['POST'])
def delete_task():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    data = request.get_json()
    if not data or "task_id" not in data or "date" not in data or "tec_id" not in data:
        return jsonify({"error": "Missing task_id, date, or tec_id"}), 400

    task_id = data["task_id"]
    date_str = data["date"]
    tec_id = data["tec_id"]

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 1. Find the exact agenda row for this employee + team + month
        cursor.execute(
            f"SELECT {day_col} AS agenda_data FROM member_daily_agenda_details_all "
            "WHERE eha_id = %s AND tec_id = %s AND mm_yy = %s",
            (empid, tec_id, mm_yy)
        )
        row = cursor.fetchone()
        if not row or not row["agenda_data"]:
            return jsonify({"error": "No agenda found for this date and team"}), 404

        payload = json.loads(row["agenda_data"])
        tasks = payload.get("text", [])

        # 2. Remove the task with the given ID
        new_tasks = [t for t in tasks if t.get("id") != task_id]
        if len(new_tasks) == len(tasks):
            return jsonify({"error": "Task ID not found"}), 404

        # 3. Update the payload
        payload["text"] = new_tasks
        payload["in_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # optional
        new_json = json.dumps(payload, ensure_ascii=False)

        cursor.execute(
            f"UPDATE member_daily_agenda_details_all SET {day_col} = %s "
            "WHERE eha_id = %s AND tec_id = %s AND mm_yy = %s",
            (new_json, empid, tec_id, mm_yy)
        )
        conn.commit()

        # 4. Return only the remaining tasks (or the full team view if needed)
        return jsonify({
            "message": "Task deleted",
            "remaining_tasks": new_tasks   # frontend can use this to update view
        }), 200

    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route("/daily_agenda_date/get", methods=["GET"])
def get_daily_agenda_date():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Date required"}), 400

    try:
        target_date = date.fromisoformat(date_str)
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn   = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute(f"""
            SELECT
                tec.tec_id,
                tha.team_name,
                ag.{day_col} AS day_data

            FROM team_emp_combination_all tec

            INNER JOIN team_header_all tha
                ON tha.tha_id = tec.tha_id

            LEFT JOIN member_daily_agenda_details_all ag
                ON ag.tec_id = tec.tec_id
                AND ag.mm_yy = %s

            WHERE tec.eha_id = %s
              AND DATE(tec.linked_on) <= %s
              AND (
                  tec.unlinked_on IS NULL
                  OR DATE(tec.unlinked_on) > %s
              )
              AND (
                  tec.valid_till IS NULL
                  OR DATE(tec.valid_till) >= %s
              )

        """, (mm_yy, empid, target_date, target_date, target_date))

        rows = cursor.fetchall()

        if not rows:
            return jsonify([]), 200

        result = []
        for row in rows:
            raw = row['day_data']

            if raw:
                payload = json.loads(raw) if isinstance(raw, str) else raw
                result.append({
                    "tec_id":        row['tec_id'],
                    "team_name":     row['team_name'],
                    "submitted":     True,
                    "submit_status": payload.get("submit_status"),
                    "in_time":       payload.get("in_time"),
                    "wa_id":         payload.get("wa_id"),
                    "tasks":         payload.get("text", [])
                })
            else:
                result.append({
                    "tec_id":    row['tec_id'],
                    "team_name": row['team_name'],
                    "submitted": False,
                    "tasks":     []
                })

        return jsonify(result), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/daily_agenda_date/team_wise', methods=['GET'])
def get_daily_agenda_team_wise():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Missing date"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn   = get_db()
    cursor = conn.cursor()

    try:

        # date-aware team fetch — mirrors daily_team_submissions/today
        cursor.execute(f"""
            SELECT
                tec.tec_id,
                tha.team_name,
                ag.{day_col} AS agenda_data

            FROM team_emp_combination_all tec

            INNER JOIN team_header_all tha
                ON tha.tha_id = tec.tha_id

            LEFT JOIN member_daily_agenda_details_all ag
                ON ag.tec_id = tec.tec_id
                AND ag.mm_yy = %s

            WHERE tec.eha_id = %s
              AND DATE(tec.linked_on) <= %s
              AND (
                  tec.unlinked_on IS NULL
                  OR DATE(tec.unlinked_on) > %s
              )
              AND (
                  tec.valid_till IS NULL
                  OR DATE(tec.valid_till) >= %s
              )

        """, (mm_yy, empid, target_date, target_date, target_date))

        rows = cursor.fetchall()

        if not rows:
            return jsonify([]), 200

        result = []
        for row in rows:
            raw = row['agenda_data']

            if raw:
                payload = json.loads(raw) if isinstance(raw, str) else raw
                result.append({
                    "team_name":     row['team_name'],
                    "tec_id":        row['tec_id'],
                    "submitted":     True,
                    "in_time":       payload.get("in_time"),
                    "submit_status": payload.get("submit_status", 0),
                    "tasks":         payload.get("text", [])
                })
            else:
                result.append({
                    "team_name": row['team_name'],
                    "tec_id":    row['tec_id'],
                    "submitted": False
                })

        return jsonify(result), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

#CALENDAR

@app.route('/daily_agenda/monthly_status', methods=['GET'])
def monthly_agenda_status():

    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    month = request.args.get('month', type=int)
    year  = request.args.get('year', type=int)

    if not month or not year:
        return jsonify({'error': 'month and year required'}), 400

    mm_yy         = f"{str(month).zfill(2)}-{year}"
    days_in_month = cal.monthrange(year, month)[1]
    today         = datetime.now().date()

    conn   = get_db()
    cursor = conn.cursor()

    try:

        # calendar for the month
        cursor.execute("""
            SELECT
                c.calendar_date,
                COALESCE(
                    (
                        SELECT h.previous_type
                        FROM company_calendar_history_all h
                        WHERE h.calendar_date = c.calendar_date
                        ORDER BY h.changed_at ASC
                        LIMIT 1
                    ),
                    c.day_type
                ) AS day_type
            FROM company_calendar_settings_all c
            WHERE MONTH(c.calendar_date) = %s
              AND YEAR(c.calendar_date)  = %s
        """, (month, year))

        calendar_map = {
            str(r['calendar_date']): r['day_type']
            for r in cursor.fetchall()
        }

        result = []

        for day in range(1, days_in_month + 1):

            date_str = f"{year}-{str(month).zfill(2)}-{str(day).zfill(2)}"
            day_date = datetime(year, month, day).date()

            # future
            if day_date > today:
                result.append({
                    'date':     date_str,
                    'status':   'future',
                    'day_type': 'working'
                })
                continue

            # holiday / weekoff
            day_type = calendar_map.get(date_str, 'working')
            if day_type in ('holiday', 'weekoff'):
                result.append({
                    'date':     date_str,
                    'status':   'grey',
                    'day_type': day_type
                })
                continue

            # get tec_ids active on exactly this day + their submission status
            cursor.execute(f"""
                SELECT
                    tec.tec_id,
                    CASE
                        WHEN ag.day_{day} IS NULL THEN 'missed'
                        WHEN JSON_EXTRACT(ag.day_{day}, '$.submit_status') = 1 THEN 'delayed'
                        ELSE 'ontime'
                    END AS day_status

                FROM employee_header_all eh

                INNER JOIN team_emp_combination_all tec
                    ON tec.eha_id = eh.eha_id
                    AND DATE(tec.linked_on) <= %s
                    AND (
                        tec.unlinked_on IS NULL
                        OR DATE(tec.unlinked_on) > %s
                    )
                    AND (
                        tec.valid_till IS NULL
                        OR DATE(tec.valid_till) >= %s
                    )

                INNER JOIN team_header_all tha
                    ON tha.tha_id = tec.tha_id

                LEFT JOIN member_daily_agenda_details_all ag
                    ON ag.tec_id = tec.tec_id
                    AND ag.mm_yy = %s

                WHERE eh.status = 1
                  AND eh.eha_id = %s

            """, (day_date, day_date, day_date, mm_yy, empid))

            day_rows = cursor.fetchall()

            # no active teams on this day
            if not day_rows:
                result.append({
                    'date':     date_str,
                    'status':   'grey',
                    'day_type': day_type
                })
                continue

            missed  = any(r['day_status'] == 'missed'  for r in day_rows)
            delayed = any(r['day_status'] == 'delayed' for r in day_rows)

            if missed:
                status = 'red'
            elif delayed:
                status = 'yellow'
            else:
                status = 'green'

            result.append({
                'date':     date_str,
                'status':   status,
                'day_type': day_type
            })

        return jsonify(result), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/daily_report/monthly_status', methods=['GET'])
def monthly_report_status():

    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    month = request.args.get('month', type=int)
    year  = request.args.get('year', type=int)

    if not month or not year:
        return jsonify({'error': 'month and year required'}), 400

    mm_yy         = f"{str(month).zfill(2)}-{year}"
    days_in_month = cal.monthrange(year, month)[1]
    today         = datetime.now().date()

    conn   = get_db()
    cursor = conn.cursor()

    try:

        # calendar for the month
        cursor.execute("""
            SELECT
                c.calendar_date,
                COALESCE(
                    (
                        SELECT h.previous_type
                        FROM company_calendar_history_all h
                        WHERE h.calendar_date = c.calendar_date
                        ORDER BY h.changed_at ASC
                        LIMIT 1
                    ),
                    c.day_type
                ) AS day_type
            FROM company_calendar_settings_all c
            WHERE MONTH(c.calendar_date) = %s
              AND YEAR(c.calendar_date)  = %s
        """, (month, year))

        calendar_map = {
            str(r['calendar_date']): r['day_type']
            for r in cursor.fetchall()
        }

        result = []

        for day in range(1, days_in_month + 1):

            date_str = f"{year}-{str(month).zfill(2)}-{str(day).zfill(2)}"
            day_date = datetime(year, month, day).date()

            # future
            if day_date > today:
                result.append({
                    'date':     date_str,
                    'status':   'future',
                    'day_type': 'working'
                })
                continue

            # holiday / weekoff
            day_type = calendar_map.get(date_str, 'working')
            if day_type in ('holiday', 'weekoff'):
                result.append({
                    'date':     date_str,
                    'status':   'grey',
                    'day_type': day_type
                })
                continue

            # get tec_ids active on exactly this day + their submission status
            cursor.execute(f"""
                SELECT
                    tec.tec_id,
                    CASE
                        WHEN ag.day_{day} IS NULL THEN 'missed'
                        WHEN JSON_EXTRACT(ag.day_{day}, '$.submit_status') = 1 THEN 'delayed'
                        ELSE 'ontime'
                    END AS day_status

                FROM employee_header_all eh

                INNER JOIN team_emp_combination_all tec
                    ON tec.eha_id = eh.eha_id
                    AND DATE(tec.linked_on) <= %s
                    AND (
                        tec.unlinked_on IS NULL
                        OR DATE(tec.unlinked_on) > %s
                    )
                    AND (
                        tec.valid_till IS NULL
                        OR DATE(tec.valid_till) >= %s
                    )

                INNER JOIN team_header_all tha
                    ON tha.tha_id = tec.tha_id

                LEFT JOIN member_daily_report_details_all ag
                    ON ag.tec_id = tec.tec_id
                    AND ag.mm_yy = %s

                WHERE eh.status = 1
                  AND eh.eha_id = %s

            """, (day_date, day_date, day_date, mm_yy, empid))

            day_rows = cursor.fetchall()

            # no active teams on this day
            if not day_rows:
                result.append({
                    'date':     date_str,
                    'status':   'grey',
                    'day_type': day_type
                })
                continue

            missed  = any(r['day_status'] == 'missed'  for r in day_rows)
            delayed = any(r['day_status'] == 'delayed' for r in day_rows)

            if missed:
                status = 'red'
            elif delayed:
                status = 'yellow'
            else:
                status = 'green'

            result.append({
                'date':     date_str,
                'status':   status,
                'day_type': day_type
            })

        return jsonify(result), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()




#REPORT
def resolve_report_window(empid, target_date, cursor):
    cursor.execute("""
        SELECT 
            TIME_FORMAT(daily_report_start, '%%H:%%i') AS daily_report_start,
            TIME_FORMAT(daily_report_end, '%%H:%%i') AS daily_report_end,
            submit_next_day
        FROM hr_general_settings_all
        WHERE valid_from <= %s
        AND (valid_till IS NULL OR valid_till >= %s)
        ORDER BY valid_from DESC
        LIMIT 1
    """, (target_date, target_date))
    gs =  cursor.fetchone()
    if not gs:
        return None, None,None,0

    window_start = gs.get("daily_report_start")
    window_end = gs.get("daily_report_end")
    window_submit_nextday = 'yes' if gs.get("submit_next_day") == 1 else 'no'
    
    
    
    
    cursor.execute("""
        SELECT 
            TIME_FORMAT(early_report_submit, '%%H:%%i') AS early_report_submit,
            TIME_FORMAT(late_report_submit,  '%%H:%%i') AS late_report_submit,
            submit_next_day
        FROM hr_special_approvals_settings_all
        WHERE eha_id = %s
        AND valid_from <= %s
        AND (valid_till IS NULL OR valid_till >= %s)
        ORDER BY valid_from DESC
        LIMIT 1
    """, (empid, target_date, target_date))
    hsa = cursor.fetchone()
    print(hsa)
    if hsa:
        window_start = hsa.get("early_report_submit")
        window_end = hsa.get("late_report_submit")
        window_submit_nextdaysp = hsa.get("submit_next_day")
    
        return window_start, window_end ,window_submit_nextdaysp,1
    
    
    return window_start, window_end ,window_submit_nextday,0

@app.route("/report_window_status", methods=["GET"])
def get_report_window():
    
    empid = session.get("empid")
    print(empid)
    if not empid:
        return jsonify({"error": "Not logged in"}), 401
    now   = datetime.now()
    today = now.date()

    conn   = get_db()
    cursor = conn.cursor()

    try:
        window_start, window_end,submit_next_day,_ = resolve_report_window(empid, today, cursor)
        if window_start is None:
            return jsonify({"error": "No active settings found"}), 500

        window_start = to_time(window_start)
        window_end   = to_time(window_end)
        if window_start is None or window_end is None:
            return jsonify({"error": "No active settings found"}), 500
        
        window_open  = datetime.combine(today, window_start)

        if submit_next_day == "yes":
            window_close = datetime.combine(today +timedelta(days=1),window_end)
        else:
            window_close = datetime.combine(today, window_end)

        
        is_open = window_open <= now <= window_close

        return jsonify({
            "is_open":      is_open,
            "window_start": window_start.strftime("%H:%M"),  # 24hr
            "window_end":   window_end.strftime("%H:%M"),     # 24hr
            "submit_next_day": bool(submit_next_day)
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())  # full error in terminal
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/daily_report/team_wise', methods=['GET'])
def get_daily_report_team_wise():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Missing date"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn   = get_db()
    cursor = conn.cursor()

    try:

        cursor.execute(f"""
            SELECT
                tec.tec_id,
                tha.team_name,
                rep.{day_col} AS report_data

            FROM team_emp_combination_all tec

            INNER JOIN team_header_all tha
                ON tha.tha_id = tec.tha_id

            LEFT JOIN member_daily_report_details_all rep
                ON rep.tec_id = tec.tec_id
                AND rep.mm_yy = %s

            WHERE tec.eha_id = %s
              AND DATE(tec.linked_on) <= %s
              AND (
                  tec.unlinked_on IS NULL
                  OR DATE(tec.unlinked_on) > %s
              )
              AND (
                  tec.valid_till IS NULL
                  OR DATE(tec.valid_till) >= %s
              )

        """, (
            mm_yy,
            empid,
            target_date,
            target_date,
            target_date
        ))

        rows = cursor.fetchall()

        if not rows:
            return jsonify([]), 200

        result = []

        for row in rows:
            raw = row['report_data']

            if raw:
                payload = json.loads(raw) if isinstance(raw, str) else raw

                result.append({
                    "team_name":     row['team_name'],
                    "tec_id":        row['tec_id'],
                    "submitted":     True,
                    "out_time":      payload.get("out_time"),
                    "submit_status": payload.get("submit_status", 0),
                    "tasks":          payload.get("text", [])
                })

            else:
                result.append({
                    "team_name": row['team_name'],
                    "tec_id":    row['tec_id'],
                    "submitted": False
                })
        print(result)

        return jsonify(result), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()



@app.route("/submit_next_day_status", methods=["GET"])
def submit_next_day_status():
    empid = session.get("empid")
    today = datetime.today().date()

    conn   = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT submit_next_day
            FROM hr_special_approvals_settings_all
            WHERE eha_id    = %s
              AND valid_from <= %s
              AND (valid_till IS NULL OR valid_till >= %s)
            ORDER BY valid_from DESC
            LIMIT 1
        """, (empid, today, today))
        special = cursor.fetchone()

        if special:
            return jsonify({
                "submit_next_day": special["submit_next_day"],
                "source": "special"
            }), 200

        cursor.execute("""
            SELECT submit_next_day
            FROM hr_general_settings_all
            WHERE valid_from <= %s
              AND (valid_till IS NULL OR valid_till >= %s)
            ORDER BY valid_from DESC
            LIMIT 1
        """, (today, today))
        general = cursor.fetchone()
        print(general)

        if general:
            return jsonify({
                "submit_next_day": general["submit_next_day"],
                "source": "general"
            }), 200

        return jsonify({"submit_next_day": None, "source": None}), 200

    except Exception as e:
        print("submit_next_day_status error:", e)
        return jsonify({"message": "Internal server error"}), 500

    finally:
        cursor.close()
        conn.close()


@app.route("/add_report", methods=["POST"])
def add_daily_report():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    data = request.get_json()
    report_date_str = data.get("date")
    tasks = data.get("tasks", [])
    tec_id = data.get("tec_id")
    now = datetime.now()

    if not report_date_str or not tasks:
        return jsonify({"error": "Missing date or tasks"}), 400

    try:
        report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    conn = get_db()
    cursor = conn.cursor()

    try:
        # 1. Resolve reporting window
        window_start, window_end, submit_next_day, is_special = resolve_report_window(empid, report_date, cursor)
        if window_start is None:
            return jsonify({"error": "No active report settings found"}), 500

        start_time = to_time(window_start)
        end_time = to_time(window_end)

        window_open = datetime.combine(report_date, start_time)
        if now < window_open:
            return jsonify({"error": "Window not open yet"}), 400

        if submit_next_day == 'yes':
            window_close = datetime.combine(report_date + timedelta(days=1), end_time)
        else:
            window_close = datetime.combine(report_date, end_time)

        is_open = window_open <= now <= window_close
        submit_status = 0 if is_open else 1

        # 2. Team assignment
        if tec_id:
            cursor.execute("""
                SELECT tec_id, tha_id
                FROM team_emp_combination_all
                WHERE eha_id = %s
                  AND tec_id = %s
                  AND unlinked_on IS NULL
                  AND (valid_till IS NULL OR valid_till >= CURDATE())
                LIMIT 1
            """, (empid, tec_id))
        else:
            cursor.execute("""
                SELECT tec_id, tha_id
                FROM team_emp_combination_all
                WHERE eha_id = %s
                  AND unlinked_on IS NULL
                  AND (valid_till IS NULL OR valid_till >= CURDATE())
                ORDER BY id DESC
                LIMIT 1
            """, (empid,))

        assign = cursor.fetchone()
        if not assign:
            return jsonify({"error": "No active team assignment found"}), 400

        tec_id = assign["tec_id"]
        tha_id = assign["tha_id"]

        mm_yy = report_date.strftime("%m-%Y")
        day_num = report_date.day
        day_col = f"day_{day_num}"

        # 3. Check existing report
        cursor.execute(f"""
            SELECT id, {day_col} AS existing_data
            FROM member_daily_report_details_all
            WHERE eha_id = %s AND mm_yy = %s AND tec_id = %s
        """, (empid, mm_yy, tec_id))
        existing_row = cursor.fetchone()

        # 4. Verify agenda exists
        cursor.execute(f"""
            SELECT JSON_UNQUOTE(JSON_EXTRACT({day_col}, '$.wa_id')) AS wa_id
            FROM member_daily_agenda_details_all
            WHERE eha_id = %s AND mm_yy = %s AND tec_id = %s
        """, (empid, mm_yy, tec_id))
        row = cursor.fetchone()
        if not row or row['wa_id'] is None:
            return jsonify({"error": "Please submit your daily agenda before reporting"}), 400
        wa_id = row['wa_id']

        # 5. Build payload
        payload = {
            "out_time": now.strftime("%Y-%m-%d %H:%M:%S"),
            "text": tasks,
            "submit_status": submit_status,
            "submitted_at": now.isoformat(),
            "submit_next_day": submit_next_day,
            "wa_id": wa_id
        }
        payload_json = json.dumps(payload, ensure_ascii=False)

        # 6. Determine if the submitted date is yesterday
        today = datetime.now().date()
        is_yesterday = (report_date == today - timedelta(days=1))

        # 7. Window enforcement: only block non‑yesterday dates when closed
        if not is_yesterday and not is_open:
            return jsonify({"error": "Window closed – cannot add/modify report"}), 403

        # 8. Insert or update
        if existing_row:
            cursor.execute(f"""
                UPDATE member_daily_report_details_all
                SET {day_col} = %s
                WHERE eha_id = %s AND mm_yy = %s AND tec_id = %s
            """, (payload_json, empid, mm_yy, tec_id))
        else:
            # Atomic ID generation
            cursor.execute("""
                INSERT INTO unique_header_all (table_name, prefix, last_id)
                VALUES ('member_daily_report_details_all', 'MDR', LAST_INSERT_ID(1))
                ON DUPLICATE KEY UPDATE last_id = LAST_INSERT_ID(last_id + 1)
            """)
            cursor.execute("SELECT LAST_INSERT_ID() AS new_id")
            new_id = cursor.fetchone()['new_id']
            mdr_id = f"MDR-{new_id:05d}"

            cursor.execute(f"""
                INSERT INTO member_daily_report_details_all
                (eha_id, mdr_id, tha_id, tec_id, mm_yy, {day_col})
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (empid, mdr_id, tha_id, tec_id, mm_yy, payload_json))

        conn.commit()
        return jsonify({
            "message": "Report submitted successfully",
            "submit_status": "on-time" if submit_status == 0 else "delayed"
        }), 200

    except Exception:
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Internal server error"}), 500

    finally:
        cursor.close()
        conn.close()


@app.route("/daily_report/get", methods=["GET"])
def get_daily_report():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Missing date"}), 400
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    mm_yy = report_date.strftime("%m-%Y")
    day_col = f"day_{report_date.day}"
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT 
            {day_col} as report_data
        FROM member_daily_report_details_all 
        WHERE eha_id = %s AND mm_yy = %s"""   
        ,(empid, mm_yy))
        row = cursor.fetchone()

        if not row or not row["report_data"]:
            return jsonify({"submitted": False, "message": "No report for this date"}), 200
        payload = json.loads(row["report_data"])
        
        # payload structure: {"out_time": "...", "tasks": [...], "submit_status": 0/1}
        return jsonify({
            "submitted": True,
            "out_time": payload.get("out_time"),
            "tasks": payload.get("tasks", []),
            "submit_status": payload.get("submit_status")
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/user_tec_combinations', methods=['GET'])
def get_user_tec_combinations():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    conn = get_db()
    cursor = conn.cursor()
    try:
        query = """
            SELECT
                tec.tec_id,
                tha.team_name,
                CONCAT_WS(' ', eha.first_name, eha.middle_name, eha.last_name) AS employee_name
            FROM team_emp_combination_all tec
            INNER JOIN team_header_all tha ON tec.tha_id = tha.tha_id
            INNER JOIN employee_header_all eha ON tec.eha_id = eha.eha_id
            WHERE tec.eha_id = %s
              AND tec.unlinked_on IS NULL
              AND (tec.valid_till IS NULL OR tec.valid_till >= CURDATE())
        """
        cursor.execute(query, (empid,))
        rows = cursor.fetchall()

        combinations = []
        for row in rows:
            combinations.append({
                "tec_id": row["tec_id"],
                "team_name": row["team_name"],
                "employee_name": row["employee_name"]
            })

        return jsonify(combinations), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/')


#PERFORMANCE REPORT///////////////////////////////////////////////////////////////////////////////////////////////////////



#EMPLOYEE

def resolve_performance_window(target_date, cursor):
    cursor.execute("""
        SELECT 
            perform_verify_start AS Start,
            perform_verify_end AS End
        FROM hr_general_settings_all
        WHERE valid_from <= %s
        AND (valid_till IS NULL OR valid_till >= %s)
        ORDER BY valid_from DESC
        LIMIT 1
    """, (target_date, target_date))
    gs =  cursor.fetchone()
    if not gs:
        return None, None

    window_start = gs.get("Start")
    window_end = gs.get("End")
    print(window_start,window_end)
     
    return window_start, window_end 

@app.route("/window_status_performance", methods=["GET"])
def get_performance_window():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401
    print(empid)

    now   = datetime.now()
    today = now.date()

    conn   = get_db()
    cursor = conn.cursor()

    try:
        now = datetime.now()
        formatted_date = now.strftime("%Y-%m-%d")
        window_start,window_end = resolve_performance_window(formatted_date,cursor)
        current_day = now.day
        is_open = 0
        if window_start <= window_end:
            # Normal window (e.g., 2 → 5)
            if window_start <= current_day <= window_end:
                is_open = 1
        else:
            # Wrap‑around window (e.g., 30 → 5)
            if current_day >= window_start or current_day <= window_end:
                is_open = 1
        print("ASJFBAKJbg")
        return jsonify({
            "is_open":      is_open,
            "window_start": window_start,  # 24hr
            "window_end":   window_end   # 24hr
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())  # full error in terminal
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@app.route("/performance_report/submit", methods=["POST"])
def submit_performance_report():
    conn = None
    try:
        if "empid" not in session:
            return jsonify({"error": "Not logged in"}), 401

        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400
        

        

        eha_id = session["empid"]

        self_rating = data.get("self_rating")
        skill_development = data.get("skill_development")
        major_tasks = data.get("major_tasks", [])
        challenges = data.get("challenges", [])
        goals = data.get("goals", [])

        conn = get_db()
        cursor = conn.cursor()

        # ----- 2. Fetch employee name -----
        cursor.execute(
            "SELECT first_name, middle_name, last_name FROM employee_header_all WHERE eha_id = %s",
            (eha_id,)
        )
        emp = cursor.fetchone()
        if not emp:
            return jsonify({"error": "Employee not found"}), 404

        first = emp["first_name"]
        middle = emp["middle_name"]
        last = emp["last_name"]
        employee_name = f"{first} {middle} {last}".strip() if middle else f"{first} {last}"

        # ----- 3. Auto‑fill dates -----
        now = datetime.now()
        review_month = now.strftime("%m/%y")   # MM/YY format (e.g. 05/26)
        review_date = now.strftime("%Y-%m-%d") # e.g. 2026-05-12

        formatted_date = now.strftime("%Y-%m-%d")
        window_start, window_end = resolve_performance_window(formatted_date, cursor)

        current_day = now.day

        is_open = False
        is_delayed = False

        if window_start <= window_end:
            # Normal window
            if window_start <= current_day <= window_end:
                is_open = True
            elif current_day > window_end:
                is_open = True
                is_delayed = True
        else:
            # Wrap-around window
            if current_day >= window_start or current_day <= window_end:
                is_open = True

        # Block before-window submission
        if not is_open:
            return jsonify({
                "error": f"Performance review window opens on day {window_start}"
            }), 400

        submit_status = 1 if is_delayed else 0

        # ----- 4. Serialize array fields (correctly) -----
        
        major_tasks_json = json.dumps(major_tasks, default=str)
        challenges_json = json.dumps(challenges, default=str)
        goals_json = json.dumps(goals, default=str)


        # Check if already submitted for this month
        cursor.execute("""
            SELECT 1
            FROM performance_report_details_all
            WHERE eha_id = %s
            AND review_month = %s
            LIMIT 1
        """, (eha_id, review_month))

        existing_report = cursor.fetchone()

        if existing_report:
            return jsonify({
                "error": "Performance report already submitted for this month"
            }), 400

        # ----- 5. Current team info (project, reporting manager, assigned projects) -----
        cursor.execute("""
            SELECT 
                t.tha_id,
                t.team_name,
                t.team_members,
                t.team_leader,
                tec.linked_on,
                tec.link_type,
                tec.designation
            FROM team_emp_combination_all tec
            JOIN team_header_all t ON tec.tha_id = t.tha_id
            WHERE tec.eha_id = %s
              AND tec.unlinked_on IS NULL
              AND (tec.valid_till IS NULL OR tec.valid_till >= NOW())
              AND t.status = 1
            ORDER BY tec.linked_on DESC
        """, (eha_id,))
        teams = cursor.fetchall()

        assigned_projects = []
        reporting_manager = json.dumps([]) 
        designation = ""                     

        if not teams:
            project = "N/A"
            reporting_manager = "N/A"
            assigned_projects = []
        else:
            # Primary team
            primary_team = teams[0]
            project = primary_team["team_name"]

            # ---------- REPORTING MANAGERS ----------
            all_leader_ids = set()
            designations = []
            for team in teams:
                if team.get("designation"):
                    designations.append({
                        "teamName": team["team_name"],
                        "designation": team["designation"]
                    })

                leader_data = []
                if team["team_leader"]:
                    try:
                        leader_data = json.loads(team["team_leader"])
                    except:
                        leader_data = []
                for item in leader_data:
                    if isinstance(item, dict):
                        lid = item.get("eha_id") or item.get("id")
                        if lid:
                            all_leader_ids.add(str(lid).strip())
                    else:
                        if item:
                            all_leader_ids.add(str(item).strip())

            leader_ids = list(all_leader_ids)
            leader_map = {}
            if leader_ids:
                format_strings = ",".join(["%s"] * len(leader_ids))
                cursor.execute(
                    f"""SELECT eha_id,
                            CONCAT_WS(' ', first_name, middle_name, last_name) AS full_name
                    FROM employee_header_all
                    WHERE eha_id IN ({format_strings})""",
                    tuple(leader_ids)
                )
                leaders = cursor.fetchall()
                leader_map = {
                    str(row["eha_id"]).strip(): row["full_name"].strip().replace("  ", " ")
                    for row in leaders
                }

            reporting_manager = json.dumps([
                {"eha_id": lid, "name": leader_map.get(lid, lid)}
                for lid in leader_ids
            ]) if leader_ids else json.dumps([])

            # Optional: set designation to the whole list if your schema requires it
            designation = json.dumps(designations)   # or "" if not needed

            # ---------- ASSIGNED PROJECTS ----------
            assigned_projects = []   # clear the pre-initialized list (or start fresh)
            for team in teams:
                # parse members
                member_ids_raw = []
                if team["team_members"]:
                    try:
                        member_data = json.loads(team["team_members"])
                    except:
                        member_data = []
                    for item in member_data:
                        if isinstance(item, dict):
                            mid = item.get("eha_id") or item.get("id")
                            if mid:
                                member_ids_raw.append(str(mid).strip())
                        else:
                            if item:
                                member_ids_raw.append(str(item).strip())

                # parse leaders
                leader_ids_raw = []
                if team["team_leader"]:
                    try:
                        leader_data = json.loads(team["team_leader"])
                    except:
                        leader_data = []
                    for item in leader_data:
                        if isinstance(item, dict):
                            lid = item.get("eha_id") or item.get("id")
                            if lid:
                                leader_ids_raw.append(str(lid).strip())
                        else:
                            if item:
                                leader_ids_raw.append(str(item).strip())

                # merge & remove self
                all_ids = set(member_ids_raw + leader_ids_raw)
                other_ids = [uid for uid in all_ids if uid != str(eha_id).strip()]

                worked_with = ""
                if other_ids:
                    fmt = ",".join(["%s"] * len(other_ids))
                    cursor.execute(
                        f"""SELECT eha_id,
                                CONCAT_WS(' ', first_name, middle_name, last_name) AS full_name
                        FROM employee_header_all
                        WHERE eha_id IN ({fmt})""",
                        tuple(other_ids)
                    )
                    others = cursor.fetchall()
                    worked_with = ", ".join(
                        row["full_name"].strip().replace("  ", " ") for row in others
                    )

                assigned_projects.append({
                    "projectName": team["team_name"],
                    "workedAs": team.get("designation") or "",
                    "workedWith": worked_with,
                    "workingFrom": (
                        team["linked_on"].strftime("%d/%m/%Y") if team["linked_on"] else ""
                    )
                })

        assigned_projects_json = json.dumps(assigned_projects)

        # ----- 6. Generate unique PR‑ID -----
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'performance_reports'")
        header = cursor.fetchone()

        if not header or header["last_id"] is None:
            cursor.execute(
                "INSERT INTO unique_header_all (table_name, prefix, last_id) VALUES ('performance_reports', 'PR', '0')"
            )
            conn.commit()
            new_last_id = 1
            pr_id = "PR-00001"
        else:
            new_last_id = int(header["last_id"]) + 1
            pr_id = f"PR-{new_last_id:05d}"
            cursor.execute(
                "UPDATE unique_header_all SET last_id = %s, modified_on = NOW() WHERE table_name = 'performance_reports'",
                (str(new_last_id),)
            )
            conn.commit()

        

        sql = """INSERT INTO performance_report_details_all 
            (pr_id, eha_id, employee_name, project,
             reporting_manager, designation, review_month,
             review_date, self_rating, skill_development,
             major_tasks, challenges, assigned_projects, goals, submit_status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

        values = (
            pr_id, eha_id, employee_name, project, reporting_manager,
            designation, review_month, review_date, self_rating,
            skill_development, major_tasks_json, challenges_json,
            assigned_projects_json, goals_json, submit_status
        )

        cursor.execute(sql, values)
        conn.commit()                    # ← note the parentheses

        return jsonify({"message": "Performance report submitted successfully"
                        , "pr_id": pr_id,
                        "submit_status":submit_status}), 200

    except Exception as e:
        print("Error submitting performance report:", e)
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": f"Error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()

@app.route('/performance_report/prefill', methods=['GET'])
def get_performance_prefill():

    conn = None

    try:

        if "empid" not in session:
            return jsonify({
                "error": "Not logged in"
            }), 401

        eha_id = session["empid"]

        conn = get_db()
        cursor = conn.cursor()

        # =====================================================
        # EMPLOYEE DETAILS
        # =====================================================

        cursor.execute("""
            SELECT
                first_name,
                middle_name,
                last_name
            FROM employee_header_all
            WHERE eha_id = %s
        """, (eha_id,))

        emp = cursor.fetchone()

        if not emp:
            return jsonify({
                "error": "Employee not found"
            }), 404

        first = emp["first_name"]
        middle = emp.get("middle_name")
        last = emp["last_name"]

        employee_name = (
            f"{first} {middle} {last}".strip()
            if middle
            else f"{first} {last}"
        )

        now = datetime.now()

        review_month = now.strftime("%m/%y")

        review_date = now.strftime("%d/%m/%Y")

        # =====================================================
        # ACTIVE TEAMS
        # =====================================================

        cursor.execute("""
            SELECT
                t.tha_id,
                t.team_name,
                t.team_members,
                t.team_leader,

                tec.linked_on,
                tec.link_type,
                tec.designation

            FROM team_emp_combination_all tec

            JOIN team_header_all t
                ON tec.tha_id = t.tha_id

            WHERE tec.eha_id = %s
              AND tec.unlinked_on IS NULL
              AND (
                    tec.valid_till IS NULL
                    OR tec.valid_till >= NOW()
                )
              AND t.status = 1

            ORDER BY tec.linked_on DESC
        """, (eha_id,))

        teams = cursor.fetchall()

        if not teams:

            return jsonify({
                "employee_name": employee_name,
                "project": "N/A",
                "reporting_manager": "N/A",
                "review_month": review_month,
                "review_date": review_date,
                "assigned_projects": []
            }), 200

        # =====================================================
        # PRIMARY TEAM
        # =====================================================

        primary_team = teams[0]

        project = primary_team["team_name"]

        # =====================================================
        # REPORTING MANAGERS
        # =====================================================

        all_leader_ids = set()

        for team in teams:

            leader_data = []

            if team["team_leader"]:

                try:
                    leader_data = json.loads(
                        team["team_leader"]
                    )
                except:
                    leader_data = []

            for item in leader_data:

                # Object format
                if isinstance(item, dict):

                    leader_id = (
                        item.get("eha_id")
                        or item.get("id")
                    )

                    if leader_id:

                        all_leader_ids.add(
                            str(leader_id).strip()
                        )

                # Direct string format
                else:

                    if item:

                        all_leader_ids.add(
                            str(item).strip()
                        )

        leader_ids = list(all_leader_ids)

        manager_names = []

        if leader_ids:

            format_strings = ",".join(
                ["%s"] * len(leader_ids)
            )

            cursor.execute(
                f"""
                SELECT
                    eha_id,

                    CONCAT_WS(
                        ' ',
                        first_name,
                        middle_name,
                        last_name
                    ) AS full_name

                FROM employee_header_all

                WHERE eha_id IN ({format_strings})
                """,
                tuple(leader_ids)
            )

            leaders = cursor.fetchall()

            leader_map = {

                str(row["eha_id"]).strip():
                row["full_name"].strip().replace("  ", " ")

                for row in leaders
            }

            manager_names = [

                leader_map.get(lid, lid)

                for lid in leader_ids

            ]

        reporting_manager = (

            ", ".join(manager_names)

            if manager_names

            else "N/A"

        )

        # =====================================================
        # ASSIGNED PROJECTS
        # =====================================================

        assigned_projects = []

        for team in teams:

            # ---------------------------------
            # TEAM MEMBERS
            # ---------------------------------

            member_data = []

            if team["team_members"]:

                try:
                    member_data = json.loads(
                        team["team_members"]
                    )
                except:
                    member_data = []

            member_ids = []

            for item in member_data:

                # Object format
                if isinstance(item, dict):

                    member_id = (
                        item.get("eha_id")
                        or item.get("id")
                    )

                    if member_id:

                        member_ids.append(
                            str(member_id).strip()
                        )

                # Direct string format
                else:

                    if item:

                        member_ids.append(
                            str(item).strip()
                        )

            # ---------------------------------
            # TEAM LEADERS
            # ---------------------------------

            team_leader_data = []

            if team["team_leader"]:

                try:
                    team_leader_data = json.loads(
                        team["team_leader"]
                    )
                except:
                    team_leader_data = []

            team_leader_ids = []

            for item in team_leader_data:

                # Object format
                if isinstance(item, dict):

                    leader_id = (
                        item.get("eha_id")
                        or item.get("id")
                    )

                    if leader_id:

                        team_leader_ids.append(
                            str(leader_id).strip()
                        )

                # Direct string format
                else:

                    if item:

                        team_leader_ids.append(
                            str(item).strip()
                        )

            # ---------------------------------
            # MERGE MEMBERS + LEADERS
            # ---------------------------------

            combined_ids = list(
                set(member_ids + team_leader_ids)
            )

            # remove self
            other_member_ids = [

                mid

                for mid in combined_ids

                if mid != str(eha_id).strip()

            ]

            worked_with = ""

            if other_member_ids:

                format_strings = ",".join(
                    ["%s"] * len(other_member_ids)
                )

                cursor.execute(
                    f"""
                    SELECT
                        eha_id,

                        CONCAT_WS(
                            ' ',
                            first_name,
                            middle_name,
                            last_name
                        ) AS full_name

                    FROM employee_header_all

                    WHERE eha_id IN ({format_strings})
                    """,
                    tuple(other_member_ids)
                )

                others = cursor.fetchall()

                other_names = [

                    row["full_name"]
                    .strip()
                    .replace("  ", " ")

                    for row in others

                ]

                worked_with = ", ".join(other_names)

            assigned_projects.append({

                "projectName": team["team_name"],

                "workedAs": team["designation"] or "",

                "workedWith": worked_with,

                "workingFrom": (
                    team["linked_on"].strftime("%d/%m/%Y")
                    if team["linked_on"]
                    else ""
                )

            })

        # =====================================================
        # RESPONSE
        # =====================================================

        return jsonify({

            "employee_name": employee_name,

            "project": project,

            "reporting_manager": reporting_manager,

            "review_month": review_month,

            "review_date": review_date,

            "assigned_projects": assigned_projects

        }), 200

    except Exception as e:

        import traceback
        traceback.print_exc()

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        if conn:
            conn.close()

@app.route("/performance_report/submission_status", methods=["GET"])
def performance_report_submission_status():
    conn = None
    try:
        if "empid" not in session:
            return jsonify({"error": "Not logged in"}), 401

        eha_id = session["empid"]

        conn = get_db()
        cursor = conn.cursor()

        # Current month in same format used during submit
        review_month = datetime.now().strftime("%m/%y")

        cursor.execute("""
            SELECT 
                pr_id,
                review_month,
                review_date,
                submit_status
            FROM performance_report_details_all
            WHERE eha_id = %s
              AND review_month = %s
            LIMIT 1
        """, (eha_id, review_month))

        report = cursor.fetchone()

        if report:
            return jsonify({
                "submitted": True,
                "pr_id": report["pr_id"],
                "review_month": report["review_month"],
                "review_date": str(report["review_date"]),
                "submit_status": report["submit_status"]
            }), 200

        return jsonify({
            "submitted": False
        }), 200

    except Exception as e:
        print("Error fetching performance report status:", e)
        return jsonify({
            "error": str(e)
        }), 500

    finally:
        if conn:
            conn.close()

@app.route('/performance_review/list', methods=['GET'])
def performance_review_list():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    month = request.args.get("month")  # MM/YY
    status = request.args.get("status", "pending")
    if not month:
        month = datetime.today().strftime("%m/%y")

    conn = get_db()
    cursor = conn.cursor()

    try:
        # ---------------------------------------------------------
        # 1. Team members (used for pending & reviewed)
        # ---------------------------------------------------------
        cursor.execute("""
            SELECT DISTINCT pr.eha_id AS member_eha_id
            FROM performance_report_details_all pr
            WHERE pr.reporting_manager LIKE %s
        """, (f"%{empid}%",))
        team_members = [r['member_eha_id'] for r in cursor.fetchall()]

        if not team_members:
            return jsonify({"employees": [], "review_deadline_days": 0}), 200

        # ---------------------------------------------------------
        # 2. Deadline days from general settings
        # ---------------------------------------------------------
        cursor.execute("""
            SELECT perform_verification
            FROM hr_general_settings_all
            WHERE valid_till IS NULL
            ORDER BY valid_from DESC LIMIT 1
        """)
        setting = cursor.fetchone()
        if not setting or setting['perform_verification'] is None:
            raise Exception("Performance verification days not configured")
        deadline_days = int(setting['perform_verification'])

        # ---------------------------------------------------------
        # 3. PENDING
        # ---------------------------------------------------------
        if status == "pending":
            cursor.execute("""
                SELECT
                    pr.pr_id,
                    pr.eha_id,
                    pr.employee_name,
                    pr.review_month,
                    pr.review_date,
                    pr.submit_status,
                    pr.created_at
                FROM performance_report_details_all pr
                LEFT JOIN performance_points_details_all pp
                    ON pp.eha_id = pr.eha_id AND pp.mm_yy = pr.review_month
                WHERE pr.reporting_manager LIKE %s
                  AND pr.review_month = %s
                  AND pr.eha_id != %s
                  AND pp.tl_submit_time IS NULL
                ORDER BY pr.created_at ASC
            """, (f"%{empid}%", month, empid))

        # ---------------------------------------------------------
        # 4. REVIEWED – NO DUPLICATES, INCLUDES pr_id
        # ---------------------------------------------------------
        elif status == "reviewed":
            cursor.execute("""
                SELECT
                    pr.pr_id,
                    pp.eha_id,
                    CONCAT_WS(' ', eh.first_name, eh.middle_name, eh.last_name) AS employee_name,
                    pp.mm_yy,
                    pp.tl_point_1, pp.tl_point_2, pp.tl_point_3, pp.tl_point_4,
                    pp.tl_point_5, pp.tl_point_6, pp.tl_point_7, pp.tl_point_8, pp.tl_point_9,
                    pp.tl_comments,
                    pp.tl_submit_time,
                    pp.tl_submit_status,
                    team.team_name,
                    pr.submit_status
                FROM performance_points_details_all pp
                INNER JOIN performance_report_details_all pr
                    ON pr.eha_id = pp.eha_id AND pr.review_month = pp.mm_yy
                INNER JOIN employee_header_all eh
                    ON eh.eha_id = pp.eha_id
                INNER JOIN (
                    -- Subquery: one active team per employee (most recent by `id`)
                    SELECT
                        tec.eha_id,
                        th.team_name
                    FROM team_emp_combination_all tec
                    INNER JOIN team_header_all th
                        ON th.tha_id = tec.tha_id
                        AND th.team_leader LIKE %s
                    WHERE tec.unlinked_on IS NULL
                    AND (tec.valid_till IS NULL
                        OR tec.valid_till >= STR_TO_DATE(CONCAT('01/', %s), '%%d/%%m/%%y'))
                    ORDER BY tec.id DESC   -- most recent assignment first
                ) AS team ON team.eha_id = pp.eha_id
                GROUP BY pp.eha_id   -- ensures one row per employee (first from ORDER BY)
                HAVING pp.mm_yy = %s
                AND pp.eha_id != %s
                ORDER BY pp.tl_submit_time DESC
            """, (
                f"%{empid}%",   # for team_leader LIKE
                month,          # for valid_till
                month,          # for pp.mm_yy
                empid           # exclude self
            ))

        # ---------------------------------------------------------
        # 5. MISSED
        # ---------------------------------------------------------
        else:  # missed
            cursor.execute("""
                SELECT DISTINCT
                    member.eha_id,
                    CONCAT_WS(' ', eh.first_name, eh.middle_name, eh.last_name) AS employee_name
                FROM team_emp_combination_all lead
                INNER JOIN team_emp_combination_all member
                    ON member.tha_id = lead.tha_id
                    AND member.link_type = 1
                    AND member.unlinked_on IS NULL
                    AND (member.valid_till IS NULL
                         OR member.valid_till >= STR_TO_DATE(CONCAT('01/', %s), '%%d/%%m/%%y'))
                INNER JOIN employee_header_all eh ON eh.eha_id = member.eha_id
                LEFT JOIN performance_report_details_all pr
                    ON pr.eha_id = member.eha_id AND pr.review_month = %s
                WHERE lead.eha_id = %s
                  AND lead.link_type = 2
                  AND lead.unlinked_on IS NULL
                  AND (lead.valid_till IS NULL
                       OR lead.valid_till >= STR_TO_DATE(CONCAT('01/', %s), '%%d/%%m/%%y'))
                  AND member.eha_id != %s
                  AND pr.pr_id IS NULL
                ORDER BY employee_name ASC
            """, (month, month, empid, month, empid))

        rows = cursor.fetchall()
        result = []

        for row in rows:
            if status == "missed":
                result.append({
                    "eha_id": row['eha_id'],
                    "employee_name": row['employee_name'],
                })
                continue

            if status == "reviewed":
                # ✅ Include pr_id here
                entry = {
                    "pr_id": row['pr_id'],               # <-- added
                    "eha_id": row['eha_id'],
                    "employee_name": row['employee_name'],
                    "review_month": row['mm_yy'],
                    "submitted_at": str(row['tl_submit_time']) if row['tl_submit_time'] else None,
                    "report_points": row['tl_point_9'],
                    "tl_reviewed_at": str(row['tl_submit_time']) if row['tl_submit_time'] else None,
                    "tl_review_status": row['tl_submit_status'],
                    "team_name": row['team_name'],
                    "submit_status": row['submit_status']  
                }
                result.append(entry)
                continue

            # status == "pending"
            created_at = row['created_at']
            deadline_date = created_at + timedelta(days=deadline_days)
            days_left = (deadline_date.date() - datetime.now().date()).days

            result.append({
                "pr_id": row['pr_id'],
                "eha_id": row['eha_id'],
                "employee_name": row['employee_name'],
                "review_month": row['review_month'],
                "review_date": str(row['review_date']),
                "submit_status": row['submit_status'],
                "submitted_at": str(created_at),
                "days_left": max(days_left, 0),
                "overdue": days_left < 0,
            })

        return jsonify({
            "employees": result,
            "review_deadline_days": deadline_days
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/performance_review/<int:pr_id>', methods=['GET'])
def performance_review_details(pr_id):

    empid = session.get("empid")

    if not empid:
        return jsonify({
            "error": "Not logged in"
        }), 401

    conn = get_db()
    cursor = conn.cursor()

    try:

        # =====================================================
        # FETCH REPORT
        # =====================================================

        cursor.execute("""
            SELECT
                pr.*,

                pp.tl_point_1,
                pp.tl_point_2,
                pp.tl_point_3,
                pp.tl_point_4,
                pp.tl_point_5,
                pp.tl_point_6,
                pp.tl_point_7,
                pp.tl_point_8,
                pp.tl_point_9,

                pp.tl_comments,
                pp.tl_reviewed_by,
                eh.first_name AS tl_reviewed_by_name,

                pp.tl_submit_time,
                pp.tl_submit_status

            FROM performance_report_details_all pr

            LEFT JOIN performance_points_details_all pp
                ON pp.eha_id = pr.eha_id
                AND pp.mm_yy = pr.review_month

            LEFT JOIN employee_header_all eh
                ON eh.eha_id = pp.tl_reviewed_by

            WHERE pr.pr_id = %s

            LIMIT 1
        """, (pr_id,))
            
        row = cursor.fetchone()

        if not row:
            return jsonify({
                "error": "Performance report not found"
            }), 404


        response = dict(row)

        # =====================================================
        # PARSE JSON FIELDS
        # =====================================================

        response["major_tasks"] = (
            json.loads(row["major_tasks"])
            if row["major_tasks"]
            else []
        )
        response["reporting_manager"] = (
            json.loads(row["reporting_manager"])
            if row["reporting_manager"]
            else []
        )
        

        response["assigned_projects"] = (
            json.loads(row["assigned_projects"])
            if row["assigned_projects"]
            else []
        )
        response["tl_comments"] = (
            json.loads(row["tl_comments"])
            if row["tl_comments"]
            else {}
        )

        designations = []
        for proj in response["assigned_projects"]:
            raw = proj.get("workedAs", "[]")
            try:
                worked_as_list = json.loads(raw) if isinstance(raw, str) else raw
            except (json.JSONDecodeError, TypeError):
                worked_as_list = []
            # Find the designation that matches this project's team name
            team_name = proj.get("projectName", "")
            match = next((item for item in worked_as_list if item.get("teamName") == team_name), None)
            designations.append({
                "teamName": team_name,
                "designation": match.get("designation", "") if match else ""
            })
        response["designations"] = designations


        response["challenges"] = (
            json.loads(row["challenges"])
            if row["challenges"]
            else []
        )

        response["goals"] = (
            json.loads(row["goals"])
            if row["goals"]
            else []
        )
        print(response)

        return jsonify(response), 200

    except Exception as e:

        import traceback
        traceback.print_exc()

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        cursor.close()
        conn.close()


@app.route("/performance_review/submit_evaluation", methods=["POST"])
def submit_evaluation():
    conn = None
    try:
        # 1. Auth
        if "empid" not in session:
            return jsonify({"error": "Not logged in"}), 401

        reviewer_id = session["empid"]

        # 2. Parse payload
        data = request.get_json()
        print(data)
        if not data:
            return jsonify({"error": "No data provided"}), 400

        pr_id = data.get("pr_id")
        eha_id = data.get("eha_id")
        mm_yy = data.get("mm_yy")               # e.g. "05/26"
        tl_points = data.get("tl_points")       # list of 8 integers
        tl_comments = data.get("tl_comments")   # dict like {"1": "good", ...}
        tl_avg = data.get("tl_average")             # optional, can be float or None

        # Basic validation
        if not pr_id or not eha_id or not mm_yy:
            return jsonify({"error": "pr_id, eha_id, and mm_yy are required"}), 400

        # 3. Connect DB
        conn = get_db()
        cursor = conn.cursor()

        # 4. Check if row already exists for this employee + month
        cursor.execute("""
            SELECT id, ppd_id FROM performance_points_details_all
            WHERE eha_id = %s AND mm_yy = %s
            LIMIT 1
        """, (eha_id, mm_yy))
        existing = cursor.fetchone()

        # 5. Serialise comments
        comments_json = json.dumps(tl_comments) if tl_comments else json.dumps({})

        if existing:
            # Update existing row
            # Example for UPDATE (assuming you added tl_avg_score column)
            cursor.execute("""
                UPDATE performance_points_details_all
                SET
                    tl_point_1 = %s, tl_point_2 = %s, tl_point_3 = %s,
                    tl_point_4 = %s, tl_point_5 = %s, tl_point_6 = %s,
                    tl_point_7 = %s, tl_point_8 = %s,
                    tl_point_9 = %s,               -- new column
                    tl_comments = %s,
                    tl_reviewed_by = %s,
                    tl_submit_time = NOW(),
                    updated_on = NOW()
                WHERE id = %s
            """, (*tl_points, tl_avg, comments_json, reviewer_id, existing["id"]))
            message = "Evaluation updated successfully"
        else:
            # Insert new row – generate a unique ppd_id
            cursor.execute(
                "SELECT * FROM unique_header_all WHERE table_name = 'performance_points'"
            )
            header = cursor.fetchone()

            if not header or header["last_id"] is None:
                cursor.execute(
                    "INSERT INTO unique_header_all (table_name, prefix, last_id) "
                    "VALUES ('performance_points', 'PPD', '0')"
                )
                conn.commit()
                new_last_id = 1
            else:
                new_last_id = int(header["last_id"]) + 1
                cursor.execute(
                    "UPDATE unique_header_all SET last_id = %s, modified_on = NOW() "
                    "WHERE table_name = 'performance_points'",
                    (str(new_last_id),)
                )
                conn.commit()

            ppd_id = f"PPD-{new_last_id:05d}"

            cursor.execute("""
                INSERT INTO performance_points_details_all
                    (ppd_id, eha_id, mm_yy,
                     tl_point_1, tl_point_2, tl_point_3,
                     tl_point_4, tl_point_5, tl_point_6,
                     tl_point_7, tl_point_8, tl_point_9,
                     tl_comments, tl_reviewed_by,
                     tl_submit_time, tl_submit_status)
                VALUES (%s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        NOW(), 0)
            """, (
                ppd_id, eha_id, mm_yy,
                *tl_points,tl_avg,
                comments_json, reviewer_id
            ))
            message = "Evaluation submitted successfully"

        conn.commit()
        return jsonify({
            "message": message,
            "ppd_id": existing["ppd_id"] if existing else ppd_id
        }), 200

    except Exception as e:
        print("Error submitting evaluation:", e)
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": f"Error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()

#HR
@app.route("/performance_review/hr_list", methods=["GET"])
def hr_performance_review_list():
    try:
        print("hello")
        # 1. Auth – only logged‑in users (later you can add HR role check)
        if "empid" not in session:
            return jsonify({"error": "Not logged in"}), 401

        # 2. Read query params
        month = request.args.get("month")            # "05/26"
        status = request.args.get("status")          # "reviewed", "missed_leader", "missed_employee"

        if not month or not status:
            return jsonify({"error": "month and status are required"}), 400

        conn = get_db()
        cursor = conn.cursor()

        # 3. Build query based on status
        if status == "reviewed":
            # TL‑reviewed reports
            cursor.execute("""
                SELECT
                    pr.pr_id,
                    pr.eha_id,
                    pr.employee_name,
                    pr.review_month,
                    pr.created_at AS submitted_at,
                    pr.submit_status,

                    pp.tl_submit_time AS tl_reviewed_at,
                    pp.tl_submit_status,
                    pp.tl_point_9 AS avg_points,

                    pp.hr_submit_time AS hr_reviewed_at,

                    eh.first_name AS reviewed_by

                FROM performance_report_details_all pr

                INNER JOIN performance_points_details_all pp
                    ON pp.eha_id = pr.eha_id
                    AND pp.mm_yy = pr.review_month

                LEFT JOIN employee_header_all eh
                    ON eh.eha_id = pp.tl_reviewed_by

                WHERE pr.review_month = %s
                AND pp.tl_submit_time IS NOT NULL

                ORDER BY pr.created_at DESC
            """, (month,))

        elif status == "missed_leader":
            # Reports submitted by employee, but TL hasn't reviewed yet

            cursor.execute("""
                SELECT
                    pr.pr_id,
                    pr.eha_id,
                    pr.employee_name,
                    pr.review_month,
                    pr.created_at AS submitted_at,
                    pr.submit_status,

                    hgs.perform_verification

                FROM performance_report_details_all pr

                LEFT JOIN performance_points_details_all pp
                    ON pp.eha_id = pr.eha_id
                    AND pp.mm_yy = pr.review_month

                CROSS JOIN hr_general_settings_all hgs

                WHERE pr.review_month = %s
                AND pp.tl_submit_time IS NULL
                AND hgs.valid_till IS NULL

                ORDER BY pr.created_at DESC
            """, (month,))

        elif status == "missed_employee":
            # Employees who never submitted for the given month
            # (Assuming all active employees are eligible; adjust if HR sees only specific ones)
            cursor.execute("""
                SELECT
                    e.eha_id,
                    CONCAT_WS(' ', e.first_name, e.middle_name, e.last_name) AS employee_name
                FROM employee_header_all e
                WHERE e.status = 1              -- active employees only
                  AND e.eha_id NOT IN (
                      SELECT pr.eha_id
                      FROM performance_report_details_all pr
                      WHERE pr.review_month = %s
                  )
                ORDER BY e.first_name, e.last_name
            """, (month,))

            
        else:
            return jsonify({"error": "Invalid status"}), 400

        employees = cursor.fetchall()
        result = []

        for row in employees:

            if status == "missed_leader":

                deadline_date = (
                    row['submitted_at']
                    + timedelta(days=row['perform_verification'])
                )

                review_overdue = (
                    datetime.now() > deadline_date
                )

                row['review_deadline'] = str(deadline_date)
                row['review_overdue'] = review_overdue

            result.append(row)
        return jsonify({"employees": employees, "review_deadline_days": 7}), 200   # deadline_days can be dynamic

    except Exception as e:
        print("Error in HR review list:", e)
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()


#QUESTIONNAIRE////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


def calculate_due_info(question_dict, today):
    ask_every = question_dict.get('ask_every_x_days') or 0
    if ask_every <= 0:
        return None, None, False

    start_date = question_dict.get('valid_from') or question_dict.get('inserted_on')
    if not start_date:
        return None, None, False

    # Convert to date
    if isinstance(start_date, str):
        start_date = datetime.fromisoformat(start_date.replace(' ', 'T')).date()
    elif isinstance(start_date, datetime):
        start_date = start_date.date()
    if isinstance(today, datetime):
        today = today.date()

    if start_date > today:
        # Not yet started
        return start_date, (start_date - today).days, False

    days_since_start = (today - start_date).days
    cycles_passed = days_since_start // ask_every
    last_due = start_date + timedelta(days=cycles_passed * ask_every)
    next_due = last_due + timedelta(days=ask_every)
    days_left = (next_due - today).days
    is_overdue = today > last_due  # if today is after the last due date, and we have no record of answer -> overdue

    return next_due, days_left, is_overdue

from flask import request, jsonify, session
from datetime import datetime, timedelta
# Assume get_db() and calculate_due_info() are defined elsewhere. 
# The due‑date calculation can also be derived from the fixed schedule; 
# I’ve integrated a simplified version directly.

@app.route("/qa/my-questions", methods=["GET"])
def get_my_questions():
    try:
        if 'empid' not in session:
            return jsonify({"message": "Unauthorized"}), 401

        conn = get_db()
        cursor = conn.cursor()

        # Role mapping
        role_mapping = {"hr": 1, "employee": 2, "director": 4}
        conditions = []
        role_value = role_mapping.get(session.get('role'))
        if role_value:
            conditions.append(f"FIND_IN_SET('{role_value}', asking_to)")
        if session.get('is_team_lead') == 1:
            conditions.append("FIND_IN_SET('3', asking_to)")

        if not conditions:
            conn.close()
            return jsonify({"questions": []}), 200

        where_clause = " OR ".join(conditions)

        # We now need inserted_on to calculate the fixed‑schedule cycle
        cursor.execute(f"""
            SELECT
                id, hqa_id, question, question_type,
                length_of_desc, radio_option, selection_option,
                asking_to, ask_every_x_days, valid_from, valid_till,
                viewable_by, inserted_on
            FROM hr_qa_settings_all
            WHERE ({where_clause})
            ORDER BY id DESC
        """)
        questions = cursor.fetchall()

        final_questions = []
        today = datetime.now().date()

        for q in questions:
            q_dict = dict(q)
            hqa_id = q_dict['hqa_id']
            ask_every = q_dict.get('ask_every_x_days') or 0
            inserted_on = q_dict['inserted_on']  # datetime or None

            should_show = False
            cycle_no = 1
            submit_status = 0
            cycle_start = None
            cycle_end = None

            if ask_every > 0 and inserted_on is not None:
                # Fixed schedule: cycle 1 starts on inserted_on.date()
                inserted_date = inserted_on.date() if isinstance(inserted_on, datetime) else inserted_on
                elapsed_days = (today - inserted_date).days

                if elapsed_days >= 0:
                    # Current cycle number based on full periods elapsed
                    cycle_no = (elapsed_days // ask_every) + 1

                    # Cycle start date (first day of this cycle)
                    cycle_start = inserted_date + timedelta(days=(cycle_no - 1) * ask_every)
                    # Cycle end date (last day of this cycle)
                    cycle_end = cycle_start + timedelta(days=ask_every - 1)

                    # Check if user already answered this specific cycle
                    cursor.execute("""
                        SELECT id FROM hr_qa_answers_all
                        WHERE hqa_id = %s
                          AND eha_id = %s
                          AND question_cycle_no = %s
                          AND status = 1
                        LIMIT 1
                    """, (hqa_id, session['empid'], cycle_no))
                    already_answered = cursor.fetchone()

                    if not already_answered:
                        should_show = True
                        # Overdue if today is past the cycle end date
                        if today > cycle_end:
                            submit_status = 1
            else:
                # Non‑recurring question (ask_every = 0 or no inserted_on)
                # Show only if never answered at all (original behaviour)
                cursor.execute("""
                    SELECT id FROM hr_qa_answers_all
                    WHERE hqa_id = %s AND eha_id = %s AND status = 1
                    LIMIT 1
                """, (hqa_id, session['empid']))
                if not cursor.fetchone():
                    should_show = True
                    cycle_no = 1
                    submit_status = 0

            if should_show:
                # Parse selection options
                if q_dict.get('selection_option'):
                    q_dict['selection_option'] = [
                        opt.strip() for opt in q_dict['selection_option'].split(',') if opt.strip()
                    ]
                else:
                    q_dict['selection_option'] = []

                # Parse asking_to and viewable_by as lists of ints
                for field in ['asking_to', 'viewable_by']:
                    if q_dict.get(field):
                        q_dict[field] = [int(x.strip()) for x in str(q_dict[field]).split(',') if x.strip()]
                    else:
                        q_dict[field] = []

                # Calculate due date info for the frontend (using the fixed schedule)
                # next_due_date = cycle_end + 1 day (start of next cycle)
                if cycle_end:
                    next_due = cycle_end + timedelta(days=1)
                    days_left = (next_due - today).days
                    is_overdue = submit_status == 1
                else:
                    next_due = None
                    days_left = None
                    is_overdue = False

                q_dict['next_due_date'] = next_due.isoformat() if next_due else None
                q_dict['days_left'] = days_left
                q_dict['is_overdue'] = is_overdue

                q_dict['question_cycle_no'] = cycle_no
                q_dict['submit_status'] = submit_status

                final_questions.append(q_dict)

        conn.close()
        return jsonify({"success": True, "questions": final_questions}), 200

    except Exception as e:
        print(f"Error in get_my_questions: {e}")
        return jsonify({"message": str(e)}), 500


@app.route("/qa/submit-answer", methods=["POST"])
def submit_answer():
    try:
        if 'empid' not in session:
            return jsonify({"message": "Unauthorized"}), 401

        data = request.get_json()
        hqa_id = data.get('hqa_id')
        answer = data.get('answer')
        answer_option = data.get('answer_option')
        submit_status = data.get('submit_status', 0)

        if not hqa_id:
            return jsonify({"message": "Question ID is required"}), 400
        if not answer and not answer_option:
            return jsonify({"message": "Answer is required"}), 400

        conn = get_db()
        cursor = conn.cursor()

        # Fetch question to get viewable_by, ask_every_x_days, and inserted_on
        cursor.execute("""
            SELECT viewable_by, ask_every_x_days, inserted_on
            FROM hr_qa_settings_all
            WHERE hqa_id = %s
        """, (hqa_id,))
        question = cursor.fetchone()
        if not question:
            conn.close()
            return jsonify({"message": "Question not found"}), 404

        visible_to = question["viewable_by"]
        ask_every = question.get("ask_every_x_days") or 0
        inserted_on = question["inserted_on"]

        # =================================================
        # CALCULATE CYCLE NUMBER BASED ON FIXED SCHEDULE
        # =================================================
        if ask_every > 0 and inserted_on is not None:
            inserted_date = inserted_on.date() if isinstance(inserted_on, datetime) else inserted_on
            today = datetime.now().date()
            elapsed_days = (today - inserted_date).days
            if elapsed_days < 0:
                cycle_no = 1  # question not started yet, but allow early answer? We'll allow cycle 1.
            else:
                cycle_no = (elapsed_days // ask_every) + 1
        else:
            cycle_no = 1  # non‑recurring

        # =================================================
        # PREVENT DUPLICATE ANSWERS FOR THE SAME CYCLE
        # =================================================
        cursor.execute("""
            SELECT id FROM hr_qa_answers_all
            WHERE hqa_id = %s
              AND eha_id = %s
              AND question_cycle_no = %s
              AND status = 1
            LIMIT 1
        """, (hqa_id, session['empid'], cycle_no))
        if cursor.fetchone():
            conn.close()
            return jsonify({
                "message": f"You have already submitted an answer for cycle {cycle_no}."
            }), 409

        # Generate new answer ID
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'hr_qa_answers_all'")
        header = cursor.fetchone()
        if not header or header['last_id'] is None:
            cursor.execute("""
                INSERT INTO unique_header_all (table_name, prefix, last_id)
                VALUES ('hr_qa_answers_all', 'HQAANS', '0')
            """)
            conn.commit()
            new_id = 1
            hqa_answer_id = "HQAANS-00001"
        else:
            new_id = int(header['last_id']) + 1
            hqa_answer_id = f"HQAANS-{new_id:05d}"
            cursor.execute("""
                UPDATE unique_header_all
                SET last_id = %s, modified_on = NOW()
                WHERE table_name = 'hr_qa_answers_all'
            """, (new_id,))
            conn.commit()

        # Insert the answer with the correct cycle number
        cursor.execute("""
            INSERT INTO hr_qa_answers_all
            (hqa_answer_id, hqa_id, eha_id, answer, answer_option,
             question_cycle_no, submit_status, answered_on, inserted_by, status, visible_to)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), %s, 1, %s)
        """, (
            hqa_answer_id,
            hqa_id,
            session['empid'],
            answer,
            answer_option,
            cycle_no,
            submit_status,
            session['empid'],
            visible_to
        ))

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Answer submitted successfully",
            "hqa_answer_id": hqa_answer_id,
            "cycle_number": cycle_no
        }), 201

    except Exception as e:
        print(f"Error in submit_answer: {e}")
        return jsonify({"message": str(e)}), 500







@app.route("/qa/tl-responses/<hqa_id>", methods=["GET"])
def get_tl_responses(hqa_id):

    try:

        if 'empid' not in session:
            return jsonify({
                "message": "Unauthorized"
            }), 401

        if session.get('is_team_lead') != 1:
            return jsonify({
                "message": "Only Team Leads allowed"
            }), 403

        conn = get_db()
        cursor = conn.cursor()

        # =========================================
        # CURRENT CYCLE CALCULATION
        # =========================================

        cursor.execute("""
            SELECT ask_every_x_days, inserted_on
            FROM hr_qa_settings_all
            WHERE hqa_id = %s
        """, (hqa_id,))

        question = cursor.fetchone()

        current_cycle = 1

        if question and question['ask_every_x_days'] and question['inserted_on']:
            inserted_on = question['inserted_on']
            if isinstance(inserted_on, datetime):
                inserted_date = inserted_on.date()
            else:
                inserted_date = inserted_on
            today = datetime.now().date()
            elapsed = (today - inserted_date).days
            if elapsed >= 0:
                current_cycle = (elapsed // question['ask_every_x_days']) + 1

        # =========================================
        # SUBMITTED RESPONSES
        # =========================================

        cursor.execute("""

            SELECT

                a.hqa_answer_id,
                a.hqa_id,
                a.eha_id,

                a.answer,
                a.answer_option,

                a.question_cycle_no,
                a.submit_status,

                a.visible_to,

                a.answered_on,

                q.question,
                q.question_type,

                e.first_name,
                e.last_name

            FROM hr_qa_answers_all a

            JOIN hr_qa_settings_all q
                ON a.hqa_id = q.hqa_id

            JOIN employee_header_all e
                ON a.eha_id = e.eha_id

            WHERE

                a.status = 1

                AND a.hqa_id = %s

                AND a.question_cycle_no = %s

                AND FIND_IN_SET(
                    '3',
                    REPLACE(a.visible_to, ' ', '')
                ) > 0

                AND EXISTS (
                    SELECT 1
                    FROM team_emp_combination_all tec
                    WHERE
                        tec.eha_id = a.eha_id
                        AND tec.unlinked_on IS NULL
                        AND tec.link_type = 1
                        AND tec.tha_id IN (
                            SELECT tha_id
                            FROM team_emp_combination_all
                            WHERE
                                eha_id = %s
                                AND link_type = 2
                                AND unlinked_on IS NULL
                        )
                )

            ORDER BY
                a.answered_on DESC

        """, (
            hqa_id,
            current_cycle,
            session['empid']
        ))

        submitted = cursor.fetchall()

        submitted_responses = []

        for r in submitted:

            r_dict = dict(r)

            r_dict['employee_name'] = (
                f"{r_dict['first_name']} "
                f"{r_dict['last_name']}"
            )

            r_dict['answered_on'] = str(
                r_dict['answered_on']
            )

            submitted_responses.append(r_dict)

        # =========================================
        # MISSED EMPLOYEES
        # =========================================

        cursor.execute("""

            SELECT

                e.eha_id,
                e.first_name,
                e.last_name

            FROM employee_header_all e

            WHERE

                e.status = 1

                AND EXISTS (
                    SELECT 1
                    FROM team_emp_combination_all tec
                    WHERE
                        tec.eha_id = e.eha_id
                        AND tec.unlinked_on IS NULL
                        AND tec.link_type = 1
                        AND tec.tha_id IN (
                            SELECT tha_id
                            FROM team_emp_combination_all
                            WHERE
                                eha_id = %s
                                AND link_type = 2
                                AND unlinked_on IS NULL
                        )
                )

                AND NOT EXISTS (
                    SELECT 1
                    FROM hr_qa_answers_all a
                    WHERE
                        a.hqa_id = %s
                        AND a.eha_id = e.eha_id
                        AND a.question_cycle_no = %s
                )

        """, (
            session['empid'],
            hqa_id,
            current_cycle
        ))

        missed = cursor.fetchall()

        missed_employees = []

        for e in missed:

            e_dict = dict(e)

            e_dict['employee_name'] = (
                f"{e_dict['first_name']} "
                f"{e_dict['last_name']}"
            )

            e_dict['question_cycle_no'] = current_cycle

            missed_employees.append(e_dict)

        conn.close()

        return jsonify({
            "success": True,
            "current_cycle": current_cycle,
            "submitted": submitted_responses,
            "missed": missed_employees
        }), 200

    except Exception as e:

        print(f"Error: {e}")

        return jsonify({
            "message": str(e)
        }), 500


@app.route("/qa/responses/<hqa_id>", methods=["GET"])
def get_responses(hqa_id):
    try:
        if 'empid' not in session:
            return jsonify({"message": "Unauthorized"}), 401

        conn = get_db()
        cursor = conn.cursor()

        # 1. Fetch question schedule for current cycle calculation
        cursor.execute("""
            SELECT ask_every_x_days, inserted_on
            FROM hr_qa_settings_all
            WHERE hqa_id = %s
        """, (hqa_id,))
        question = cursor.fetchone()

        current_cycle = 1
        if question and question['ask_every_x_days'] and question['inserted_on']:
            inserted_on = question['inserted_on']
            if isinstance(inserted_on, datetime):
                inserted_date = inserted_on.date()
            else:
                inserted_date = inserted_on
            today = datetime.now().date()
            elapsed = (today - inserted_date).days
            if elapsed >= 0:
                current_cycle = (elapsed // question['ask_every_x_days']) + 1

        submitted_responses = []
        added_answer_ids = set()

        role_mapping = {"hr": "1", "director": "4"}
        role_visibility = role_mapping.get(session.get('role'))

        # HR / DIRECTOR ANSWERS
        if role_visibility:
            cursor.execute(f"""
                SELECT a.hqa_answer_id, a.hqa_id, a.eha_id,
                       a.answer, a.answer_option,
                       a.question_cycle_no, a.submit_status,
                       a.visible_to, a.answered_on,
                       q.question, q.question_type,
                       e.first_name, e.last_name
                FROM hr_qa_answers_all a
                JOIN hr_qa_settings_all q ON a.hqa_id = q.hqa_id
                JOIN employee_header_all e ON a.eha_id = e.eha_id
                WHERE a.status = 1
                  AND a.hqa_id = %s
                  AND a.question_cycle_no = %s
                  AND FIND_IN_SET('{role_visibility}', a.visible_to)
                ORDER BY a.answered_on DESC
            """, (hqa_id, current_cycle))

            for r in cursor.fetchall():
                r_dict = dict(r)
                if r_dict['hqa_answer_id'] in added_answer_ids:
                    continue
                added_answer_ids.add(r_dict['hqa_answer_id'])
                r_dict['employee_name'] = f"{r_dict['first_name']} {r_dict['last_name']}"
                r_dict['answered_on'] = str(r_dict['answered_on'])
                submitted_responses.append(r_dict)

        # TEAM LEAD ANSWERS
        if session.get('is_team_lead') == 1:
            cursor.execute("""
                SELECT a.hqa_answer_id, a.hqa_id, a.eha_id,
                       a.answer, a.answer_option,
                       a.question_cycle_no, a.submit_status,
                       a.visible_to, a.answered_on,
                       q.question, q.question_type,
                       e.first_name, e.last_name
                FROM hr_qa_answers_all a
                JOIN hr_qa_settings_all q ON a.hqa_id = q.hqa_id
                JOIN employee_header_all e ON a.eha_id = e.eha_id
                JOIN team_emp_combination_all tec ON a.eha_id = tec.eha_id
                WHERE a.status = 1
                  AND a.hqa_id = %s
                  AND a.question_cycle_no = %s
                  AND FIND_IN_SET('3', a.visible_to)
                  AND tec.unlinked_on IS NULL
                  AND tec.link_type = 1
                  AND tec.tha_id IN (
                      SELECT tha_id FROM team_emp_combination_all
                      WHERE eha_id = %s AND link_type = 2 AND unlinked_on IS NULL
                  )
                ORDER BY a.answered_on DESC
            """, (hqa_id, current_cycle, session['empid']))

            for r in cursor.fetchall():
                r_dict = dict(r)
                if r_dict['hqa_answer_id'] in added_answer_ids:
                    continue
                added_answer_ids.add(r_dict['hqa_answer_id'])
                r_dict['employee_name'] = f"{r_dict['first_name']} {r_dict['last_name']}"
                r_dict['answered_on'] = str(r_dict['answered_on'])
                submitted_responses.append(r_dict)

        # MISSED EMPLOYEES (current cycle)
        missed_employees = []
        cursor.execute("""
            SELECT e.eha_id, e.first_name, e.last_name
            FROM employee_header_all e
            WHERE e.status = 1
              AND NOT EXISTS (
                  SELECT 1 FROM hr_qa_answers_all a
                  WHERE a.hqa_id = %s
                    AND a.eha_id = e.eha_id
                    AND a.question_cycle_no = %s
              )
        """, (hqa_id, current_cycle))

        for e in cursor.fetchall():
            e_dict = dict(e)
            e_dict['employee_name'] = f"{e_dict['first_name']} {e_dict['last_name']}"
            e_dict['question_cycle_no'] = current_cycle
            missed_employees.append(e_dict)

        conn.close()

        return jsonify({
            "success": True,
            "current_cycle": current_cycle,
            "submitted": submitted_responses,
            "missed": missed_employees
        }), 200

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500


@app.route("/qa/viewable-questions", methods=["GET"])
def get_viewable_questions():

    try:

        if 'empid' not in session:
            return jsonify({
                "message": "Unauthorized"
            }), 401

        conn = get_db()
        cursor = conn.cursor()

        final_questions = []
        print("kajbgsakjbgkajsbkjb")

        added_questions = set()
        rolll = session.get('role')
        print(rolll)
        # =========================================
        # HR QUESTIONS
        # =========================================

        if session.get('role') == 'hr':

            cursor.execute("""
                SELECT

                    hqa_id,
                    question,
                    question_type,

                    length_of_desc,
                    radio_option,
                    selection_option,

                    asking_to,
                    ask_every_x_days,

                    viewable_by,

                    inserted_on

                FROM hr_qa_settings_all

                WHERE FIND_IN_SET(%s, viewable_by)

                ORDER BY inserted_on DESC
            """, ("1",))

            hr_questions = cursor.fetchall()

            for q in hr_questions:

                q_dict = dict(q)

                if q_dict['hqa_id'] in added_questions:
                    continue

                added_questions.add(
                    q_dict['hqa_id']
                )

                # Selection Options
                if q_dict.get('selection_option'):

                    q_dict['selection_option'] = [
                        opt.strip()
                        for opt in q_dict[
                            'selection_option'
                        ].split(',')
                        if opt.strip()
                    ]

                else:

                    q_dict['selection_option'] = []

                # Asking To
                if q_dict.get('asking_to'):

                    q_dict['asking_to'] = [
                        int(x.strip())
                        for x in str(
                            q_dict['asking_to']
                        ).split(',')
                        if x.strip()
                    ]

                else:

                    q_dict['asking_to'] = []

                # Viewable By
                if q_dict.get('viewable_by'):

                    q_dict['viewable_by'] = [
                        int(x.strip())
                        for x in str(
                            q_dict['viewable_by']
                        ).split(',')
                        if x.strip()
                    ]

                else:

                    q_dict['viewable_by'] = []

                # Inserted On
                if q_dict.get('inserted_on'):

                    q_dict['inserted_on'] = str(
                        q_dict['inserted_on']
                    )

                final_questions.append(q_dict)

        # =========================================
        # DIRECTOR QUESTIONS
        # =========================================

        if session.get('role') == 'director':

            cursor.execute("""
                SELECT

                    hqa_id,
                    question,
                    question_type,

                    length_of_desc,
                    radio_option,
                    selection_option,

                    asking_to,
                    ask_every_x_days,

                    viewable_by,

                    inserted_on

                FROM hr_qa_settings_all

                WHERE FIND_IN_SET(%s, viewable_by)

                ORDER BY inserted_on DESC
            """, ("4",))

            director_questions = cursor.fetchall()

            for q in director_questions:

                q_dict = dict(q)

                if q_dict['hqa_id'] in added_questions:
                    continue

                added_questions.add(
                    q_dict['hqa_id']
                )

                # Selection Options
                if q_dict.get('selection_option'):

                    q_dict['selection_option'] = [
                        opt.strip()
                        for opt in q_dict[
                            'selection_option'
                        ].split(',')
                        if opt.strip()
                    ]

                else:

                    q_dict['selection_option'] = []

                # Asking To
                if q_dict.get('asking_to'):

                    q_dict['asking_to'] = [
                        int(x.strip())
                        for x in str(
                            q_dict['asking_to']
                        ).split(',')
                        if x.strip()
                    ]

                else:

                    q_dict['asking_to'] = []

                # Viewable By
                if q_dict.get('viewable_by'):

                    q_dict['viewable_by'] = [
                        int(x.strip())
                        for x in str(
                            q_dict['viewable_by']
                        ).split(',')
                        if x.strip()
                    ]

                else:

                    q_dict['viewable_by'] = []

                # Inserted On
                if q_dict.get('inserted_on'):

                    q_dict['inserted_on'] = str(
                        q_dict['inserted_on']
                    )

                final_questions.append(q_dict)

        # =========================================
        # TEAM LEAD QUESTIONS
        # =========================================

        if session.get('is_team_lead') == 1:

            cursor.execute("""
                SELECT

                    hqa_id,
                    question,
                    question_type,

                    length_of_desc,
                    radio_option,
                    selection_option,

                    asking_to,
                    ask_every_x_days,

                    viewable_by,

                    inserted_on

                FROM hr_qa_settings_all

                WHERE FIND_IN_SET(%s, viewable_by)

                ORDER BY inserted_on DESC
            """, ("3",))

            tl_questions = cursor.fetchall()

            for q in tl_questions:

                q_dict = dict(q)

                if q_dict['hqa_id'] in added_questions:
                    continue

                added_questions.add(
                    q_dict['hqa_id']
                )

                # Selection Options
                if q_dict.get('selection_option'):

                    q_dict['selection_option'] = [
                        opt.strip()
                        for opt in q_dict[
                            'selection_option'
                        ].split(',')
                        if opt.strip()
                    ]

                else:

                    q_dict['selection_option'] = []

                # Asking To
                if q_dict.get('asking_to'):

                    q_dict['asking_to'] = [
                        int(x.strip())
                        for x in str(
                            q_dict['asking_to']
                        ).split(',')
                        if x.strip()
                    ]

                else:

                    q_dict['asking_to'] = []

                # Viewable By
                if q_dict.get('viewable_by'):

                    q_dict['viewable_by'] = [
                        int(x.strip())
                        for x in str(
                            q_dict['viewable_by']
                        ).split(',')
                        if x.strip()
                    ]

                else:

                    q_dict['viewable_by'] = []

                # Inserted On
                if q_dict.get('inserted_on'):

                    q_dict['inserted_on'] = str(
                        q_dict['inserted_on']
                    )

                final_questions.append(q_dict)

        conn.close()

        return jsonify({
            "success": True,
            "questions": final_questions
        }), 200

    except Exception as e:

        print(f"Error: {e}")

        return jsonify({
            "message": str(e)
        }), 500






@app.route("/gettasks",methods=["GET"])
def get_tasks():
    empid = session.get("empid")
    if(not empid):
        return jsonify({"Error":"Not logged in"}),401
    
    task_date = request.args.get("date")
    if not task_date:
        task_date = datetime.today().date()         #datetime 2026-04-23 14:35:22.123456 .date 2026-04-23

    task_date = request.args.get("date")
    if task_date:
        today = datetime.strptime(task_date,"%Y-%m-%d")
    else:
        today = datetime.today()
    day_col = f"day_{today.day}"   #dynamic column name
    month = today.month
    year = today.year

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            f"SELECT taskid, {day_col} as content, status, month, year FROM tasks WHERE empid = %s AND month = %s AND year = %s",
            (empid, month, year)        #,for tuple as date should not be changed
        )              

        row = cursor.fetchone()
        result = []

        if not row or not row["content"]:
            return jsonify({"message":"No Tasks"}),200
        
        tasks_list = [{"taskid": i, "content": t, "status": row["status"]} for i, t in enumerate(json.loads(row["content"]))]
        return jsonify(tasks_list),200
    except Exception as e:
        return jsonify({"error":str(e)}),500
    finally:
        cursor.close()
        conn.close()




@app.route("/me")
def me():
    if "empid" in session:
        return jsonify({
            "empid":session["empid"],
            "role" :session["role"]
        })
    return jsonify({
        "message":"not logged in"
    }),401



#HR SETTINGS/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

@app.route("/hr-settings", methods=["GET"])
def get_hr_settings():
    try:
        setting_type = request.args.get('type', 'present')
        setting_for = request.args.get('for', 'general')
        today = datetime.now().date()
        
        conn = get_db()
        cursor = conn.cursor()
        
        if setting_for == "special":
            table = "hr_special_approvals_settings_all"
            base_query = f"""
                SELECT s.*, e.first_name, e.last_name
                FROM {table} s
                LEFT JOIN employee_header_all e ON s.eha_id = e.eha_id
            """
            alias = "s"
        else:
            table = "hr_general_settings_all"
            base_query = f"SELECT * FROM {table}"
            alias = ""
        
        if setting_type == "present":
            if alias:
                query = base_query + f" WHERE {alias}.valid_from <= %s AND ({alias}.valid_till IS NULL OR {alias}.valid_till >= %s) ORDER BY {alias}.valid_from DESC"
                params = (today, today)
            else:
                query = base_query + " WHERE valid_from <= %s AND (valid_till IS NULL OR valid_till > %s) ORDER BY valid_from DESC"
                params = (today, today)
        elif setting_type == "upcoming":
            if alias:
                query = base_query + f" WHERE {alias}.valid_from > %s ORDER BY {alias}.valid_from ASC"
                params = (today,)
            else:
                query = base_query + " WHERE valid_from > %s ORDER BY valid_from ASC"
                params = (today,)
        else:  
            if alias:
                query = base_query + f" WHERE {alias}.valid_till IS NOT NULL AND {alias}.valid_till < %s ORDER BY {alias}.valid_till DESC"
                params = (today,)
            else:
                query = base_query + " WHERE valid_till IS NOT NULL AND valid_till < %s ORDER BY valid_till DESC"
                params = (today,)
        
        cursor.execute(query, params)
        settings = cursor.fetchall()
        conn.close()

        result = []
        for s in settings:
            item = dict(s)
            for key, value in item.items():
                if isinstance(value, timedelta):              
                    total_seconds = int(value.total_seconds())
                    hours, remainder = divmod(total_seconds, 3600)
                    minutes = remainder // 60
                    item[key] = f"{hours:02d}:{minutes:02d}"
                elif isinstance(value, datetime):        # datetime BEFORE date
                    item[key] = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, date):            # date AFTER datetime
                    item[key] = value.isoformat()
                elif isinstance(value, time):
                    item[key] = value.strftime("%H:%M")
                # (anything else is left as-is, e.g., int, str)
            result.append(item)
            print(result)
        
        return jsonify({"settings": result}), 200
    except Exception as e:
        print(f"Error in get_hr_settings: {e}")
        return jsonify({"message": str(e)}), 500


@app.route("/create_hrgen", methods=["POST"])
def create_hr_settings():
    try:
        if session.get('role') not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        data = request.get_json()
        today = datetime.now().date()
        print(data)
        
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE hr_general_settings_all
            SET valid_till = %s
            WHERE valid_till IS NULL
        """, (today,))
        
        
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'hr_general_settings_all'")
        header = cursor.fetchone()
        
        if not header or header['last_id'] is None:
            cursor.execute("INSERT INTO unique_header_all (table_name, prefix, last_id) VALUES ('hr_general_settings_all', 'HGS', '0')")
            conn.commit()
            new_id = 1
            hgs_id = "HGS-00001"
        else:
            new_id = int(header['last_id']) + 1
            hgs_id = f"HGS-{new_id:05d}"
            cursor.execute("UPDATE unique_header_all SET last_id = %s, modified_on = NOW() WHERE table_name = 'hr_general_settings_all'", (str(new_id),))
        
        cursor.execute("""
            INSERT INTO hr_general_settings_all (
                hgs_id, 
                daily_agenda_start, 
                daily_agenda_end, 
                daily_report_start, 
                daily_report_end, 
                submit_next_day, 
                perform_verify_start, 
                perform_verify_end, 
                perform_verification,
                valid_from,
                valid_till
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NULL)
        """, (
            hgs_id, 
            data.get('daily_agenda_start'), 
            data.get('daily_agenda_end'), 
            data.get('daily_report_start'), 
            data.get('daily_report_end'), 
            data.get('submit_next_day'), 
            data.get('perform_verify_start'), 
            data.get('perform_verify_end'), 
            data.get('perform_verification'),
            today
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Settings created", "hgs_id": hgs_id}), 201
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500

@app.route("/hr-settings/update/<setting_id>", methods=["PUT"])
def update_hr_settings(setting_id):
    try:
        setting_for = request.args.get('for', 'general')
        data = request.json

        conn = get_db()
        cursor = conn.cursor()

        if setting_for == "general":
            cursor.execute("""
                UPDATE hr_general_settings_all SET
                    daily_agenda_start = %s,
                    daily_agenda_end = %s,
                    daily_report_start = %s,
                    daily_report_end = %s,
                    submit_next_day = %s,
                    perform_verify_start = %s,
                    perform_verify_end = %s,
                    perform_verification = %s,
                    valid_from = %s
                WHERE hgs_id = %s
            """, (
                data.get("daily_agenda_start"),
                data.get("daily_agenda_end"),
                data.get("daily_report_start"),
                data.get("daily_report_end"),
                data.get("submit_next_day"),
                data.get("perform_verify_start"),
                data.get("perform_verify_end"),
                data.get("perform_verification"),
                data.get("valid_from"),
                setting_id
            ))
        else:
            cursor.execute("""
                UPDATE hr_special_approvals_settings_all SET
                    early_agenda_submit = %s,
                    late_agenda_submit = %s,
                    early_report_submit = %s,
                    late_report_submit = %s,
                    submit_next_day = %s,
                    valid_from = %s,
                    valid_till = %s
                WHERE hsa_id = %s
            """, (
                data.get("early_agenda_submit"),
                data.get("late_agenda_submit"),
                data.get("early_report_submit"),
                data.get("late_report_submit"),
                data.get("submit_next_day"),
                data.get("valid_from"),
                data.get("valid_till"),
                setting_id
            ))

        conn.commit()
        conn.close()
        return jsonify({"message": "Updated successfully"}), 200
    except Exception as e:
        print(f"Update error: {e}")
        return jsonify({"message": str(e)}), 500
    

@app.route("/qa/questions", methods=["GET", "POST"])
def qa_questions():
    try:

        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403

        conn = get_db()
        cursor = conn.cursor()

        # ========================= GET =========================
        if request.method == "GET":

            cursor.execute("""
                SELECT 
                    id,
                    hqa_id,
                    question,
                    question_type,
                    length_of_desc,
                    selection_option,
                    viewable_by,
                    valid_from,
                    valid_till,
                    asking_to,
                    ask_every_x_days

                FROM hr_qa_settings_all

                ORDER BY id DESC
            """)

            questions = cursor.fetchall()

            conn.close()

            result = []

            for q in questions:

                q_dict = dict(q)

                if q_dict.get('valid_from'):
                    q_dict['valid_from'] = str(q_dict['valid_from'])

                if q_dict.get('valid_till'):
                    q_dict['valid_till'] = str(q_dict['valid_till'])

                # Selection Options
                if q_dict.get('selection_option'):

                    q_dict['selection_option'] = [
                        opt.strip()
                        for opt in q_dict['selection_option'].split(',')
                        if opt.strip()
                    ]

                else:
                    q_dict['selection_option'] = []

                # Viewable By
                if q_dict.get('viewable_by'):

                    q_dict['viewable_by'] = [
                        int(x.strip())
                        for x in str(q_dict['viewable_by']).split(',')
                        if x.strip()
                    ]

                else:
                    q_dict['viewable_by'] = []

                # Asking To
                if q_dict.get('asking_to'):

                    q_dict['asking_to'] = [
                        int(x.strip())
                        for x in str(q_dict['asking_to']).split(',')
                        if x.strip()
                    ]

                else:
                    q_dict['asking_to'] = []

                result.append(q_dict)

            return jsonify({
                "questions": result
            }), 200

        # ========================= CREATE =========================
        elif request.method == "POST":

            data = request.get_json()
            print(data)
            print("akjgbaskjgba")

            viewable_by = ','.join(
                map(str, data.get('viewable_by', []))
            )

            selection_option = ','.join(
                data.get('selection_option', [])
            )


            asking_to = ','.join(
                map(str, data.get('asking_to', []))
            )

            ask_every_x_days = data.get('ask_every_x_days')

            # Generate HQA ID
            cursor.execute("""
                SELECT *
                FROM unique_header_all
                WHERE table_name = 'hr_qa_settings_all'
            """)

            header = cursor.fetchone()

            if not header or header['last_id'] is None:

                cursor.execute("""
                    INSERT INTO unique_header_all
                    (
                        table_name,
                        prefix,
                        last_id
                    )
                    VALUES
                    (
                        'hr_qa_settings_all',
                        'HQA',
                        '0'
                    )
                """)

                conn.commit()

                new_id = 1
                hqa_id = "HQA-00001"

            else:

                new_id = int(header['last_id']) + 1

                hqa_id = f"HQA-{new_id:05d}"

                cursor.execute("""
                    UPDATE unique_header_all
                    SET
                        last_id = %s,
                        modified_on = NOW()

                    WHERE table_name = 'hr_qa_settings_all'
                """, (new_id,))

                conn.commit()

            # Insert Question
            cursor.execute("""
                INSERT INTO hr_qa_settings_all
                (
                    hqa_id,
                    question,
                    question_type,
                    length_of_desc,
                    selection_option,
                    viewable_by,
                    asking_to,
                    ask_every_x_days,
                    entered_by
                )

                VALUES
                (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )
            """, (
                hqa_id,
                data.get('question'),
                data.get('question_type'),
                str(data.get('length_of_desc', '')),
                selection_option,
                viewable_by,
                asking_to,
                ask_every_x_days,
                session.get('empid', 'UNKNOWN')
            ))

            conn.commit()

            conn.close()

            return jsonify({
                "success": True,
                "message": "Question created",
                "hqa_id": hqa_id
            }), 201

    except Exception as e:

        print(f"Error: {e}")

        return jsonify({
            "message": str(e)
        }), 500
    

@app.route("/qa/questions/<identifier>", methods=["DELETE"])
def delete_qa_question(identifier):
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Delete by either id or hqa_id
        cursor.execute("""
            DELETE FROM hr_qa_settings_all 
            WHERE id = %s OR hqa_id = %s
        """, (identifier, identifier))
        
        conn.commit()
        affected_rows = cursor.rowcount
        conn.close()
        
        if affected_rows > 0:
            return jsonify({"success": True, "message": "Question deleted successfully"}), 200
        else:
            return jsonify({"message": "Question not found"}), 404
            
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500
    
def expire_overlapping_settings(cursor, table_name, new_valid_from, eha_id=None):
    """Expire settings that overlap with the new valid_from date.
    For special approvals, also filter by eha_id."""
    if eha_id:
        # For special approvals - expire only for that employee
        if new_valid_from:
            cursor.execute(f"""
                UPDATE {table_name}
                SET valid_till = DATE_SUB(%s, INTERVAL 1 DAY)
                WHERE (valid_till IS NULL OR valid_till >= %s)
                AND valid_from < %s
                AND eha_id = %s
            """, (new_valid_from, new_valid_from, new_valid_from, eha_id))
        else:
            cursor.execute(f"""
                UPDATE {table_name}
                SET valid_till = CURDATE()
                WHERE valid_till IS NULL OR valid_till >= CURDATE()
                AND eha_id = %s
            """, (eha_id,))
    else:
        # For general settings - expire all
        if new_valid_from:
            cursor.execute(f"""
                UPDATE {table_name}
                SET valid_till = DATE_SUB(%s, INTERVAL 1 DAY)
                WHERE (valid_till IS NULL OR valid_till >= %s)
                AND valid_from < %s
            """, (new_valid_from, new_valid_from, new_valid_from))
        else:
            cursor.execute(f"""
                UPDATE {table_name}
                SET valid_till = CURDATE()
                WHERE valid_till IS NULL OR valid_till >= CURDATE()
            """)

@app.route("/create_special_approval", methods=["POST"])
def create_special_approval():
    try:
        if session['role'] not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        data = request.get_json()
        
        conn = get_db()
        cursor = conn.cursor()

        expire_overlapping_settings(cursor, 'hr_special_approvals_settings_all', data.get('valid_from'), data.get('eha_id'))
        
        cursor.execute("SELECT * FROM unique_header_all WHERE table_name = 'hr_special_approvals_settings_all'")
        header = cursor.fetchone()
        
        if not header or header['last_id'] is None:
            cursor.execute("INSERT INTO unique_header_all (table_name, prefix, last_id) VALUES ('hr_special_approvals_settings_all', 'HSA', '1')")
            conn.commit()
            new_id = 1
            hsa_id = "HSA-00001"
        else:
            new_id = int(header['last_id']) + 1
            hsa_id = f"HSA-{new_id:05d}"
            cursor.execute("UPDATE unique_header_all SET last_id = %s, modified_on = NOW() WHERE table_name = 'hr_special_approvals_settings_all'", (new_id,))
            conn.commit()
        
        cursor.execute("""
            INSERT INTO hr_special_approvals_settings_all (
                hsa_id,
                eha_id,
                early_agenda_submit,
                late_agenda_submit,
                early_report_submit,
                late_report_submit,
                submit_next_day,
                valid_from,
                valid_till
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            hsa_id,
            data.get('eha_id'),
            data.get('early_agenda_submit'),
            data.get('late_agenda_submit'),
            data.get('early_report_submit'),
            data.get('late_report_submit'),
            data.get('submit_next_day'),
            data.get('valid_from'),
            data.get('valid_till')
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Special approval created", "hsa_id": hsa_id}), 201
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": str(e)}), 500
    
@app.route("/hr-settings/special/<hsa_id>", methods=["DELETE"])
def delete_special_approval(hsa_id):
    try:
        if session.get('role') not in ['admin', 'hr', 'director']:
            return jsonify({"message": "Access Denied"}), 403
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT hsa_id FROM hr_special_approvals_settings_all WHERE hsa_id = %s", (hsa_id,))
        existing = cursor.fetchone()
        if not existing:
            conn.close()
            return jsonify({"message": "Special approval not found"}), 404
        
        cursor.execute("DELETE FROM hr_special_approvals_settings_all WHERE hsa_id = %s", (hsa_id,))
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Deleted successfully"}), 200
    except Exception as e:
        print(f"Error deleting special approval: {e}")
        return jsonify({"message": str(e)}), 500
    

#Calendar settings 

@app.route('/company_calendar', methods=['GET'])
def get_calendar():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    if not month or not year:
        return jsonify({'error': 'month and year are required'}), 400

    try:
        
        conn = get_db()
        cursor =conn.cursor()
        cursor.execute("""
            SELECT calendar_date, day_type, holiday_name, remarks, is_default
            FROM company_calendar_settings_all
            WHERE MONTH(calendar_date) = %s AND YEAR(calendar_date) = %s
            ORDER BY calendar_date ASC
        """, (month, year))

        rows = cursor.fetchall()

        # convert date to string for JSON
        for row in rows:
            row['calendar_date'] = row['calendar_date'].strftime('%Y-%m-%d')

        return jsonify(rows), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/company_calendar/update', methods=['POST'])
def update_calendar_day():
    data = request.get_json()

    date       = data.get('date')
    day_type   = data.get('day_type')
    holiday_name = data.get('holiday_name')
    remarks    = data.get('remarks')
    changed_by = session.get('eha_id')  # from your session

    if not date or not day_type:
        return jsonify({'error': 'date and day_type are required'}), 400

    # block past date edits
    if date < datetime.today().strftime('%Y-%m-%d'):
        return jsonify({'error': 'Cannot edit past dates'}), 400

    try:
        conn = get_db()
        cursor =conn.cursor()

        # get current state before update
        cursor.execute("""
            SELECT day_type, holiday_name 
            FROM company_calendar_settings_all
            WHERE calendar_date = %s
        """, (date,))
        existing = cursor.fetchone()

        # upsert
        cursor.execute("""
            INSERT INTO company_calendar_settings_all 
                (calendar_date, day_type, holiday_name, remarks, is_default)
            VALUES (%s, %s, %s, %s, FALSE)
            ON DUPLICATE KEY UPDATE
                day_type     = VALUES(day_type),
                holiday_name = VALUES(holiday_name),
                remarks      = VALUES(remarks),
                is_default   = FALSE
        """, (date, day_type, holiday_name, remarks))

        # log to history
        cursor.execute("""
            INSERT INTO company_calendar_history_all
                (calendar_date, day_type, previous_type, holiday_name, previous_holiday_name, remarks, changed_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            date,
            day_type,
            existing['day_type'] if existing else None,
            holiday_name,
            existing['holiday_name'] if existing else None,
            remarks,
            changed_by
        ))

        cursor.close()
        conn.commit()
        conn.close()
        return jsonify({'message': 'Calendar updated successfully'}), 200

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/company_calendar/template', methods=['GET'])
def download_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2026 Holidays"

        # Headers
        headers = ['Date', 'Day Type', 'Holiday Name', 'Remarks']
        ws.append(headers)

        # Header styling
        header_fill = PatternFill(
            start_color='1F4E78',
            end_color='1F4E78',
            fill_type='solid'
        )

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Sample holiday data
        holidays = [
            ['2026-01-01', 'Holiday', "New Year's Day", ''],
            ['2026-01-13', 'Holiday', 'Lohri', ''],
            ['2026-01-14', 'Holiday', 'Makar Sankranti / Pongal', ''],
            ['2026-01-23', 'Holiday', 'Netaji Jayanti', ''],
            ['2026-01-26', 'Holiday', 'Republic Day', ''],
            ['2026-02-15', 'Holiday', 'Vasant Panchami', ''],
            ['2026-02-19', 'Holiday', 'Shivaji Maharaj Jayanti', ''],
            ['2026-03-04', 'Holiday', 'Holi', ''],
            ['2026-03-20', 'Holiday', 'Ugadi / Gudi Padwa', ''],
        ]

        for row in holidays:
            ws.append(row)

        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Dropdown for Day Type
        dv = DataValidation(
            type="list",
            formula1='"Working,Holiday,Weekend"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        dv.add('B2:B500')

        # Center align date + day type
        for row in ws.iter_rows(min_row=2, max_row=500, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name='calendar_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/company_calendar/upload', methods=['POST'])
def upload_calendar():
    file = request.files.get('file')
    changed_by = session.get('eha_id')

    if not file:
        return jsonify({'error': 'No file provided'}), 400

    conn = None

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        if not rows:
            return jsonify({'error': 'File is empty'}), 400

        conn = get_db()
        cursor = conn.cursor()

        success = 0
        skipped = 0
        failed = []

        today = datetime.today().date()

        for i, row in enumerate(rows, start=2):
            try:
                raw_date = row[0]

                if not raw_date:
                    continue

                # convert excel/string date properly
                if isinstance(raw_date, datetime):
                    calendar_date = raw_date.date()
                else:
                    calendar_date = datetime.strptime(
                        str(raw_date).strip(),
                        '%Y-%m-%d'
                    ).date()

                # skip past dates
                if calendar_date < today:
                    skipped += 1
                    continue

                day_type = str(row[1]).strip().lower() if row[1] else 'working'
                holiday_name = str(row[2]).strip() if row[2] else None
                remarks = str(row[3]).strip() if row[3] else None

                valid_types = ['working', 'holiday', 'weekoff', 'half_day']

                if day_type not in valid_types:
                    failed.append({
                        'row': i,
                        'reason': f'Invalid day_type: {day_type}'
                    })
                    continue

                # existing record
                cursor.execute("""
                    SELECT day_type, holiday_name
                    FROM company_calendar_settings_all
                    WHERE calendar_date = %s
                """, (calendar_date,))

                existing = cursor.fetchone()

                # insert/update only future dates
                cursor.execute("""
                    INSERT INTO company_calendar_settings_all
                        (calendar_date, day_type, holiday_name, remarks, is_default)
                    VALUES (%s, %s, %s, %s, FALSE)
                    ON DUPLICATE KEY UPDATE
                        day_type = VALUES(day_type),
                        holiday_name = VALUES(holiday_name),
                        remarks = VALUES(remarks),
                        is_default = FALSE
                """, (
                    calendar_date,
                    day_type,
                    holiday_name,
                    remarks
                ))

                # history log only if changed
                if (
                    not existing or
                    existing['day_type'] != day_type or
                    existing['holiday_name'] != holiday_name
                ):
                    cursor.execute("""
                        INSERT INTO company_calendar_history_all
                            (
                                calendar_date,
                                day_type,
                                previous_type,
                                holiday_name,
                                previous_holiday_name,
                                remarks,
                                changed_by
                            )
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (
                        calendar_date,
                        day_type,
                        existing['day_type'] if existing else None,
                        holiday_name,
                        existing['holiday_name'] if existing else None,
                        remarks,
                        changed_by
                    ))

                success += 1

            except Exception as row_err:
                failed.append({
                    'row': i,
                    'reason': str(row_err)
                })

        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({
            'message': f'{success} future rows uploaded successfully',
            'skipped_past_dates': skipped,
            'failed': failed
        }), 200

    except Exception as e:
        if conn:
            conn.rollback()

        return jsonify({
            'error': str(e)
        }), 500 

@app.route('/company_calendar/generate-default', methods=['POST'])
def generate_default_calendar():
    data = request.get_json()
    year = data.get('year')

    if not year:
        return jsonify({'error': 'year is required'}), 400

    try:
        from calendar import monthrange
        import datetime

        today = datetime.date.today()
        conn = get_db()
        cursor = conn.cursor()

        # delete all future dates for this year only
        cursor.execute("""
            DELETE FROM company_calendar_settings_all
            WHERE YEAR(calendar_date) = %s
            AND calendar_date >= %s
        """, (year, today))

        deleted = cursor.rowcount

        # regenerate from today onwards
        rows = []
        for month in range(1, 13):
            _, days_in_month = monthrange(year, month)
            saturday_count = 0

            for day in range(1, days_in_month + 1):
                date = datetime.date(year, month, day)

                if date < today:
                    continue  # skip past dates entirely

                weekday = date.weekday()  # 0=Mon, 6=Sun

                if weekday == 6:  # Sunday
                    day_type = 'weekoff'
                elif weekday == 5:  # Saturday
                    saturday_count += 1
                    day_type = 'weekoff' if saturday_count in [2, 4] else 'working'
                else:
                    day_type = 'working'

                rows.append((str(date), day_type))

        # bulk insert fresh
        cursor.executemany("""
            INSERT INTO company_calendar_settings_all
                (calendar_date, day_type, is_default)
            VALUES (%s, %s, TRUE)
        """, rows)

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({
            'message': f'Default calendar generated for {year}',
            'deleted': deleted,
            'inserted': len(rows)
        }), 200

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500




#SUBMISSIONS///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

def _parse_day_col(raw):
    if not raw:
        return None
    try:
        return json.loads(raw) if isinstance(raw, str) else raw
    except (json.JSONDecodeError, TypeError):
        return None

def _build_record(emp_row):
    parsed = _parse_day_col(emp_row.get("day_data"))
    if parsed:
        status = "Delayed" if parsed.get("submit_status", 0) == 1 else "Submitted"
        time   = parsed.get("in_time")
        text   = parsed.get("text")
        wa_id  = parsed.get("wa_id")
    else:
        status = "Missed"
        time = text = wa_id = None
    return {
        "eha_id":      emp_row["eha_id"],
        "first_name":  emp_row["first_name"],
        "middle_name": emp_row.get("middle_name"),
        "last_name":   emp_row["last_name"],
        "status":      status,
        "time":        time,
        "text":        text,
        "wa_id":       wa_id,
    }

@app.route("/team_agenda_submissions",methods=["GET"])
def team_agenda_submissions():
    date_str = request.args.get("date", "")
    tha_id = request.args.get("tha_id", "")
    tab = request.args.get("tab", "all").lower()

    if not tha_id:
        return jsonify({"message": "tha_id is required"}), 400
    
    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"message": "date must be YYYY-MM-DD"}), 400

    if target_date > datetime.today().date():
        return jsonify({"message": "Future dates are not allowed"}), 400
    
    if tab not in ("all", "delayed", "missed","submitted"):
        return jsonify({"message": "tab must be: all | delayed | missed |submitted"}), 400
    
    try:
        page = max(1, int(request.args.get("page", 1)))
        per_page = min(100, max(1, int(request.args.get("per_page", 20))))
    except ValueError:
        return jsonify({"message": "Invalid pagination params"}), 400
    
    mm_yy   = target_date.strftime("%m-%Y")          # e.g. "05-2026"
    day_col = f"day_{target_date.day}"               # e.g. "day_8"
    offset  = (page - 1) * per_page

    conn = get_db()
    cursor = conn.cursor()

    try:
        base_from = """
            FROM team_emp_combination_all tec
            JOIN employee_header_all e
              ON e.eha_id = tec.eha_id
             AND e.status = 1
            LEFT JOIN member_daily_agenda_details_all a
              ON a.tec_id = tec.tec_id
             AND a.tha_id = tec.tha_id
             AND a.mm_yy  = %s
            WHERE tec.tha_id = %s
            AND tec.unlinked_on is NULL
        """
        select_fields = f"""
            SELECT
                e.eha_id,
                e.first_name,
                e.middle_name,
                e.last_name,
                a.{day_col} AS day_data
        """

        if tab == "all":
            where_extra = ""
        elif tab == "submitted":
            where_extra = f"""
                AND a.{day_col} IS NOT NULL
                AND JSON_UNQUOTE(JSON_EXTRACT(a.{day_col}, '$.submit_status')) = '0'
            """
            

        elif tab == "delayed":
            where_extra = f"""
                AND a.{day_col} IS NOT NULL
                AND JSON_UNQUOTE(JSON_EXTRACT(a.{day_col}, '$.submit_status')) = '1'
            """
        else:  # missed
            where_extra = f" AND (a.tec_id IS NULL OR a.{day_col} IS NULL)"
        
        sql = f"""
            {select_fields}
            {base_from}
            {where_extra}
            ORDER BY e.first_name, e.last_name
            LIMIT %s OFFSET %s
        """
        params = (mm_yy, tha_id, per_page, offset)

        cursor.execute(sql, params)
        rows = cursor.fetchall()

        records = []
        for row in rows:
            day_data = row["day_data"]
            # Parse string to dict if necessary
            if isinstance(day_data, str):
                try:
                    day_data = json.loads(day_data)
                except (json.JSONDecodeError, TypeError):
                    day_data = None
            records.append({
                "eha_id": row["eha_id"],
                "first_name": row["first_name"],
                "middle_name": row["middle_name"],
                "last_name": row["last_name"],
                "day_data": day_data   # now an object or None
            })

        return jsonify({"records": records}), 200

    except Exception as e:
        print("Team Agenda VIEW ERROR :",e)
        return jsonify({"message":"Internal Server error"})
    finally:
        cursor.close()
        conn.close()






















@app.route("/agenda_submissions", methods=["GET"])
def agenda_submissions():
    date_str = request.args.get("date", "")
    tab      = request.args.get("tab", "all").lower()

    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = min(100, max(1, int(request.args.get("per_page", 20))))
    except ValueError:
        return jsonify({"message": "Invalid pagination params"}), 400

    if tab not in ("all", "delayed", "missed"):
        return jsonify({"message": "tab must be: all | delayed | missed"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"message": "date must be YYYY-MM-DD"}), 400

    if target_date > datetime.today().date():
        return jsonify({"message": "Future dates are not allowed"}), 400

    mm_yy   = target_date.strftime("%m-%Y")
    day_col = f"day_{target_date.day}"
    offset  = (page - 1) * per_page

    conn   = get_db()
    cursor = conn.cursor()

    try:
        base = f"""
            FROM employee_header_all e
            LEFT JOIN member_daily_agenda_details_all a
                   ON a.eha_id = e.eha_id
                  AND a.mm_yy  = %s
            WHERE e.status = 1
        """
        select = f"""
            SELECT
                e.eha_id,
                e.first_name,
                e.middle_name,
                e.last_name,
                a.{day_col} AS day_data
        """

        if tab == "all":
            sql    = f"{select} {base} ORDER BY e.first_name, e.last_name LIMIT %s OFFSET %s"
            params = (mm_yy, per_page, offset)

        elif tab == "delayed":
            sql    = f"""
                {select} {base}
                  AND a.{day_col} IS NOT NULL
                  AND JSON_UNQUOTE(JSON_EXTRACT(a.{day_col}, '$.submit_status')) = '1'
                ORDER BY e.first_name, e.last_name LIMIT %s OFFSET %s
            """
            params = (mm_yy, per_page, offset)

        else:  # missed
            sql    = f"""
                {select} {base}
                  AND (a.eha_id IS NULL OR a.{day_col} IS NULL)
                ORDER BY e.first_name, e.last_name LIMIT %s OFFSET %s
            """
            params = (mm_yy, per_page, offset)

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        return jsonify({"records": [_build_record(r) for r in rows]}), 200

    except Exception as e:
        print("agenda_submissions error:", e)
        return jsonify({"message": "Internal server error"}), 500

    finally:
        cursor.close()
        conn.close()


def _build_report_record(row):
    parsed = _parse_day_col(row.get("day_data"))
    special = row.get("submission_label") == "Special Approval" 
    submit_next_day ="No"

    if parsed:
        submit_status = parsed.get("submit_status", 0)
        status = "Delayed" if submit_status == 1 else "Submitted"
        time   = parsed.get("submitted_at") or parsed.get("out_time")
        text   = parsed.get("tasks")
        wa_id  = parsed.get("wa_id")
        submit_next_day = parsed.get("submit_next_day","no")
    else:
        status = "Missed"
        time   = None
        text   = None
        wa_id  = None
    print(submit_next_day)

    return {
        "eha_id":            row["eha_id"],
        "first_name":        row["first_name"],
        "middle_name":       row.get("middle_name"),
        "last_name":         row["last_name"],
        "status":            status,
        "is_special":        special,  
        "time":              time,
        "text":              text,
        "wa_id":             wa_id,
        "submit_next_day":  submit_next_day
    }

@app.route("/report_submissions", methods=["GET"])
def daily_report_submissions():
    date_str = request.args.get("date", "")
    tab      = request.args.get("tab", "all").lower()

    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = min(100, max(1, int(request.args.get("per_page", 20))))
    except ValueError:
        return jsonify({"message": "Invalid pagination params"}), 400

    if tab not in ("all", "delayed", "missed"):
        return jsonify({"message": "tab must be: all | delayed | missed"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"message": "date must be YYYY-MM-DD"}), 400

    today = datetime.today().date()

    if target_date > today:
        return jsonify({"message": "Future dates are not allowed"}), 400

    mm_yy    = target_date.strftime("%m-%Y")
    day_col  = f"day_{target_date.day}"
    prev_date   = target_date - timedelta(days=1)
    prev_mm_yy  = prev_date.strftime("%m-%Y")
    prev_col    = f"day_{prev_date.day}"
    is_today    = target_date == today
    offset      = (page - 1) * per_page
    

    conn   = get_db()
    cursor = conn.cursor()

    try:
        if tab == "all":
            sql = f"""
                SELECT
                    e.eha_id, e.first_name, e.middle_name, e.last_name,
                    a.{day_col} AS day_data
                FROM employee_header_all e
                LEFT JOIN member_daily_report_details_all a
                    ON a.eha_id = e.eha_id
                    AND a.mm_yy = %s
                WHERE e.status = 1
                ORDER BY e.first_name, e.last_name
                LIMIT %s OFFSET %s
            """
            params = (mm_yy, per_page, offset)

        elif tab == "delayed":
            sql = f"""
                SELECT e.eha_id, e.first_name, e.middle_name, e.last_name,
                    a.{day_col} AS day_data
                FROM employee_header_all e
                LEFT JOIN member_daily_report_details_all a
                    ON a.eha_id = e.eha_id AND a.mm_yy = %s
                WHERE e.status = 1
                AND a.{day_col} IS NOT NULL
                AND JSON_EXTRACT(a.{day_col}, '$.submit_status') = 1
                ORDER BY e.first_name, e.last_name
                LIMIT %s OFFSET %s
            """
            params = (mm_yy, per_page, offset)

        else:  # missed
            sql = f"""
                SELECT e.eha_id, e.first_name, e.middle_name, e.last_name,
                    NULL AS day_data
                FROM employee_header_all e
                LEFT JOIN member_daily_report_details_all a
                    ON a.eha_id = e.eha_id AND a.mm_yy = %s
                WHERE e.status = 1
                AND (a.eha_id IS NULL OR a.{day_col} IS NULL)
                ORDER BY e.first_name, e.last_name
                LIMIT %s OFFSET %s
            """
            params = (mm_yy, per_page, offset)

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        return jsonify({"records": [_build_report_record(r) for r in rows]}), 200

    except Exception as e:
        print("daily_report_submissions error:", e)
        return jsonify({"message": "Internal server error"}), 500

    finally:
        cursor.close()
        conn.close()

#NEW
@app.route("/teammate_submissions",methods=["GET"])
def teammate_submissions():
    empid = session.get("empid")
    eha_id = request.args.get("eha_id")
    date_str = request.args.get("date", "")
    sub_type = request.args.get("type", "agenda").lower()

    if not eha_id:
        return jsonify({"message": "eha_id required"}), 400

    if sub_type not in ("agenda", "report"):
        return jsonify({"message": "type must be agenda | report"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"message": "date must be YYYY-MM-DD"}), 400

    if target_date > datetime.today().date():
        return jsonify({"message": "Future dates are not allowed"}), 400

    conn   = get_db()
    cursor = conn.cursor()

    try:
        #check if in team
        cursor.execute("""
            SELECT COUNT(*) AS cnt
            FROM team_emp_combination_all c1
            INNER JOIN team_emp_combination_all c2
                    ON c1.tha_id = c2.tha_id
            WHERE c1.eha_id      = %s
              AND c2.eha_id      = %s
              AND c1.unlinked_on IS NULL
              AND c2.unlinked_on IS NULL
        """, (empid, eha_id))
        check = cursor.fetchone()
        if not check or check["cnt"] == 0:
            return jsonify({"message": "Access denied"}), 403
    
        mm_yy   = target_date.strftime("%m-%Y")
        day_col = f"day_{target_date.day}"
        table   = "member_daily_agenda_details_all" if sub_type == "agenda" else "member_daily_report_details_all"

        cursor.execute(f"""
            SELECT {day_col} AS day_data
            FROM {table}
            WHERE eha_id = %s AND mm_yy = %s
        """, (eha_id, mm_yy))
        row = cursor.fetchone()

        day_data = None
        if row and row["day_data"]:
            try:
                day_data = json.loads(row["day_data"]) if isinstance(row["day_data"], str) else row["day_data"]
            except Exception:
                day_data = None

        return jsonify({
            "eha_id":   eha_id,
            "date":     date_str,
            "type":     sub_type,
            "data":     day_data,
            "status":   "Submitted" if day_data and day_data.get("submit_status", 0) == 0
                        else "Delayed" if day_data and day_data.get("submit_status") == 1
                        else "Missed"
        }), 200

    except Exception as e:
        print("teammate_submissions error:", e)
        return jsonify({"message": "Internal server error"}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/daily_agendas_all/today',methods=['GET'])
def daily_agendas__all_today():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Missing date parameter"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn   = get_db()
    cursor = conn.cursor()

    try:

        # check calendar for this date
        cursor.execute("""
            SELECT
                COALESCE(
                    (
                        SELECT h.previous_type
                        FROM company_calendar_history_all h
                        WHERE h.calendar_date = %s
                        ORDER BY h.changed_at ASC
                        LIMIT 1
                    ),
                    c.day_type
                ) AS day_type
            FROM company_calendar_settings_all c
            WHERE c.calendar_date = %s
        """, (target_date, target_date))

        cal_row  = cursor.fetchone()
        day_type = cal_row['day_type'] if cal_row else 'working'

        # fetch employees with date-bound team assignments + agenda + report
        cursor.execute(f"""
    SELECT
        eh.eha_id,
        CONCAT(eh.first_name, ' ', eh.last_name) AS name,

        tec.tec_id,
        tec.tha_id,
        tha.team_name,
        tec.unlinked_on,

        ag.{day_col} AS day_today,
        rp.{day_col} AS report_today,

        (
            SELECT submit_next_day
            FROM hr_general_settings_all
            WHERE (
                valid_till IS NULL
                OR DATE(valid_till) >= %s
            )
            LIMIT 1
        ) AS general_submit_next_day,

        (
            SELECT submit_next_day
            FROM hr_special_approvals_settings_all sa
            WHERE sa.eha_id = eh.eha_id
            AND (
                sa.valid_till IS NULL
                OR DATE(sa.valid_till) >= %s
            )
            LIMIT 1
        ) AS special_submit_next_day

    FROM employee_header_all eh

    LEFT JOIN team_emp_combination_all tec
        ON tec.eha_id = eh.eha_id
        AND DATE(tec.linked_on) <= %s
        AND (
            tec.unlinked_on IS NULL
            OR DATE(tec.unlinked_on) > %s
        )
        AND (
            tec.valid_till IS NULL
            OR DATE(tec.valid_till) >= %s
        )

    LEFT JOIN team_header_all tha
        ON tha.tha_id = tec.tha_id

    LEFT JOIN member_daily_agenda_details_all ag
        ON ag.tec_id = tec.tec_id
        AND ag.mm_yy = %s

    LEFT JOIN member_daily_report_details_all rp
        ON rp.tec_id = tec.tec_id
        AND rp.mm_yy = %s

    WHERE eh.status = 1

    ORDER BY eh.eha_id, tec.tha_id

""", (
    target_date,  # general settings valid_till
    target_date,  # special settings valid_till

    target_date,  # linked_on
    target_date,  # unlinked_on
    target_date,  # tec valid_till

    mm_yy,        # agenda mm_yy
    mm_yy         # report mm_yy
))

        rows = cursor.fetchall()

        def parse(raw):
            if not raw:
                return None
            return json.loads(raw) if isinstance(raw, str) else raw

        employees_dict = {}
        for row in rows:
            eha_id = row['eha_id']

            submit_next_day = (
                "yes"
                if row["general_submit_next_day"] == 1
                or row["special_submit_next_day"] == "yes"
                else "no"
            )


            if eha_id not in employees_dict:
                employees_dict[eha_id] = {
                    "eha_id":    eha_id,
                    "name":      row['name'],
                    "teams":     [],
                    "submit_next_day": submit_next_day
                }

            if row['tec_id']:
                employees_dict[eha_id]["teams"].append({
                    "tec_id":        row['tec_id'],
                    "tha_id":        row['tha_id'],
                    "team_name":     row['team_name'],
                    "is_historical": row['unlinked_on'] is not None,
                    "day_today":     parse(row['day_today']),
                    "report_today":  parse(row['report_today']),
                    "submit_next_day": submit_next_day,

                })

        return jsonify({
            "employees":  list(employees_dict.values()),
            "day_type":   day_type,       # 'working', 'holiday', 'weekoff'
            "date":       str(target_date)
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()




@app.route('/daily_agendas/today', methods=['GET'])
def daily_agendas_today():
    # Only admins / managers can see all employees – adjust session check as needed
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Missing date parameter"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn   = get_db()
    cursor = conn.cursor()
    try:
        # 1. Fetch all active employees with their active team assignments
        cursor.execute(f"""
            SELECT
                eh.eha_id,
                CONCAT(eh.first_name, ' ', eh.last_name) AS name,
                tec.tec_id,
                tec.tha_id,
                tha.team_name,
                da.{day_col} AS day_today
            FROM employee_header_all eh
            LEFT JOIN team_emp_combination_all tec
                ON tec.eha_id = eh.eha_id
                AND tec.unlinked_on IS NULL
                AND (tec.valid_till IS NULL OR tec.valid_till >= CURDATE())
            LEFT JOIN team_header_all tha
                ON tha.tha_id = tec.tha_id
            LEFT JOIN member_daily_agenda_details_all da
                ON da.tec_id = tec.tec_id
                AND da.mm_yy = %s
            WHERE eh.status = 1
            ORDER BY eh.eha_id, tec.tha_id
        """, (mm_yy,))
        rows = cursor.fetchall()

        # 2. Restructure into employees -> teams
        employees_dict = {}
        for row in rows:
            eha_id = row["eha_id"]
            if eha_id not in employees_dict:
                employees_dict[eha_id] = {
                    "eha_id": eha_id,
                    "name": row["name"],
                    "teams": []
                }
            # Only add team if the employee has one (tec_id not null)
            if row["tec_id"]:
                team_data = {
                    "tec_id": row["tec_id"],
                    "tha_id": row["tha_id"],
                    "team_name": row["team_name"],
                    "day_today": None
                }
                # Parse the day column if it exists
                raw_day = row["day_today"]
                if raw_day:
                    try:
                        team_data["day_today"] = json.loads(raw_day) if isinstance(raw_day, str) else raw_day
                    except:
                        team_data["day_today"] = raw_day      # fallback, shouldn't happen
                employees_dict[eha_id]["teams"].append(team_data)

        employees_list = list(employees_dict.values())
        return jsonify({"employees": employees_list}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route("/daily_agendas/monthly", methods=["GET"])
def get_monthly_agendas():
    mm_yy = request.args.get("month")
    print(mm_yy)
    if not mm_yy:
        mm_yy = datetime.today().strftime("%m-%Y")

    day_columns = ", ".join([f"da.day_{d}" for d in range(1, 31 + 1)])

    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(f"""SELECT
            eh.eha_id,
            CONCAT(eh.first_name, ' ', eh.last_name) AS name,
            tec.tec_id,
            tec.tha_id,
            th.team_name,
            {day_columns}
        FROM employee_header_all eh
        LEFT JOIN team_emp_combination_all tec
            ON tec.eha_id = eh.eha_id
            AND tec.unlinked_on IS NULL
            AND (tec.valid_till IS NULL OR tec.valid_till >= CURDATE())
        LEFT JOIN team_header_all th
            ON th.tha_id = tec.tha_id
        LEFT JOIN member_daily_agenda_details_all da
            ON da.tec_id = tec.tec_id
            AND da.mm_yy = %s
        WHERE eh.status = 1
        ORDER BY eh.eha_id, tec.tha_id
    """, (mm_yy,))
        rows = cursor.fetchall()
        cursor.close()
        
    except Exception as e:
        print("get_monthly_agendas error:", str(e))
        return jsonify({"error": str(e)}), 500

    emp_map = {}
    for row in rows:
        eha_id    = row["eha_id"]
        name      = row["name"]
        tec_id    = row["tec_id"]
        tha_id    = row["tha_id"]
        team_name = row["team_name"]

        if eha_id not in emp_map:
            emp_map[eha_id] = {"eha_id": eha_id, "name": name, "teams": []}

        days = {}
        for i in range(1, 32):
            raw = row.get(f"day_{i}")
            if isinstance(raw, str):
                try:
                    days[str(i)] = json.loads(raw)
                except:
                    days[str(i)] = None
            else:
                days[str(i)] = raw

        if tec_id:
            emp_map[eha_id]["teams"].append({
                "tec_id":    tec_id,
                "tha_id":    tha_id,
                "team_name": team_name,
                "days":      days,
            })

    return jsonify({"employees": list(emp_map.values())}), 200

@app.route("/daily_tasks/monthly", methods=["GET"])
def get_monthly_tasks():
    mm_yy = request.args.get("month")
    if not mm_yy:
        mm_yy = datetime.today().strftime("%m-%Y")

    # parse mm_yy to get month and year
    try:
        month, year = mm_yy.split("-")
        month = int(month)
        year  = int(year)
    except:
        return jsonify({"error": "Invalid month format, expected MM-YYYY"}), 400

    days_in_month = cal.monthrange(year, month)[1]
    first_day     = f"{year}-{str(month).zfill(2)}-01"
    last_day      = f"{year}-{str(month).zfill(2)}-{str(days_in_month).zfill(2)}"

    day_columns        = ", ".join([f"da.day_{d}" for d in range(1, days_in_month + 1)])
    report_day_columns = ", ".join([f"rp.day_{d} AS report_day_{d}" for d in range(1, days_in_month + 1)])

    conn   = get_db()
    cursor = conn.cursor()

    try:

        # get calendar for the month with history awareness
        cursor.execute("""
            SELECT
                c.calendar_date,
                COALESCE(
                    (
                        SELECT h.previous_type
                        FROM company_calendar_history_all h
                        WHERE h.calendar_date = c.calendar_date
                        ORDER BY h.changed_at ASC
                        LIMIT 1
                    ),
                    c.day_type
                ) AS day_type
            FROM company_calendar_settings_all c
            WHERE MONTH(c.calendar_date) = %s
              AND YEAR(c.calendar_date)  = %s
        """, (month, year))

        calendar_map = {
            str(r['calendar_date']): r['day_type']
            for r in cursor.fetchall()
        }

        # working days only — for frontend to know max possible submissions
        working_days = [
            d for d in range(1, days_in_month + 1)
            if calendar_map.get(
                f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}",
                'working'
            ) not in ('holiday', 'weekoff')
        ]

        # employees with date-bound team filter + agenda + report
        cursor.execute(f"""
            SELECT
                eh.eha_id,
                CONCAT(eh.first_name, ' ', eh.last_name) AS name,

                tec.tec_id,
                tec.tha_id,
                tec.linked_on,
                tec.unlinked_on,
                tec.valid_till,

                th.team_name,

                {day_columns},
                {report_day_columns},

                (
                    SELECT gs.submit_next_day
                    FROM hr_general_settings_all gs
                    WHERE (
                        gs.valid_till IS NULL
                        OR DATE(gs.valid_till) >= CURDATE()
                    )
                    LIMIT 1
                ) AS general_submit_next_day,

                (
                    SELECT sa.submit_next_day
                    FROM hr_special_approvals_settings_all sa
                    WHERE sa.eha_id = eh.eha_id
                    AND (
                        sa.valid_till IS NULL
                        OR DATE(sa.valid_till) >= CURDATE()
                    )
                    LIMIT 1
                ) AS special_submit_next_day

            FROM employee_header_all eh

            LEFT JOIN team_emp_combination_all tec
                ON tec.eha_id = eh.eha_id
                AND DATE(tec.linked_on) <= %s
                AND (
                    tec.unlinked_on IS NULL
                    OR DATE(tec.unlinked_on) > %s
                )
                AND (
                    tec.valid_till IS NULL
                    OR DATE(tec.valid_till) >= %s
                )

            LEFT JOIN team_header_all th
                ON th.tha_id = tec.tha_id

            LEFT JOIN member_daily_agenda_details_all da
                ON da.tec_id = tec.tec_id
                AND da.mm_yy = %s

            LEFT JOIN member_daily_report_details_all rp
                ON rp.tec_id = tec.tec_id
                AND rp.mm_yy = %s

            WHERE eh.status = 1

            ORDER BY eh.eha_id, tec.tha_id

        """, (
            last_day,
            first_day,
            first_day,
            mm_yy,
            mm_yy
        ))

        rows = cursor.fetchall()

        emp_map = {}
        for row in rows:
            eha_id    = row['eha_id']
            tec_id    = row['tec_id']

            submit_next_day = (
                row["special_submit_next_day"]
                if row["special_submit_next_day"] is not None
                else (
                    "yes"
                    if row["general_submit_next_day"] == 1
                    else "no"
                )
            )

            if eha_id not in emp_map:
                emp_map[eha_id] = {
                    "eha_id": eha_id,
                    "name":   row['name'],
                    "teams":  []
                }

            if not tec_id:
                continue

            linked_on  = row['linked_on']
            unlinked   = row['unlinked_on']
            valid_till = row['valid_till']

            days        = {}
            report_days = {}

            for d in range(1, days_in_month + 1):
                date_str = f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}"
                day_type = calendar_map.get(date_str, 'working')

                # skip holiday/weekoff days
                if day_type in ('holiday', 'weekoff'):
                    days[str(d)]        = None
                    report_days[str(d)] = None
                    continue

                # check if tec was active on this specific day
                day_date = datetime(year, month, d).date()
                lo = linked_on.date() if hasattr(linked_on, 'date') else linked_on
                ul = unlinked.date() if unlinked and hasattr(unlinked, 'date') else unlinked
                vt = valid_till.date() if valid_till and hasattr(valid_till, 'date') else valid_till

                if lo and lo > day_date:
                    continue  # not joined yet
                if ul and ul <= day_date:
                    continue  # already left
                if vt and vt < day_date:
                    continue  # expired

                # agenda
                raw = row.get(f'day_{d}')
                if isinstance(raw, str):
                    try:    days[str(d)] = json.loads(raw)
                    except: days[str(d)] = None
                else:
                    days[str(d)] = raw

                # report
                raw_rp = row.get(f'report_day_{d}')
                if isinstance(raw_rp, str):
                    try:    report_days[str(d)] = json.loads(raw_rp)
                    except: report_days[str(d)] = None
                else:
                    report_days[str(d)] = raw_rp

            has_active_day = False

            for d in range(1, days_in_month + 1):

                day_date = datetime(year, month, d).date()

                lo = linked_on.date() if hasattr(linked_on, 'date') else linked_on
                ul = unlinked.date() if unlinked and hasattr(unlinked, 'date') else unlinked
                vt = valid_till.date() if valid_till and hasattr(valid_till, 'date') else valid_till

                if lo and lo > day_date:
                    continue

                if ul and ul <= day_date:
                    continue

                if vt and vt < day_date:
                    continue

                has_active_day = True
                break

            if not has_active_day:
                continue

            emp_map[eha_id]["teams"].append({
                "tec_id":      tec_id,
                "tha_id":      row['tha_id'],
                "team_name":   row['team_name'],
                "days":        days,
                "report_days": report_days,
                "is_historical": row['unlinked_on'] is not None,
                "submit_next_day": submit_next_day,
                "unlinked_on":   str(row['unlinked_on'].date()) if row['unlinked_on'] else None
            })



        return jsonify({
            "employees":    list(emp_map.values()),
            "working_days": working_days,   # list of day numbers that are working days
            "calendar":     calendar_map,   # full calendar map for frontend use
            "mm_yy":        mm_yy
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/daily_tasks/monthly/download', methods=['GET'])
def download_monthly_tasks():

    mm_yy = request.args.get("month")

    if not mm_yy:
        mm_yy = datetime.today().strftime("%m-%Y")

    try:
        month, year = mm_yy.split("-")
        month = int(month)
        year  = int(year)

    except:
        return jsonify({"error": "Invalid month format"}), 400

    days_in_month = cal.monthrange(year, month)[1]

    first_day = f"{year}-{str(month).zfill(2)}-01"
    last_day  = f"{year}-{str(month).zfill(2)}-{str(days_in_month).zfill(2)}"

    day_columns = ", ".join([
        f"da.day_{d}" for d in range(1, days_in_month + 1)
    ])

    report_day_columns = ", ".join([
        f"rp.day_{d} AS report_day_{d}"
        for d in range(1, days_in_month + 1)
    ])

    conn   = get_db()
    cursor = conn.cursor()

    try:

        # ---------------------------------------------------
        # Calendar Settings
        # ---------------------------------------------------

        cursor.execute("""
            SELECT
                c.calendar_date,

                COALESCE(
                    (
                        SELECT h.previous_type
                        FROM company_calendar_history_all h
                        WHERE h.calendar_date = c.calendar_date
                        ORDER BY h.changed_at ASC
                        LIMIT 1
                    ),
                    c.day_type
                ) AS day_type

            FROM company_calendar_settings_all c

            WHERE MONTH(c.calendar_date) = %s
            AND YEAR(c.calendar_date) = %s
        """, (month, year))

        calendar_map = {
            str(r['calendar_date']): r['day_type']
            for r in cursor.fetchall()
        }
        today = date.today()
        working_days = [
            d for d in range(1, days_in_month + 1)

            if (
                calendar_map.get(
                    f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}",
                    'working'
                ) not in ('holiday', 'weekoff')
            )

            and date(year, month, d) <= today
        ]

        # ---------------------------------------------------
        # Employee + Team + Agenda + Reports
        # ---------------------------------------------------

        cursor.execute(f"""
            SELECT

                eh.eha_id,
                CONCAT(eh.first_name, ' ', eh.last_name) AS name,

                tec.tec_id,
                tec.tha_id,
                tec.linked_on,
                tec.unlinked_on,
                tec.valid_till,

                th.team_name,

                {day_columns},
                {report_day_columns},

                (
                    SELECT gs.submit_next_day
                    FROM hr_general_settings_all gs
                    WHERE (
                        gs.valid_till IS NULL
                        OR DATE(gs.valid_till) >= CURDATE()
                    )
                    LIMIT 1
                ) AS general_submit_next_day,

                (
                    SELECT sa.submit_next_day
                    FROM hr_special_approvals_settings_all sa
                    WHERE sa.eha_id = eh.eha_id
                    AND (
                        sa.valid_till IS NULL
                        OR DATE(sa.valid_till) >= CURDATE()
                    )
                    LIMIT 1
                ) AS special_submit_next_day

            FROM employee_header_all eh

            LEFT JOIN team_emp_combination_all tec
                ON tec.eha_id = eh.eha_id
                AND DATE(tec.linked_on) <= %s
                AND (
                    tec.unlinked_on IS NULL
                    OR DATE(tec.unlinked_on) > %s
                )
                AND (
                    tec.valid_till IS NULL
                    OR DATE(tec.valid_till) >= %s
                )

            LEFT JOIN team_header_all th
                ON th.tha_id = tec.tha_id

            LEFT JOIN member_daily_agenda_details_all da
                ON da.tec_id = tec.tec_id
                AND da.mm_yy = %s

            LEFT JOIN member_daily_report_details_all rp
                ON rp.tec_id = tec.tec_id
                AND rp.mm_yy = %s

            WHERE eh.status = 1

            ORDER BY eh.eha_id, tec.tha_id

        """, (
            last_day,
            first_day,
            first_day,
            mm_yy,
            mm_yy
        ))

        rows = cursor.fetchall()

        emp_map = {}

        for row in rows:

            eha_id = row['eha_id']
            tec_id = row['tec_id']

            submit_next_day = (
                row["special_submit_next_day"]
                if row["special_submit_next_day"] is not None
                else ("yes" if row["general_submit_next_day"] == 1 else "no")
            )

            if eha_id not in emp_map:
                emp_map[eha_id] = {
                    "eha_id": eha_id,
                    "name": row['name'],
                    "teams": []
                }

            if not tec_id:
                continue

            linked_on  = row['linked_on']
            unlinked   = row['unlinked_on']
            valid_till = row['valid_till']

            days        = {}
            report_days = {}

            for d in range(1, days_in_month + 1):

                date_str = f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}"

                day_type = calendar_map.get(date_str, 'working')

                if day_type in ('holiday', 'weekoff'):
                    days[str(d)]        = None
                    report_days[str(d)] = None
                    continue

                day_date = datetime(year, month, d).date()

                lo = linked_on.date() if hasattr(linked_on, 'date') else linked_on
                ul = unlinked.date()  if unlinked and hasattr(unlinked, 'date') else unlinked
                vt = valid_till.date() if valid_till and hasattr(valid_till, 'date') else valid_till

                if lo and lo > day_date:
                    continue

                if ul and ul <= day_date:
                    continue

                if vt and vt < day_date:
                    continue

                raw = row.get(f'day_{d}')

                if isinstance(raw, str):
                    try:
                        days[str(d)] = json.loads(raw)
                    except:
                        days[str(d)] = None
                else:
                    days[str(d)] = raw

                raw_rp = row.get(f'report_day_{d}')

                if isinstance(raw_rp, str):
                    try:
                        report_days[str(d)] = json.loads(raw_rp)
                    except:
                        report_days[str(d)] = None
                else:
                    report_days[str(d)] = raw_rp

            # -------------------------------------------
            # Active Team Check
            # -------------------------------------------

            has_active_day = False

            for d in range(1, days_in_month + 1):

                day_date = datetime(year, month, d).date()

                lo = linked_on.date() if hasattr(linked_on, 'date') else linked_on
                ul = unlinked.date() if unlinked and hasattr(unlinked, 'date') else unlinked
                vt = valid_till.date() if valid_till and hasattr(valid_till, 'date') else valid_till

                if lo and lo > day_date:
                    continue

                if ul and ul <= day_date:
                    continue

                if vt and vt < day_date:
                    continue

                has_active_day = True
                break

            if not has_active_day:
                continue

            emp_map[eha_id]["teams"].append({
                "tec_id": tec_id,
                "tha_id": row['tha_id'],
                "team_name": row['team_name'],
                "days": days,
                "report_days": report_days,
                "submit_next_day": submit_next_day,
                "linked_on": linked_on,
                "unlinked_on": unlinked,
                "valid_till": valid_till
            })

        # ---------------------------------------------------
        # Team Status Logic
        # ---------------------------------------------------

        def monthly_team_status(team, day):

            agenda = team['days'].get(str(day))
            report = team['report_days'].get(str(day))

            def get_status(d):

                if not d:
                    return "mis"

                if d.get('submit_status') == 0:
                    return "sub"

                if d.get('submit_status') == 1:
                    return "del"

                return "mis"

            agenda_st = get_status(agenda)

            if not report:

                today_date = date.today()
                cell_date  = date(year, month, day)

                diff_days = (today_date - cell_date).days

                if team['submit_next_day'] == "yes" and diff_days in (0, 1):
                    report_st = "pending"
                else:
                    report_st = "mis"

            else:
                report_st = get_status(report)

            if agenda_st == "mis" or report_st == "mis":
                return "mis"

            if agenda_st == "del" or report_st == "del":
                return "del"

            if report_st == "pending":
                return "pending"

            if agenda_st == "sub" or report_st == "sub":
                return "sub"

            return "mis"

        # ---------------------------------------------------
        # Employee Day Status
        # ---------------------------------------------------

        def get_monthly_status(emp, day):

            date_key = f"{year}-{str(month).zfill(2)}-{str(day).zfill(2)}"

            if calendar_map.get(date_key, 'working') in ('holiday', 'weekoff'):
                return None

            statuses = []

            for t in emp['teams']:

                lo = t['linked_on']
                ul = t['unlinked_on']
                vt = t['valid_till']

                day_date = date(year, month, day)

                lo = lo.date() if lo and hasattr(lo, 'date') else lo
                ul = ul.date() if ul and hasattr(ul, 'date') else ul
                vt = vt.date() if vt and hasattr(vt, 'date') else vt

                if lo and lo > day_date:
                    continue

                if ul and ul <= day_date:
                    continue

                if vt and vt < day_date:
                    continue

                statuses.append(monthly_team_status(t, day))

            if not statuses:
                return None

            if any(s == "del" for s in statuses):
                return "del"

            if any(s == "pending" for s in statuses):
                return "pending"

            if any(s == "sub" for s in statuses):
                return "sub"

            return "mis"

        # ---------------------------------------------------
        # Build XLSX
        # ---------------------------------------------------

        wb = Workbook()
        ws = wb.active

        ws.title = "Monthly Summary"

        ws.append([
            "Employee ID",
            "Employee Name",
            "Weekoffs",
            "Holidays",
            "Working Days",
            "Attended",
            "Delayed",
            "Missed"
        ])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for eha_id, emp in emp_map.items():

            attended = 0
            delayed  = 0
            missed   = 0

            holidays = 0
            weekoffs = 0

            # ---------------------------------------
            # Count Holidays & Weekoffs
            # ---------------------------------------

            for d in range(1, days_in_month + 1):

                date_key = f"{year}-{str(month).zfill(2)}-{str(d).zfill(2)}"

                day_type = calendar_map.get(date_key, 'working')

                if day_type == "holiday":
                    holidays += 1

                elif day_type == "weekoff":
                    weekoffs += 1

            # ---------------------------------------
            # Count Statuses
            # ---------------------------------------

            for day in working_days:

                st = get_monthly_status(emp, day)

                if st is None:
                    continue

                if st == "sub":
                    attended += 1

                elif st == "del":
                    delayed += 1

                elif st == "mis":
                    missed += 1

            ws.append([
                eha_id,
                emp["name"],
                weekoffs,
                holidays,
                len(working_days),
                attended,
                delayed,
                missed
            ])

        # ---------------------------------------------------
        # Auto Width
        # ---------------------------------------------------

        for col in ws.columns:

            max_length = 0

            col_letter = col[0].column_letter

            for cell in col:
                try:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(
                max_length + 4,
                30
            )

        # ---------------------------------------------------
        # Download
        # ---------------------------------------------------

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"monthly_summary_{mm_yy}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        import traceback
        traceback.print_exc()

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        cursor.close()
        conn.close()

@app.route("/daily_report/monthly", methods=["GET"])
def get_monthly_reports():
    mm_yy = request.args.get("month")
    print(mm_yy)
    if not mm_yy:
        mm_yy = datetime.today().strftime("%m-%Y")

    day_columns = ", ".join([f"da.day_{d}" for d in range(1, 31 + 1)])

    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(f"""SELECT
            eh.eha_id,
            CONCAT(eh.first_name, ' ', eh.last_name) AS name,
            tec.tec_id,
            tec.tha_id,
            th.team_name,
            {day_columns}
        FROM employee_header_all eh
        LEFT JOIN team_emp_combination_all tec
            ON tec.eha_id = eh.eha_id
            AND tec.unlinked_on IS NULL
            AND (tec.valid_till IS NULL OR tec.valid_till >= CURDATE())
        LEFT JOIN team_header_all th
            ON th.tha_id = tec.tha_id
        LEFT JOIN member_daily_report_details_all da
            ON da.tec_id = tec.tec_id
            AND da.mm_yy = %s
        WHERE eh.status = 1
        ORDER BY eh.eha_id, tec.tha_id
    """, (mm_yy,))
        rows = cursor.fetchall()
        cursor.close()
        
    except Exception as e:
        print("get_monthly_report error:", str(e))
        return jsonify({"error": str(e)}), 500

    emp_map = {}
    for row in rows:
        eha_id    = row["eha_id"]
        name      = row["name"]
        tec_id    = row["tec_id"]
        tha_id    = row["tha_id"]
        team_name = row["team_name"]

        if eha_id not in emp_map:
            emp_map[eha_id] = {"eha_id": eha_id, "name": name, "teams": []}

        days = {}
        for i in range(1, 32):
            raw = row.get(f"day_{i}")
            if isinstance(raw, str):
                try:
                    days[str(i)] = json.loads(raw)
                except:
                    days[str(i)] = None
            else:
                days[str(i)] = raw

        if tec_id:
            emp_map[eha_id]["teams"].append({
                "tec_id":    tec_id,
                "tha_id":    tha_id,
                "team_name": team_name,
                "days":      days,
            })

    return jsonify({"employees": list(emp_map.values())}), 200

@app.route('/daily_report/today', methods=['GET'])
def daily_report_today():
    # Only admins / managers can see all employees – adjust session check as needed
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Missing date parameter"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")
    

    conn   = get_db()
    cursor = conn.cursor()
    try:
        # 1. Fetch all active employees with their active team assignments
        cursor.execute(f"""
            SELECT
                eh.eha_id,
                CONCAT(eh.first_name, ' ', eh.last_name) AS name,
                tec.tec_id,
                tec.tha_id,
                tha.team_name,
                da.{day_col} AS day_today
            FROM employee_header_all eh
            LEFT JOIN team_emp_combination_all tec
                ON tec.eha_id = eh.eha_id
                AND tec.unlinked_on IS NULL
                AND (tec.valid_till IS NULL OR tec.valid_till >= CURDATE())
            LEFT JOIN team_header_all tha
                ON tha.tha_id = tec.tha_id
            LEFT JOIN member_daily_report_details_all da
                ON da.tec_id = tec.tec_id
                AND da.mm_yy = %s
            WHERE eh.status = 1
            ORDER BY eh.eha_id, tec.tha_id
        """, (mm_yy,))
        rows = cursor.fetchall()
        print(rows)
        

        # 2. Restructure into employees -> teams
        employees_dict = {}
        for row in rows:
            eha_id = row["eha_id"]
            if eha_id not in employees_dict:
                employees_dict[eha_id] = {
                    "eha_id": eha_id,
                    "name": row["name"],
                    "teams": []
                }
            # Only add team if the employee has one (tec_id not null)
            if row["tec_id"]:
                team_data = {
                    "tec_id": row["tec_id"],
                    "tha_id": row["tha_id"],
                    "team_name": row["team_name"],
                    "day_today": None
                }
                # Parse the day column if it exists
                raw_day = row["day_today"]
                if raw_day:
                    try:
                        team_data["day_today"] = json.loads(raw_day) if isinstance(raw_day, str) else raw_day
                    except:
                        team_data["day_today"] = raw_day      # fallback, shouldn't happen
                employees_dict[eha_id]["teams"].append(team_data)
        

        employees_list = list(employees_dict.values())
        return jsonify({"employees": employees_list}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()























#TEAM VIEW//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

@app.route("/my_teams", methods=["GET"])
def my_teams():
    empid = session.get("empid")
    date_str = request.args.get("date")
    if date_str:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        target_date = date.today()

    conn   = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT
                t.tha_id,
                t.team_name,
                t.status,
                MAX(c.link_type) AS my_link_type,
                MAX(c.unlinked_on) AS unlinked_on

            FROM team_emp_combination_all c

            INNER JOIN team_header_all t
                ON t.tha_id = c.tha_id

            WHERE c.eha_id = %s

            -- employee existed in team
            AND DATE(c.linked_on) <= %s

            AND (
                c.unlinked_on IS NULL
                OR DATE(c.unlinked_on) > %s
            )

            -- team existed on selected date
            AND DATE(t.inserted_on) <= %s

            AND (
                t.valid_till IS NULL
                OR DATE(t.valid_till) >= %s
            )

            GROUP BY t.tha_id, t.team_name, t.status

        """, (
            empid,
            target_date,
            target_date,
            target_date,
            target_date
        ))
        my_teams = cursor.fetchall()

        result = []
        for team in my_teams:
            tha_id = team["tha_id"]

            # Create a fresh cursor for each inner query
            cursor2 = conn.cursor()
            cursor2.execute("""
                SELECT
                    c.tec_id,
                    c.eha_id,
                    c.link_type,
                    c.designation,

                    e.first_name,
                    e.middle_name,
                    e.last_name,
                    e.expertise,

                    c.linked_on,
                    c.unlinked_on

                FROM team_emp_combination_all c

                INNER JOIN employee_header_all e
                    ON e.eha_id = c.eha_id

                WHERE c.tha_id = %s

                -- employee had already joined
                AND DATE(c.linked_on) <= %s

                -- employee should still belong to team
                AND (
                    c.unlinked_on IS NULL
                    OR DATE(c.unlinked_on) > %s
                )

                GROUP BY c.eha_id

                ORDER BY e.first_name

            """, (
                tha_id,
                target_date,
                target_date
            ))
            members = cursor2.fetchall()
            cursor2.close()

            result.append({
                "tha_id":       team["tha_id"],
                "team_name":    team["team_name"],
                "my_link_type": team["my_link_type"],
                "is_historical": team["unlinked_on"] is not None,
                "status": team["status"],
                "members": [
                    {
                        "tec_id":      m["tec_id"],
                        "eha_id":      m["eha_id"],
                        "first_name":  m["first_name"],
                        "middle_name": m["middle_name"],
                        "last_name":   m["last_name"],
                        "expertise":   m["expertise"],
                        "role":        "Leader" if m["link_type"] == 2 else "Member"
                    }
                    for m in members
                ]
            })

        return jsonify({"teams": result}), 200

    except Exception as e:
        print("my_teams error:", e)
        return jsonify({"message": "Internal server error"}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/daily_team_submissions/today', methods=['GET'])
def daily_team_submissions():
    empid = session.get("empid")
    if not empid:
        return jsonify({"error": "Not logged in"}), 401

    
    date_str = request.args.get("date")
    tha_id = request.args.get("tha_id")
    if not date_str:
        return jsonify({"error": "Missing date parameter"}), 400

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        return jsonify({"error": "Invalid date format"}), 400

    day_col = f"day_{target_date.day}"
    mm_yy   = target_date.strftime("%m-%Y")

    conn   = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute(f"""
            SELECT
                eh.eha_id,
                CONCAT(eh.first_name, ' ', eh.last_name) AS name,

                tec.tec_id,
                tec.tha_id,

                tha.team_name,
                tec.unlinked_on,

                ag.{day_col} AS agenda_today,
                rp.{day_col} AS report_today

            FROM employee_header_all eh

            INNER JOIN team_emp_combination_all tec
                ON tec.eha_id = eh.eha_id

                AND DATE(tec.linked_on) <= %s

                AND (
                    tec.unlinked_on IS NULL
                    OR DATE(tec.unlinked_on) > %s
                )

                AND (
                    tec.valid_till IS NULL
                    OR DATE(tec.valid_till) >= %s
                )

            INNER JOIN team_header_all tha
                ON tha.tha_id = tec.tha_id

            LEFT JOIN member_daily_agenda_details_all ag
                ON ag.tec_id = tec.tec_id
                AND ag.mm_yy = %s

            LEFT JOIN member_daily_report_details_all rp
                ON rp.tec_id = tec.tec_id
                AND rp.mm_yy = %s

            WHERE eh.status = 1
            AND tec.tha_id = %s

            ORDER BY eh.eha_id

        """, (
            target_date,
            target_date,
            target_date,
            mm_yy,
            mm_yy,
            tha_id
        ))

        rows = cursor.fetchall()

        def parse(raw):
            if not raw:
                return None
            return json.loads(raw) if isinstance(raw, str) else raw

        employees_dict = {}
        for row in rows:
            eha_id = row["eha_id"]

            if eha_id not in employees_dict:
                employees_dict[eha_id] = {
                    "eha_id": eha_id,
                    "name":   row["name"],
                    "teams":  []
                }

            if row["tec_id"]:
                employees_dict[eha_id]["teams"].append({
                    "tec_id":       row["tec_id"],
                    "tha_id":       row["tha_id"],
                    "team_name":    row["team_name"],
                    "agenda_today": parse(row["agenda_today"]),
                    "report_today": parse(row["report_today"]),
                    "is_historical": row["unlinked_on"] is not None,
                })
        print("heloo")
        print(list(employees_dict.values()))

        return jsonify({"employees": list(employees_dict.values())}), 200

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()








#debug///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////


@app.route("/users")
def get_users():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM user") #Match from the database
    users = cursor.fetchone()                                                                                                #this is a dict(object) now for
    conn.close()
    result = []

    for u in users:
        result.append({
            "empid": u["empid"],
            "name": u.name,
            "email": u.email,
            "role": u.role,
            "password": u.password,
            "age": u.age,
            "gender": u.gender,
            "DOB":u.DOB,
            "marital_status":u.marital_status,
            "present_loc": u.present_loc,
            "permanent_loc": u.permanent_loc,
            "aadhar": u.aadhar,
            "pan": u.pan,
            "mobile_no": u.mobile_no,
            "bank_acc": u.bank_acc,
            "joining_date": str(u.joining_date) if u.joining_date else None,
            "tenth_school": u.tenth_school,
            "tenth_board": u.tenth_board,
            "tenth_year": u.tenth_year,
            "tenth_marks": u.tenth_marks,
            "twelfth_school": u.twelfth_school,
            "twelfth_board": u.twelfth_board,
            "twelfth_year": u.twelfth_year,
            "twelfth_marks": u.twelfth_marks,
            "ug_college": u.ug_college,
            "ug_degree": u.ug_degree,
            "ug_year": u.ug_year,
            "ug_marks": u.ug_marks,
            "pg_college": u.pg_college,
            "pg_degree": u.pg_degree,
            "pg_year": u.pg_year,
            "pg_marks": u.pg_marks,
            "father_name": u.father_name,
            "father_no": u.father_no,
            "father_location": u.father_location,
            "mother_name": u.mother_name,
            "mother_no": u.mother_no,
            "mother_location": u.mother_location,
            "aadhar_image": u.aadhar_image,
            "pan_image": u.pan_image,
            "tenth_marksheet": u.tenth_marksheet,
            "twelfth_marksheet": u.twelfth_marksheet,
            "ug_degree_image": u.ug_degree_image,
            "pg_degree_image": u.pg_degree_image,
            "photo": u.photo
        })

    return jsonify(result)




    
@app.route("/uploads/documents/<filename>")
def serve_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],filename)          #Send from dir(FROM WHERE,WHAT)


@app.route("/dailywork")
def daily_work():
    tasks = Tasks.query.all()
    result = []

    for t in tasks:
        result.append({
            "taskid": t.taskid,
            "content": t.content,
            "status": t.status,
            "empid" : t.empid
        })

    return jsonify(result)











    


if __name__ =="__main__":
    app.run(debug=True)

